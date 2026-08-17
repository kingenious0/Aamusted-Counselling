import os, sys
import threading
import time
import re

# ── Field-level Local-DB Encryption (GTEC Compliance) ──────────────────────
try:
    from crypto_utils import (
        encrypt_field, decrypt_field,
        STUDENT_SENSITIVE_FIELDS, BOOKING_SENSITIVE_FIELDS,
        CASENOTE_SENSITIVE_FIELDS, REFERRAL_SENSITIVE_FIELDS,
        SESSION_SENSITIVE_FIELDS
    )
    _CRYPTO_AVAILABLE = True
except ImportError:
    _CRYPTO_AVAILABLE = False
    def encrypt_field(v): return v
    def decrypt_field(v): return v
    print("[STARTUP] crypto_utils not loaded — fields stored as plaintext")

# Ensure core directory is in the path for imports
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
else:
    base_dir = os.path.dirname(os.path.abspath(__file__))

core_dir = os.path.join(base_dir, 'core')
if os.path.exists(core_dir) and core_dir not in sys.path:
    sys.path.insert(0, core_dir)
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Response, send_file, make_response
from functools import wraps
import sqlite3
import csv
import io
import json
from datetime import datetime, timedelta
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
try:
    import auto_report_writer
    from auto_report_writer import scheduler, toggle_scheduler, manual_generate_report
    REPORTER_ENABLED = True
except ImportError:
    REPORTER_ENABLED = False
    scheduler = None
    def toggle_scheduler(enable): pass
    def manual_generate_report(): print("Report generator disabled.")

import uuid
import node_config 
import sync_engine
from sync_engine import sync_bp, sync_manager, trigger_sync_immediate


# Initialize Node Config on Startup
current_node_config = node_config.load_config()
print(
    f"--- Node Identity: {current_node_config['node_id']} ({current_node_config['node_role']}) ---")

app = Flask(__name__)
# Register Sync Blueprint
app.register_blueprint(sync_bp)

# Start Automated Sync Background Thread
sync_manager.start()


# ==========================================
# DATE/TIME UTILS
# ==========================================
def clean_date_string(date_val):
    """
    Cleans a date value that might be an ISO string, a partial string, 
    or a full timestamp like 'Thu, 19 Mar 2026 00:00:00 GMT'.
    Returns a standard YYYY-MM-DD string or original if unparseable.
    """
    if not date_val:
        return datetime.now().strftime('%Y-%m-%d')
    
    raw_str = str(date_val).strip()
    
    # 1. Try ISO pattern YYYY-MM-DD (e.g., 2026-03-19)
    import re
    iso_match = re.search(r'(\d{4}-\d{2}-\d{2})', raw_str)
    if iso_match:
        return iso_match.group(1)
        
    # 2. Try parsing "Thu, 19 Mar 2026..." (Supabase/Bridge format)
    try:
        clean_str = raw_str.replace(' GMT', '').replace(',', '')
        if ' ' in clean_str:
            parts = clean_str.split(' ')
            # Example: ['Thu', '19', 'Mar', '2026', ...]
            if len(parts) >= 4:
                day_str = f"{parts[1]} {parts[2]} {parts[3]}"
                d_obj = datetime.strptime(day_str, '%d %b %Y')
                return d_obj.strftime('%Y-%m-%d')
    except Exception:
        pass
        
    return raw_str

def clean_time_string(time_val):
    """Cleans time to standard display format (e.g., 2:00 PM)."""
    if not time_val:
        return "09:00 AM"
    
    raw_str = str(time_val).strip().upper()
    try:
        if 'PM' in raw_str or 'AM' in raw_str:
            # Handle "2:00PM" or "2:00 P.M."
            t_str = raw_str.replace('.', '')
            if ' ' not in t_str:
                t_str = t_str.replace('AM', ' AM').replace('PM', ' PM')
            t_obj = datetime.strptime(t_str, '%I:%M %p')
            return t_obj.strftime('%I:%M %p')
        else:
            # Handle 24h: "14:22" or "14:22:00"
            parts = raw_str.split(':')
            if len(parts) >= 2:
                t_obj = datetime.strptime(f"{parts[0]}:{parts[1]}", '%H:%M')
                return t_obj.strftime('%I:%M %p')
    except Exception:
        pass
    
    return raw_str

app.secret_key = 'super_secret_key_for_dev_only'  # Change for production
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

# Global flag to prevent concurrent initialization
_db_initialization_lock = False
_db_initialized = False


def ensure_database_initialized():
    """Ensure database exists and has all required tables - atomic, single-call initialization"""
    global _db_initialization_lock, _db_initialized

    # If already initialized successfully, skip
    if _db_initialized:
        return

    # If currently initializing, wait (prevent concurrent calls)
    if _db_initialization_lock:
        import time
        wait_count = 0
        while _db_initialization_lock and wait_count < 50:  # Wait up to 5 seconds
            time.sleep(0.1)
            wait_count += 1
        if _db_initialized:
            return

    # Set lock to prevent concurrent initialization
    _db_initialization_lock = True

    try:
        # Determine database path (Vercel-safe: /tmp on serverless, local otherwise)
        try:
            import db_setup as _db_setup_mod
            db_path = _db_setup_mod.get_db_path()
        except Exception:
            # Fallback: same env-aware logic
            import tempfile
            if os.environ.get('VERCEL') or os.environ.get('AWS_LAMBDA_FUNCTION_NAME'):
                db_path = os.path.join(tempfile.gettempdir(), 'counseling.db')
            elif getattr(sys, 'frozen', False):
                db_path = os.path.join(os.path.dirname(sys.executable), 'counseling.db')
            else:
                db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'counseling.db')

        # Required tables for the application
        required_tables = [
            'Appointment', 'Student', 'Counsellor', 'session', 'app_settings', 
            'Referral', 'reports', 'BookingRequest', 'users', 'audit_logs', 
            'Notification', 'SMSQueue'
        ]

        # Check if database exists and has all required tables
        needs_init = False

        if not os.path.exists(db_path):
            print(f"[STARTUP] Database not found at: {db_path}")
            needs_init = True
        else:
            # Check for all required tables in a single connection
            try:
                check_conn = sqlite3.connect(db_path, timeout=5.0)
                cursor = check_conn.cursor()
                cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'")
                existing_tables = {row[0].lower() for row in cursor.fetchall()}
                check_conn.close()

                # Check if all required tables exist (case-insensitive)
                missing_tables = []
                for req_table in required_tables:
                    if req_table.lower() not in existing_tables:
                        missing_tables.append(req_table)

                if missing_tables:
                    print(f"[STARTUP] Missing tables: {missing_tables}")
                    needs_init = True
                else:
                    print(
                        f"[STARTUP] Database check passed - all required tables exist")
                    _db_initialized = True
                    _db_initialization_lock = False
                    return
            except Exception as e:
                print(f"[STARTUP] Error checking database: {e}")
                needs_init = True

        # Initialize database if needed (only once, atomically)
        if needs_init:
            print(f"[STARTUP] Initializing database at: {db_path}")
            try:
                # Import and run initialization (uses same path logic)
                import db_setup
                db_setup.init_db()

                # Verify initialization succeeded - check for Appointment table
                verify_conn = sqlite3.connect(db_path, timeout=5.0)
                verify_cursor = verify_conn.cursor()
                verify_cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'")
                verify_tables = {row[0].lower()
                                 for row in verify_cursor.fetchall()}
                verify_conn.close()

                # Check if all required tables now exist (case-insensitive check)
                missing_after_init = []
                for req_table in required_tables:
                    if req_table.lower() not in verify_tables:
                        missing_after_init.append(req_table)

                if missing_after_init:
                    print(
                        f"[STARTUP] ERROR: Initialization incomplete! Missing tables: {missing_after_init}")
                    print(f"[STARTUP] Available tables: {verify_tables}")
                    raise Exception(
                        f"Database initialization incomplete. Missing tables: {missing_after_init}")
                else:
                    print(
                        "[STARTUP] Database initialized successfully - all required tables verified")
                    _db_initialized = True

            except Exception as e:
                print(f"[STARTUP] ERROR initializing database: {e}")
                import traceback
                traceback.print_exc()
                # Don't set _db_initialized = True so it can retry
                raise

    finally:
        _db_initialization_lock = False


# Initialize database when module is loaded
ensure_database_initialized()

# ── One-time migration: add accepted_at to BookingRequest if not present ──
try:
    import db_setup as _ds
    _migration_conn = sqlite3.connect(_ds.get_db_path(), timeout=5.0)
    _migration_conn.execute("ALTER TABLE BookingRequest ADD COLUMN accepted_at TEXT")
    _migration_conn.commit()
    _migration_conn.close()
    print("[MIGRATION] Added accepted_at column to BookingRequest")
except Exception:
    pass  # Column already exists or table not yet created — ignore


# ══════════════════════════════════════════════════════════════════════════════
# SILENT AUTO-UPDATER — Checks GitHub 30s after startup & applies updates
# Protected files are NEVER touched: counseling.db, node_config.json
# ══════════════════════════════════════════════════════════════════════════════
def _run_silent_auto_update():
    """Background daemon: silently pulls latest code from GitHub on startup."""
    import time as _time
    import zipfile as _zipfile
    import io as _io
    import shutil as _shutil

    _REPO_OWNER   = "kingenious0"
    _REPO_NAME    = "Aamusted-Counselling"
    _BRANCH       = "main"
    _API_URL      = f"https://api.github.com/repos/{_REPO_OWNER}/{_REPO_NAME}/commits/{_BRANCH}"
    _ZIP_URL      = f"https://github.com/{_REPO_OWNER}/{_REPO_NAME}/archive/refs/heads/{_BRANCH}.zip"
    _SHA_FILE     = os.path.join(base_dir, 'current_sha.txt')
    _TEMP_DIR     = os.path.join(base_dir, '_update_temp')

    # Files/folders that must NEVER be overwritten or even touched
    _PROTECTED    = {
        'counseling.db', 'node_config.json', 'current_sha.txt',
        '.git', '_update_temp', 'PATCH.py',
        # Distribution/packaging folders — contain .bat files that trigger
        # Windows Defender WinError 225 and must be skipped entirely
        'AAMUSTED_Universal_Distribution_Old_v2',
        'USTED_Counseling_System_Distribution',
        'AAMUSTED_Universal_Distribution',
        'deployment',
    }
    _HEADERS      = {'User-Agent': 'AAMUSTED-AutoUpdater/2.0'}

    while True:
        # Wait for system to fully start before doing anything (or wait between cycles)
        _time.sleep(30)
        
        try:
            import requests as _req

            # 1. Read local SHA
            _local_sha = ""
            if os.path.exists(_SHA_FILE):
                with open(_SHA_FILE, 'r') as _f:
                    _local_sha = _f.read().strip()

            # 2. Get latest remote SHA
            _api_resp = _req.get(_API_URL, headers=_HEADERS, timeout=15)
            if _api_resp.status_code != 200:
                # Silently fail, likely no internet or GitHub down
                pass
            else:
                _remote_sha = _api_resp.json().get('sha', '')
                _known_outdated_github_sha = '77a1dfd657b3da5c29a3ffc09055da13222cd34c'
                if _remote_sha and _remote_sha != _local_sha and _remote_sha != _known_outdated_github_sha:
                    print(f"[AUTO-UPDATE] New version found: {_local_sha[:8]} -> {_remote_sha[:8]}. Downloading...")

                    # 3. Download ZIP
                    _zip_resp = _req.get(_ZIP_URL, headers=_HEADERS, timeout=180)
                    if _zip_resp.status_code == 200:
                        # 4. Extract
                        if os.path.exists(_TEMP_DIR):
                            _shutil.rmtree(_TEMP_DIR)
                        os.makedirs(_TEMP_DIR)

                        _z = _zipfile.ZipFile(_io.BytesIO(_zip_resp.content))
                        _z.extractall(_TEMP_DIR)

                        _extracted = os.listdir(_TEMP_DIR)
                        if _extracted:
                            _source = os.path.join(_TEMP_DIR, _extracted[0])

                            # 5. Copy files one-by-one — skip protected
                            _copied = 0
                            for _item in os.listdir(_source):
                                if _item in _PROTECTED:
                                    continue
                                _src = os.path.join(_source, _item)
                                _dst = os.path.join(base_dir, _item)
                                try:
                                    if os.path.isdir(_src):
                                        if os.path.exists(_dst):
                                            _shutil.rmtree(_dst)
                                        _shutil.copytree(_src, _dst, ignore=_shutil.ignore_patterns('*.bat', '*.exe', '*.msi'))
                                    else:
                                        _shutil.copy2(_src, _dst)
                                    _copied += 1
                                except:
                                    pass

                            # 6. Save new SHA
                            with open(_SHA_FILE, 'w') as _f:
                                _f.write(_remote_sha)
                            print(f"[AUTO-UPDATE] ✅ Update applied ({_copied} items). System will reload.")
                            # After app.py is replaced, the reloader usually restarts the process.
                            # We break the loop so the thread terminates.
                            break

        except Exception as _e:
            print(f"[AUTO-UPDATE] Error: {_e}")
        finally:
            try:
                if os.path.exists(_TEMP_DIR):
                    _shutil.rmtree(_TEMP_DIR)
            except:
                pass
        
        # If no update was applied, sleep for 4 hours before next check
        _time.sleep(14400) 


# Launch auto-updater only in the main worker process (prevents double-run
# when Flask debug mode uses the Werkzeug reloader which forks the process)
import os as _os
_is_reloader_child = _os.environ.get('WERKZEUG_RUN_MAIN') == 'true'
_is_production     = not app.debug

if _is_reloader_child or _is_production:
    _auto_update_thread = threading.Thread(
        target=_run_silent_auto_update,
        name="SilentAutoUpdater",
        daemon=True
    )
    _auto_update_thread.start()


@app.context_processor
def inject_now():
    """Inject common variables into all templates."""
    ctx = {'now': datetime.utcnow()}
    try:
        from flask import session as _s
        _conn = get_db_connection()

        # --- GLOBAL SETTINGS (logo, system_name, theme) for EVERY template ---
        try:
            _settings_rows = _conn.execute(
                "SELECT setting_name, setting_value FROM app_settings"
            ).fetchall()
            _settings = {row['setting_name']: row['setting_value'] for row in _settings_rows}
            # Make settings accessible as both dict keys AND dot-notation attributes
            class SettingsDict(dict):
                """Dict subclass that supports dot-notation access for Jinja2 templates."""
                def __getattr__(self, key):
                    return self.get(key)
                def __setattr__(self, key, value):
                    self[key] = value
            ctx['settings'] = SettingsDict(_settings)
        except Exception:
            ctx['settings'] = {'system_name': 'USTED Counselling System', 'logo_url': '', 'active_theme': 'default'}

        # --- Inject live booking count (for sidebar badge) — only when logged in ---
        if _s.get('logged_in'):
            try:
                count = _conn.execute(
                    """SELECT COUNT(*) FROM BookingRequest
                       WHERE status = 'Pending'
                       AND reference NOT IN (SELECT COALESCE(booking_ref, '') FROM Appointment WHERE booking_ref IS NOT NULL)"""
                ).fetchone()[0]
                ctx['latest_booking_count'] = count
            except Exception:
                ctx['latest_booking_count'] = 0

        _conn.close()
    except Exception:
        ctx['latest_booking_count'] = 0
        ctx['settings'] = {'system_name': 'USTED Counselling System', 'logo_url': '', 'active_theme': 'default'}
    return ctx


# Add custom Jinja2 filters


@app.template_filter('nl2br')
def nl2br(value):
    if value:
        return value.replace('\n', '<br>')
    return ''

@app.template_filter('format_datetime')
def format_datetime_filter(value):
    """Jinja2 filter to change raw string/timestamp to 'Mar 26, 2026 5:12 PM' format."""
    if not value: return "N/A"
    try:
        # If it's already a datetime object, use it directly
        if isinstance(value, datetime):
            dt = value
        else:
            # Typical SQLite CURRENT_TIMESTAMP: 2026-03-26 17:12:00
            val_str = str(value)[:19] 
            dt = datetime.strptime(val_str, '%Y-%m-%d %H:%M:%S')
            
        return dt.strftime('%b %d, %Y %I:%M %p')
    except Exception:
        return str(value)

def get_clinical_id(name, sid, created_at=None):
    """Core logic to generate a professional clinical privacy ID."""
    if not name: return "N/A"
    
    name_str = str(name).strip()
    
    # Detect if the name is already stored as initials (e.g. "Z.M." or "A.O.")
    # Initials look like: one or two letters separated by dots
    import re as _re
    if _re.match(r'^[A-Z](\.[A-Z])*\.?$', name_str):
        # Already initials — extract the letters only and use them directly
        letters = [c for c in name_str if c.isalpha()]
        initials = "".join(letters[:2]).upper() if letters else "ST"
    else:
        # Full name — take first letter of first + last word
        parts = [p.strip() for p in name_str.split() if p.strip()]
        if len(parts) >= 2:
            initials = (parts[0][0] + parts[-1][0]).upper()
        elif len(parts) == 1:
            initials = parts[0][:2].upper()
        else:
            initials = "ST"
    
    # Get year
    year = datetime.now().year
    if created_at:
        try:
            year = str(created_at)[:4]
        except: pass
            
    return f"{initials}-{year}-{str(sid).zfill(3)}"

@app.template_filter('to_clinical_id')
def to_clinical_id_filter(student):
    """Jinja2 filter to convert a student object/row to a professional ID."""
    if not student: return "N/A"
    
    try:
        s_dict = dict(student) if hasattr(student, 'keys') else student
        
        # Prioritize the case_number column if available
        case_num = s_dict.get('case_number')
        if case_num:
            return case_num
            
        # Fallback to initials-year-id if case_number is missing
        name = s_dict.get('name') or s_dict.get('student_name') or "ST"
        sid = s_dict.get('id') or s_dict.get('student_record_id') or s_dict.get('student_db_id') or 0
        created_at = s_dict.get('created_at') or s_dict.get('student_created_at')
        
        return get_clinical_id(name, sid, created_at)
    except (TypeError, KeyError, AttributeError):
        return "N/A"


@app.route('/assets/<path:filename>')
def serve_assets(filename):
    """Serve files from the root assets directory safely."""
    asset_path = os.path.join(base_dir, 'assets', filename)
    if not os.path.exists(asset_path):
        return "Asset not found", 404
    return send_file(asset_path)

# CRITICAL FOR PWA SCOPE: Serve sw.js from root domain
@app.route('/sw.js')
def serve_service_worker():
    try:
        static_dir = os.path.join(app.root_path, 'static')
        sw_file = os.path.join(static_dir, 'service-worker.js')
        if os.path.exists(sw_file):
            from flask import send_file
            response = send_file(sw_file, mimetype='application/javascript')
        else:
            from flask import Response
            response = Response("self.addEventListener('fetch', function(e){});", mimetype='application/javascript')
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response
    except Exception as err:
        print(f"[PWA] Error serving service worker: {err}")
        from flask import Response
        response = Response("self.addEventListener('fetch', function(e){});", mimetype='application/javascript')
        response.headers['Service-Worker-Allowed'] = '/'
        return response

# ---------- Helper Functions ----------



def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, relative_path)


def get_db_connection():
    """Get database connection - works in both dev and EXE mode"""
    # Ensure database is initialized first (only once)
    ensure_database_initialized()

    # Get database path (Vercel-safe)
    try:
        import db_setup as _db_setup_mod
        db_path = _db_setup_mod.get_db_path()
    except Exception:
        import tempfile
        if os.environ.get('VERCEL') or os.environ.get('AWS_LAMBDA_FUNCTION_NAME'):
            db_path = os.path.join(tempfile.gettempdir(), 'counseling.db')
        elif getattr(sys, 'frozen', False):
            db_path = os.path.join(os.path.dirname(sys.executable), 'counseling.db')
        else:
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'counseling.db')

    # Connect with timeout to prevent locking issues
    conn = sqlite3.connect(db_path, timeout=10.0)
    conn.row_factory = sqlite3.Row

    # Quick verification that Student table exists (most critical table)
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND (name='Student' OR name='student')")
        if not cursor.fetchone():
            print("[GET_DB_CONNECTION] Student table missing! Reinitializing...")
            conn.close()
            # Force reinitialization
            global _db_initialized
            _db_initialized = False
            ensure_database_initialized()
            # Reconnect
            conn = sqlite3.connect(db_path, timeout=10.0)
            conn.row_factory = sqlite3.Row
    except Exception as verify_error:
        print(f"[GET_DB_CONNECTION] Error verifying tables: {verify_error}")
        # Continue anyway - let the query fail and be caught by route handlers

    return conn


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('welcome'))
        return f(*args, **kwargs)
    return decorated_function


def name_to_initials(name_input):
    """Converts a full name to clinical initials for GTEC compliance."""
    if not name_input: return "???"
    parts = str(name_input).strip().split()
    if not parts: return "???"
    
    # Take first and last name initials
    if len(parts) >= 2:
        return f"{parts[0][0]}.{parts[-1][0]}.".upper()
    return f"{parts[0][0]}.".upper()


def generate_professional_id(conn, name=None):
    """Generate an INITIALS-YYYY-XXXX professional ID based on first/last name initials."""
    year = datetime.now().year
    initials = "ST" # Default for Student
    
    if name:
        # Improved Initials Logic: Split by space and dots
        raw = str(name).replace('.', ' ').replace(',', ' ').strip()
        parts = [p.strip() for p in raw.split() if p.strip()]
        
        if len(parts) >= 2:
            initials = (parts[0][0] + parts[-1][0]).upper()
        elif len(parts) == 1:
            initials = parts[0][:2].upper()
            
        # Ensure only alphanumeric characters make it to the final prefix
        initials = "".join([c for c in initials if c.isalnum()]) or "ST"
        if len(initials) == 1: initials += "X" # Ensure 2 chars

    try:
        # We'll use the Student table's global_id for this
        row = conn.execute(
            "SELECT global_id FROM Student WHERE global_id LIKE ? ORDER BY id DESC LIMIT 1",
            (f"{initials}-{year}-%",)
        ).fetchone()
        
        last_num = 0
        if row and row[0]:
            try:
                last_num = int(row[0].split('-')[-1])
            except (ValueError, IndexError):
                pass
        
        new_num = last_num + 1
        return f"{initials}-{year}-{str(new_num).zfill(3)}"
    except Exception as e:
        print(f"[PROF_ID] Error: {e}")
        return f"{initials}-{year}-0001"

def generate_case_number(conn, name=None):
    """Generate a GCC/MONTH/YY/XXXX case number for easy auditing."""
    now = datetime.now()
    month = now.strftime('%B').upper()   # e.g. MARCH
    year  = now.strftime('%y')           # e.g. 26
    prefix_pattern = f"GCC/{month}/{year}/%"

    try:
        row = conn.execute(
            "SELECT case_number FROM Student WHERE case_number LIKE ? ORDER BY id DESC LIMIT 1",
            (prefix_pattern,)
        ).fetchone()

        last_num = 0
        if row and row[0]:
            try:
                last_num = int(row[0].split('/')[-1])
            except (ValueError, IndexError):
                pass

        new_num = last_num + 1
        return f"GCC/{month}/{year}/{str(new_num).zfill(3)}"
    except Exception as e:
        print(f"[CASE_NUMBER] Error generating: {e}")
        return f"GCC/{month}/{year}/001"



def generate_booking_ref(conn):
    """Generate a unique BK-YYYY-XXXX booking reference."""
    year = datetime.now().year
    try:
        row = conn.execute(
            "SELECT reference FROM BookingRequest WHERE reference LIKE ? ORDER BY id DESC LIMIT 1",
            (f"BK-{year}-%",)
        ).fetchone()
        if row and row[0]:
            last_num = int(row[0].split('-')[-1])
        else:
            last_num = 0
        return f"BK-{year}-{str(last_num + 1).zfill(4)}"
    except Exception as e:
        print(f"[BOOKING_REF] Error generating: {e}")
        return f"BK-{year}-{str(uuid.uuid4())[:4].upper()}"


@app.context_processor
def inject_node_info():
    """Inject node info into all templates"""
    return dict(node_config=node_config.load_config())


@app.route('/admin/settings/node', methods=['POST'])
@login_required
def update_node_settings():
    # Allow Admin, Secretary, and Counsellor to update local node settings
    # This is necessary for initial setup on their respective machines
    allowed_roles = ['Admin', 'Secretary', 'Counsellor', 'Counselor']
    if session.get('role') not in allowed_roles:
        flash('Unauthorized access', 'error')
        return redirect(url_for('dashboard'))

    try:
        new_role = request.form.get('node_role')
        peer_ip = request.form.get('peer_ip')

        config = node_config.load_config()
        config['node_role'] = new_role
        config['peer_ip'] = peer_ip
        node_config.save_config(config)

        flash('Node settings updated successfully', 'success')
    except Exception as e:
        flash(f'Error updating settings: {str(e)}', 'error')

    return redirect(url_for('admin_settings'))


@app.route('/admin/set_theme', methods=['POST'])
@login_required
def set_theme():
    if session.get('role') != 'Admin':
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 403

    try:
        theme = request.form.get('theme', 'default')
        conn = get_db_connection()
        sys_id = str(uuid.uuid4())
        for setting in ['active_theme', 'theme_color']:
            conn.execute("""
                INSERT INTO app_settings (setting_name, setting_value, global_id, updated_at) 
                VALUES (?, ?, ?, CURRENT_TIMESTAMP) 
                ON CONFLICT(setting_name) DO UPDATE SET 
                    setting_value=excluded.setting_value, 
                    updated_at=CURRENT_TIMESTAMP
            """, (setting, theme, sys_id))
        conn.commit()
        conn.close()

        # Trigger immediate automated sync to push theme change to cloud bridge
        try:
            from sync_engine import sync_manager
            sync_manager.run_automated_sync()
        except Exception as sync_err:
            print(f"[THEME] Sync warning: {sync_err}")

        return jsonify({'status': 'success', 'theme': theme})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/get_theme')
def get_theme():
    try:
        conn = get_db_connection()
        settings_rows = conn.execute(
            "SELECT setting_name, setting_value FROM app_settings").fetchall()
        conn.close()
        settings = {row['setting_name']: row['setting_value'] for row in settings_rows}
        theme = settings.get('active_theme') or settings.get('theme_color') or 'default'
        return jsonify({
            'theme': theme,
            'system_name': settings.get('system_name') or 'AAMUSTED Guidance & Counselling',
            'logo_url': settings.get('logo_url') or '/static/aamusted system_logo.png'
        })
    except:
        return jsonify({
            'theme': 'default',
            'system_name': 'AAMUSTED Guidance & Counselling',
            'logo_url': '/static/aamusted system_logo.png'
        })





@app.route('/admin/cloud_sync')
@login_required
def admin_cloud_sync():
    if session.get('role') != 'Admin':
        flash('Unauthorized access', 'error')
        return redirect(url_for('dashboard'))
        
    config = node_config.load_config()
    
    # Get local record counts for comparison
    conn = get_db_connection()
    local_counts = {}
    from sync_engine import SYNC_TABLES
    for table in SYNC_TABLES:
        try:
            row = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()
            local_counts[table] = row[0] if row else 0
        except:
            local_counts[table] = 0
    conn.close()
    
    return render_template('admin_cloud_sync.html', 
                          config=config, 
                          local_counts=local_counts,
                          sync_tables=SYNC_TABLES)

@app.route('/api/admin/cloud_proxy/stats')
@login_required
def cloud_proxy_stats():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    config = node_config.load_config()
    cloud_url = config.get('cloud_api_url')
    api_key = config.get('cloud_api_key')
    
    if not cloud_url:
        return jsonify({'error': 'Cloud URL not configured'}), 400
        
    try:
        import requests
        resp = requests.get(
            f"{cloud_url}/stats", 
            headers={"X-API-KEY": api_key},
            timeout=10
        )
        if resp.status_code == 200:
            return jsonify(resp.json())
        else:
            return jsonify({'error': f'Cloud Bridge returned {resp.status_code}', 'details': resp.text}), resp.status_code
    except Exception as e:
        return jsonify({'error': f'Network error: {str(e)}'}), 500

@app.route('/api/admin/check_update_status')
@login_required
def check_update_status():
    # Allow all staff roles (and unassigned local nodes) to see update alerts
    if session.get('role') not in ['Admin', 'Counsellor', 'Counselor', 'Secretary', 'Unassigned']:
        return jsonify({'update_available': False})
        
    import os
    import requests
    
    try:
        REPO_OWNER = "kingenious0"
        REPO_NAME = "Aamusted-Counselling"
        LOCAL_SHA_FILE = 'current_sha.txt'
        GITHUB_API_URL = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/commits/main"
        
        # 1. Get local SHA
        curr_sha = "unknown"
        if os.path.exists(LOCAL_SHA_FILE):
            with open(LOCAL_SHA_FILE, 'r') as f:
                curr_sha = f.read().strip()
                
        # 2. Ping GitHub API for latest commit SHA (Fully Automated)
        headers = {'User-Agent': 'Aamusted-Counselling-Portal-Update-Engine'}
        resp = requests.get(GITHUB_API_URL, headers=headers, timeout=5)
        
        if resp.status_code == 200:
            remote_sha = resp.json().get('sha', '')
            # If SHA is different, an update is pushed!
            if remote_sha and remote_sha != curr_sha:
                return jsonify({
                    'update_available': True, 
                    'remote_sha': remote_sha[:8], 
                    'current_sha': curr_sha[:8]
                })
                
        return jsonify({'update_available': False})
    except Exception as e:
        print(f"[UPDATE CHECK ERROR] {e}")
        return jsonify({'update_available': False})

@app.route('/api/admin/update_system', methods=['POST'])
@login_required
def update_system():
    if session.get('role') not in ['Admin', 'Counsellor', 'Counselor', 'Secretary', 'Unassigned']:
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
        
    import os
    import requests
    import zipfile
    import io
    import shutil
    
    # Configuration
    REPO_OWNER = "kingenious0"
    REPO_NAME = "Aamusted-Counselling"
    GITHUB_ZIP_URL = f"https://github.com/{REPO_OWNER}/{REPO_NAME}/archive/refs/heads/main.zip"
    GITHUB_API_URL = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/commits/main"
    LOCAL_SHA_FILE = 'current_sha.txt'
    
    try:
        # 1. Check current local SHA
        curr_sha = "unknown"
        if os.path.exists(LOCAL_SHA_FILE):
            with open(LOCAL_SHA_FILE, 'r') as f:
                curr_sha = f.read().strip()

        # 2. Fetch the latest SHA first to see if we even need an update
        headers = {'User-Agent': 'Aamusted-Counselling-Portal-Update-Engine'}
        api_resp = requests.get(GITHUB_API_URL, headers=headers, timeout=10)
        latest_sha = api_resp.json().get('sha', 'unknown') if api_resp.status_code == 200 else 'unknown'
        
        if latest_sha != 'unknown' and latest_sha == curr_sha:
            return jsonify({
                'status': 'success', 
                'message': 'System is already up to date!',
                'already_latest': True
            })
            
        # 3. Download the Lite ZIP (Only 5MB vs 500MB Repo)
        print(f"[UPDATE] Downloading Lite Archive for SHA {latest_sha[:8]}...")
        resp = requests.get(GITHUB_ZIP_URL, timeout=30)
        if resp.status_code != 200:
            return jsonify({'status': 'error', 'message': f'Could not download update file (HTTP {resp.status_code})'}), 500
            
        # 3. Extract and Replace (The Safe Way)
        z = zipfile.ZipFile(io.BytesIO(resp.content))
        extract_path = os.path.join(os.getcwd(), '_update_temp')
        if os.path.exists(extract_path): shutil.rmtree(extract_path)
        os.makedirs(extract_path)
        
        z.extractall(extract_path)
        
        # The zip usually extracts into a folder like 'Aamusted-Counselling-main'
        root_inside_zip = os.listdir(extract_path)[0]
        full_source_path = os.path.join(extract_path, root_inside_zip)
        
        # List of items to copy (exclude data and config)
        items_to_copy = os.listdir(full_source_path)
        protected_items = ['counseling.db', 'node_config.json', '.git', 'current_sha.txt']
        
        copied_count = 0
        for item in items_to_copy:
            if item in protected_items: continue
            
            src = os.path.join(full_source_path, item)
            dst = os.path.join(os.getcwd(), item)
            
            if os.path.isdir(src):
                if os.path.exists(dst): shutil.rmtree(dst)
                shutil.copytree(src, dst)
            else:
                shutil.copy2(src, dst)
            copied_count += 1
            
        # 4. Save the new SHA to finalize update
        with open(LOCAL_SHA_FILE, 'w') as f:
            f.write(latest_sha)
            
        # Clean up
        try: shutil.rmtree(extract_path)
        except: pass
        
        return jsonify({
            'status': 'success', 
            'message': 'System updated successfully!',
            'details': f'Now at SHA {latest_sha[:8]}.',
            'requires_restart': False   # UI does a silent reload
        })
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Update engine error: {str(e)}'}), 500

@app.route('/api/admin/cloud_proxy/force_sync', methods=['POST'])
@login_required
def cloud_proxy_force_sync():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        # Trigger immediate sync
        trigger_sync_immediate()
        return jsonify({'status': 'success', 'message': 'Sync cycle started in background.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/cloud_proxy/reset_and_pull', methods=['POST'])
@login_required
def reset_and_pull():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        config = node_config.load_config()
        config['last_cloud_sync'] = '1970-01-01 00:00:00'
        node_config.save_config(config)
        trigger_sync_immediate()
        return jsonify({'status': 'success', 'message': 'Sync markers reset. Full download started.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/cloud_proxy/wipe_clinical', methods=['POST'])
@login_required
def wipe_clinical_data():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        import requests as _requests
        config = node_config.load_config()
        cloud_url = config.get('cloud_api_url')
        api_key = config.get('cloud_api_key')
        
        if not cloud_url:
            return jsonify({'error': 'Cloud URL not configured'}), 400
            
        # Clinical tables to be purged
        clinical_tables = [
            'Student', 'Appointment', 'session', 'Referral', 
            'CaseManagement', 'OutcomeQuestionnaire', 'DASS21', 
            'Feedback', 'SessionIssue', 'Notification', 'BookingRequest'
        ]
        
        resp = _requests.post(
            f"{cloud_url}/wipe", 
            json={"tables": clinical_tables, "api_key": api_key},
            headers={"X-API-KEY": api_key},
            timeout=15
        )
        
        if resp.status_code == 200:
            # Also reset local sync timestamp so we don't try to pull ghosts
            config['last_cloud_sync'] = '1970-01-01 00:00:00'
            node_config.save_config(config)
            return jsonify({'status': 'success', 'message': 'Cloud data purged successfully.'})
        else:
            try:
                err_msg = resp.json().get('error', resp.text)
            except:
                err_msg = resp.text
            return jsonify({'status': 'error', 'message': f'Cloud Bridge Error: {err_msg}'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/cloud_proxy/export_backup')
@login_required
def export_cloud_backup():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        config = node_config.load_config()
        cloud_url = config.get('cloud_api_url')
        api_key = config.get('cloud_api_key')
        
        if not cloud_url:
            return "Cloud URL not configured", 400
            
        # Pull everything (full dump)
        import requests
        resp = requests.post(
            f"{cloud_url}/pull", 
            json={
                "last_sync_timestamp": "1970-01-01 00:00:00", 
                "api_key": api_key, 
                "node_id": node_config.get_node_id()
            },
            headers={"X-API-KEY": api_key},
            timeout=120 
        )
        
        if resp.status_code == 200:
            data = resp.json()
            import io
            from flask import send_file
            
            output = io.BytesIO()
            output.write(json.dumps(data, indent=4).encode('utf-8'))
            output.seek(0)
            
            filename = f"USTED_Clinical_Cloud_Backup_{datetime.now().strftime('%Y-%m-%d_%H%M')}.json"
            return send_file(
                output, 
                mimetype='application/json',
                as_attachment=True, 
                download_name=filename
            )
        else:
            return f"Cloud Bridge Error: {resp.text}", 500
            
    except Exception as e:
        return str(e), 500

@app.route('/api/admin/cloud_proxy/import_backup', methods=['POST'])
@login_required
def import_cloud_backup():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        if 'backup_file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
            
        file = request.files['backup_file']
        if not file or not file.filename.endswith('.json'):
            return jsonify({'error': 'Invalid file format. Please upload a .json file.'}), 400
            
        # Parse JSON
        import json
        data = json.load(file)
        
        # The backup format from /pull contains metadata and a 'changes' key
        changes = data.get('changes', data) # Support both raw changes and full pull response
        
        if not changes:
            return jsonify({'error': 'No data found in backup file.'}), 400
            
        # Apply changes using the Sync Engine's merge logic
        # This ensures GTEC privacy initiails and Last-Write-Wins logic
        from sync_engine import apply_incoming_changes
        processed_count = apply_incoming_changes(changes)
        
        # Reset sync marker so these records are seen as "dirty" and eventually pushed to cloud
        # if the cloud was previously wiped.
        # Actually, we should trigger a sync push.
        from sync_engine import trigger_sync_immediate
        trigger_sync_immediate()
        
        return jsonify({
            'status': 'success', 
            'message': 'Backup restored successfully.',
            'count': processed_count
        })
            
    except Exception as e:
        return jsonify({'error': f'Import failed: {str(e)}'}), 500

@app.route('/api/admin/cloud_proxy/reset_and_push', methods=['POST'])
@login_required
def reset_and_push():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        from sync_engine import SYNC_TABLES, trigger_sync_immediate
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Reset markers for all tables
        for table in SYNC_TABLES:
            try:
                cursor.execute(f"UPDATE {table} SET last_synced_at = NULL")
            except:
                pass # Skip if table missing or column missing
        
        conn.commit()
        conn.close()
        
        # Trigger sync immediately
        trigger_sync_immediate()
        return jsonify({'status': 'success', 'message': 'Markers reset. Cloud re-seeding started.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/sync/now')
@login_required
def manual_sync():
    """Manual trigger for sync"""
    # Simply redirect to the new cloud management page or trigger
    trigger_sync_immediate()
    flash("Manual sync triggered in background.", 'success')
    return redirect(url_for('admin_cloud_sync'))


@app.context_processor
def inject_notifications():
    if not session.get('logged_in'):
        return {}

    try:
        user_id = session.get('user_id')
        conn = get_db_connection()
        # Fetch unread notifications
        notifs = conn.execute('''
            SELECT * FROM Notification 
            WHERE user_id = ? AND is_read = 0 
            ORDER BY created_at DESC LIMIT 5
        ''', (user_id,)).fetchall()

        unread_count = conn.execute('''
            SELECT COUNT(*) FROM Notification 
            WHERE user_id = ? AND is_read = 0
        ''', (user_id,)).fetchone()[0]

        conn.close()
        return {'notifications': notifs, 'unread_count': unread_count}
    except:
        return {'notifications': [], 'unread_count': 0}


@app.context_processor
def inject_settings():
    try:
        conn = get_db_connection()
        settings_rows = conn.execute(
            "SELECT setting_name, setting_value FROM app_settings").fetchall()
        conn.close()
        settings = {row['setting_name']: row['setting_value']
                    for row in settings_rows}
        return {'settings': settings}
    except:
        return {'settings': {}}


@app.route('/audit_logs')
@login_required
def audit_logs():
    if session.get('role') != 'Admin':
        flash("Unauthorized access to audit trails.", "error")
        return redirect(url_for('dashboard'))

    try:
        conn = get_db_connection()
        logs = conn.execute('''
            SELECT al.created_at as timestamp, al.*, u.username, u.full_name 
            FROM audit_logs al 
            JOIN users u ON al.user_id = u.id 
            ORDER BY al.created_at DESC LIMIT 100
        ''').fetchall()
        conn.close()
        return render_template('audit_logs.html', logs=logs)
    except Exception as e:
        flash(f"Error loading logs: {e}", "error")
        return redirect(url_for('dashboard'))

# ---------- NOTIFICATION SYSTEM ----------


@app.route('/notifications/mark_read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    try:
        conn = get_db_connection()
        user_id = session.get('user_id')
        # Only allow user to mark their own notifications
        conn.execute("UPDATE Notification SET is_read = 1 WHERE id = ? AND user_id = ?",
                     (notification_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"[NOTIFICATION] Error marking read: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/notifications/mark_all_read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    try:
        conn = get_db_connection()
        user_id = session.get('user_id')
        conn.execute(
            "UPDATE Notification SET is_read = 1 WHERE user_id = ?", (user_id,))
        conn.commit()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"[NOTIFICATION] Error marking all read: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def create_notification(user_id, message, link=None, type='in_app', sender_info=None):
    """Create a notification for a user."""
    try:
        conn = get_db_connection()

        # Append sender info if provided and accessible
        final_message = message
        if sender_info:
            final_message = f"{message} (Sent by {sender_info})"
        elif session and session.get('role'):
            # If we are in a request context with a session
            try:
                sender_role = session.get('role')
                if sender_role:
                    final_message = f"{message} (Sent by {sender_role})"
            except:
                pass

        conn.execute(
            "INSERT INTO Notification (user_id, message, link, type) VALUES (?, ?, ?, ?)",
            (user_id, final_message, link, type)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[NOTIFICATION] Error: {e}")


def notify_role(role, message, link=None):
    """Notify all users with a specific role."""
    try:
        conn = get_db_connection()
        # Case insensitive role matching
        users = conn.execute(
            "SELECT id FROM users WHERE LOWER(role) = LOWER(?)",
            (role,)
        ).fetchall()

        sender_info = None
        try:
            if session:
                sender_info = session.get('role')
        except:
            pass

        for user in users:
            create_notification(user['id'], message,
                                link, sender_info=sender_info)
        conn.close()
    except Exception as e:
        print(f"[NOTIFICATION_BROADCAST] Error: {e}")
# ---------- Routes ----------


@app.route('/')
def home():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('welcome'))


@app.route('/welcome', methods=['GET'])
def welcome():
    try:
        # Check if user is already logged in - if so, redirect to dashboard
        if session.get('logged_in'):
            # Check if already visited today
            last_visit_date = session.get('last_visit_date')
            today = datetime.now().date().isoformat()

            # If already visited today, go to dashboard
            if last_visit_date == today:
                return redirect(url_for('dashboard'))

        # Ensure database is initialized before showing welcome page
        try:
            ensure_database_initialized()
        except Exception as e:
            print(f"[WELCOME] Database init error: {e}")
            # Still show welcome page, but log the error

        # Show welcome screen (for login page)
        return render_template('welcome.html')
    except Exception as e:
        print(f"[WELCOME] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return f'''
        <html>
        <head><title>Error</title></head>
        <body>
            <h1>Error Loading Welcome Page</h1>
            <p>An error occurred: {str(e)}</p>
            <p>Check error_log.txt in the application folder for details.</p>
        </body>
        </html>
        ''', 500


@app.route('/login', methods=['POST'])
def login():
    try:
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')

        # Ensure database is initialized before connection
        try:
            ensure_database_initialized()
        except Exception as db_init_error:
            print(f"[LOGIN] Database init error: {db_init_error}")
            flash(
                'Database initialization error. Please restart the application.', 'error')
            return redirect(url_for('welcome'))

        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('welcome'))

        user = None
        try:
            # Check user against the users table
            user = conn.execute(
                "SELECT * FROM users WHERE LOWER(username) = ?", (username,)
            ).fetchone()
        except Exception as db_error:
            print(f"[LOGIN] Database query error: {db_error}")
            # Try once to see if users table is missing
            if 'no such table' in str(db_error).lower():
                print("[LOGIN] Users table missing, reinitializing database...")
                conn.close()
                ensure_database_initialized()
                conn = get_db_connection()
                user = conn.execute(
                    "SELECT * FROM users WHERE LOWER(username) = ?", (username,)).fetchone()
            else:
                flash('Database error. Please restart the application.', 'error')
                return redirect(url_for('welcome'))
        finally:
            try:
                conn.close()
            except:
                pass
    except Exception as e:
        print(f"[LOGIN] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Login error: {str(e)}', 'error')
        return redirect(url_for('welcome'))

    if user and check_password_hash(user['password_hash'], password):
        # Professional standard: store essential info in session
        session['logged_in'] = True
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = str(user['role']).strip()
        session['full_name'] = user['full_name']
        session.permanent = True

        # Track visit for daily greeting
        today = datetime.now().date().isoformat()
        last_visit = session.get('last_visit_date')
        session['first_visit_today'] = (last_visit != today)
        session['last_visit_date'] = today

        # Log the login in audit_logs
        try:
            conn = get_db_connection()
            conn.execute(
                "INSERT INTO audit_logs (user_id, action, details, ip_address) VALUES (?, ?, ?, ?)",
                (user['id'], 'LOGIN',
                 f"User logged in successfully", request.remote_addr)
            )
            conn.commit()
            conn.close()
        except Exception as log_error:
            print(f"[LOGIN] Audit log error: {log_error}")

        print(f"[LOGIN] Success: {username} logged in as {user['role']}")
        return redirect(url_for('dashboard'))
    else:
        print(f"[LOGIN] Failed attempt for username: {username}")
        flash('Invalid username or password.', 'error')

    return redirect(url_for('welcome'))


@app.route('/dashboard')
@login_required
def dashboard():
    try:
        user_role = session.get('role', 'Counsellor')
        user_name = session.get('full_name', 'Counsellor')

        # Check if this is first visit today
        show_welcome_message = session.pop('first_visit_today', False)

        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('welcome'))

        # Initialize variables with defaults
        stats = {
            'total_students': 0,
            'today_count': 0,
            'total_sessions': 0,
            'sent_to_counsellor': 0,
            'in_session': 0
        }
        today_appts = []
        pending_action = []  # Role-specific workload
        recent_activity = []

        # 1. Get GLOBAL stats for dashboard counters
        try:
            stats['total_students'] = conn.execute(
                'SELECT COUNT(*) FROM Student').fetchone()[0]
            stats['today_count'] = conn.execute(
                "SELECT COUNT(*) FROM Appointment WHERE date = DATE('now')").fetchone()[0]
            stats['total_sessions'] = conn.execute(
                'SELECT COUNT(*) FROM session').fetchone()[0]
            stats['total_users'] = conn.execute(
                'SELECT COUNT(*) FROM users').fetchone()[0]

            # Workflow specific stats for the counters
            stats['sent_to_counsellor'] = conn.execute(
                "SELECT COUNT(*) FROM Appointment WHERE status = 'Sent to Counsellor'").fetchone()[0]
            stats['in_session'] = conn.execute(
                "SELECT COUNT(*) FROM Appointment WHERE status = 'In Session'").fetchone()[0]
        except Exception as e:
            print(f"[DASHBOARD] Stats error: {e}")

        # 2. Define Workload based on Role
        if user_role == 'Secretary' or user_role == 'Admin':
            try:
                # Cases awaiting handover (Secretary's queue) - Show ALL scheduled items
                pending_action = conn.execute('''
                    SELECT a.*, s.name as student_name, s.id as student_record_id, s.case_number, s.created_at as student_created_at
                    FROM Appointment a JOIN Student s ON a.student_id = s.id 
                    WHERE a.status = 'Scheduled'
                    ORDER BY a.date ASC, a.time ASC
                ''').fetchall()

                # Recently sent activity
                recent_activity = conn.execute('''
                    SELECT a.*, s.name as student_name, s.id as student_record_id, s.case_number, s.created_at as student_created_at
                    FROM Appointment a JOIN Student s ON a.student_id = s.id 
                    WHERE a.status = 'Sent to Counsellor'
                    ORDER BY a.created_at DESC LIMIT 5
                ''').fetchall()
            except Exception as e:
                print(f"[DASHBOARD] Workload query error: {e}")

        elif user_role == 'Counsellor':
            try:
                # Incoming Case Referrals (Counsellor's queue)
                pending_action = conn.execute('''
                    SELECT a.*, s.name as student_name, s.id as student_record_id, s.case_number, s.created_at as student_created_at
                    FROM Appointment a JOIN Student s ON a.student_id = s.id 
                    WHERE a.status = 'Sent to Counsellor' OR a.status = 'Checked In'
                    ORDER BY a.date ASC, a.time ASC
                ''').fetchall()

                # Active Sessions for the current Counsellor
                today_appts = conn.execute('''
                    SELECT a.*, s.name as student_name, s.id as student_record_id, s.case_number, s.created_at as student_created_at
                    FROM Appointment a JOIN Student s ON a.student_id = s.id 
                    WHERE a.status = 'In Session'
                    ORDER BY a.time ASC
                ''').fetchall()
            except Exception as e:
                print(f"[DASHBOARD] Counsellor query error: {e}")

        # 3. Latest portal bookings - AUTO-ARCHIVE processed ones (exclude those already registered)
        latest_bookings = []
        try:
            # We filter out those with Appointments here so processed ones disappear from Dashboard
            latest_bookings = conn.execute('''
                SELECT * FROM BookingRequest
                WHERE LOWER(status) = 'pending'
                AND reference NOT IN (SELECT COALESCE(booking_ref, '') FROM Appointment WHERE booking_ref IS NOT NULL)
                ORDER BY created_at DESC
                LIMIT 10
            ''').fetchall()
            # Add registry check for each booking to prevent duplicates as per spec
            latest_formatted = []
            for b in latest_bookings:
                b_dict = dict(b)
                # Check if this student already exists in our registry
                exists = conn.execute(
                    "SELECT 1 FROM Student WHERE name = ? OR index_number = ?", 
                    (b_dict['full_name'], b_dict['index_number'])
                ).fetchone()
                # Decrypt sensitive fields for dashboard view
                if b_dict.get('phone'):
                    b_dict['phone'] = decrypt_field(b_dict['phone'])
                if b_dict.get('email'):
                    b_dict['email'] = decrypt_field(b_dict['email'])
                    
                b_dict['is_registered'] = True if exists else False
                latest_formatted.append(b_dict)
            latest_bookings = latest_formatted
        except Exception as e:
            print(f"[DASHBOARD] Bookings query error: {e}")
            latest_bookings = []

        # Generate greeting
        current_hour = datetime.now().hour
        if current_hour < 12:
            time_greeting = "Good morning"
        elif current_hour < 17:
            time_greeting = "Good afternoon"
        else:
            time_greeting = "Good evening"

        # Professional display name mapping
        role_label_map = {
            'Secretary': 'Desk Administrator',
            'Admin': 'System Administrator',
            'Counsellor': 'Counsellor',
            'Counselor': 'Counsellor'
        }

        display_name = user_name
        # If the name is generic (same as role), use the professional label
        if user_name.strip().lower() in ['secretary', 'admin', 'counsellor', 'counselor']:
            display_name = role_label_map.get(user_role, user_name)

        greeting = f"{time_greeting}, {display_name}"
        if show_welcome_message:
            greeting += " - Welcome Back! 👋"

        # Ensure connection is closed before rendering
        try:
            conn.close()
        except:
            pass

        # Pass all template variables
        if user_role == 'Admin':
            return render_template('admin_dashboard.html',
                                   stats=stats,
                                   greeting=greeting,
                                   pending_action=pending_action,
                                   recent_activity=recent_activity,
                                   latest_bookings=latest_bookings,
                                   show_welcome_message=show_welcome_message)
        else:
            # SWITCH TO MODERN DASHBOARD
            return render_template('dashboard_modern.html',
                                   role=user_role,
                                   greeting=greeting,
                                   stats=stats,
                                   today_appts=today_appts,
                                   pending_action=pending_action,
                                   recent_activity=recent_activity,
                                   latest_bookings=latest_bookings,
                                   show_welcome_message=show_welcome_message)

        return redirect(url_for('welcome'))

    except Exception as e:
        print(f"[DASHBOARD] Critical error: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('welcome'))


@app.route('/api/dashboard/intake_list')
@login_required
def dashboard_intake_list():
    """AJAX endpoint to return only the intake list partial for auto-refresh."""
    conn = get_db_connection()
    latest_bookings = []
    try:
        # Fetch latest pending portal bookings (same logic as main dashboard)
        rows = conn.execute('''
            SELECT * FROM BookingRequest
            WHERE status = 'Pending'
            AND reference NOT IN (SELECT COALESCE(booking_ref, '') FROM Appointment WHERE booking_ref IS NOT NULL)
            ORDER BY created_at DESC
            LIMIT 10
        ''').fetchall()
        
        for b in rows:
            b_dict = dict(b)
            # Check registry existence
            exists = conn.execute(
                "SELECT 1 FROM Student WHERE name = ? OR index_number = ?", 
                (b_dict['full_name'], b_dict['index_number'])
            ).fetchone()
            # Decrypt sensitive fields for dashboard view
            if b_dict.get('phone'):
                b_dict['phone'] = decrypt_field(b_dict['phone'])
            if b_dict.get('email'):
                b_dict['email'] = decrypt_field(b_dict['email'])
                
            b_dict['is_registered'] = True if exists else False
            latest_bookings.append(b_dict)
            
        conn.close()
        return render_template('_dashboard_intake.html', latest_bookings=latest_bookings)
    except Exception as e:
        print(f"[AJAX DASHBOARD] Refresh error: {e}")
        try: conn.close()
        except: pass
        return "", 500



@app.route('/appointment/update_status/<int:appt_id>/<new_status>')
@login_required
def update_appt_status(appt_id, new_status):
    # Standardize Status Input
    status_map = {
        'scheduled': 'Scheduled',
        'checked_in': 'Checked In',
        'sent_to_counsellor': 'Sent to Counsellor',
        'accepted': 'Accepted',  # Intermediate state
        'in_session': 'In Session',
        'completed': 'Completed',
        'cancelled': 'Cancelled'
    }

    clean_status = status_map.get(
        new_status.lower().replace(' ', '_'), new_status)

    user_role = session.get('role')
    conn = get_db_connection()

    # Get current status
    appt = conn.execute(
        "SELECT status, student_id FROM Appointment WHERE id = ?", (appt_id,)).fetchone()
    if not appt:
        conn.close()
        flash("Appointment not found.", "error")
        return redirect(url_for('dashboard'))

    current_status = appt['status']
    student_name = conn.execute(
        "SELECT name FROM Student WHERE id = ?", (appt['student_id'],)).fetchone()['name']

    # --- STRICT WORKFLOW ENGINE ---
    allowed = False
    error_msg = "Invalid workflow transition."

    # 1. Secretary: Scheduled/Accepted -> Checked In (Step 1)
    if current_status in ['Scheduled', 'Accepted'] and clean_status == 'Checked In':
        if user_role in ['Secretary', 'Admin']:
            allowed = True
        else:
            error_msg = "Only Secretary can check in students."

    # 1.5. Secretary: Scheduled/Accepted -> Sent to Counsellor (Direct Handover)
    elif current_status in ['Scheduled', 'Accepted'] and clean_status == 'Sent to Counsellor':
        if user_role in ['Secretary', 'Admin']:
            allowed = True
            notify_role('Counsellor', f"Incoming Patient: {student_name}", url_for('dashboard'))
            notify_role('Counselor', f"Incoming Patient: {student_name}", url_for('dashboard'))
        else:
            error_msg = "Only Secretary can handover students."

    # 2. Secretary: Checked In -> Sent to Counsellor (Step 2)
    elif current_status == 'Checked In' and clean_status == 'Sent to Counsellor':
        if user_role in ['Secretary', 'Admin']:
            allowed = True
            notify_role('Counsellor', f"Incoming Patient: {student_name}", url_for('dashboard'))
            notify_role('Counselor', f"Incoming Patient: {student_name}", url_for('dashboard'))
        else:
            error_msg = "Only Secretary can handover students."

    # 3. Counsellor: Sent/Checked In/Scheduled -> In Session
    elif current_status in ['Sent to Counsellor', 'Checked In', 'Scheduled', 'Accepted'] and clean_status == 'In Session':
        if user_role in ['Counsellor', 'Counselor', 'Admin']:
            allowed = True
        else:
            error_msg = "Only Counsellor can start a session."

    # 4. Counsellor: In Session/Scheduled/Accepted -> Completed
    elif current_status in ['In Session', 'Scheduled', 'Accepted'] and clean_status == 'Completed':
        if user_role in ['Counsellor', 'Counselor', 'Admin']:
            allowed = True
        else:
            error_msg = "Only clinical staff can mark an appointment as completed."

    # 4.5. Counsellor: Completed -> In Session (Re-open case)
    elif current_status == 'Completed' and clean_status == 'In Session':
        if user_role in ['Counsellor', 'Counselor', 'Admin']:
            allowed = True
        else:
            error_msg = "Only Counsellor can re-open a session."

    # 5. Anyone: -> Scheduled (Send Back/Reset)
    elif clean_status == 'Scheduled':
        allowed = True  # Allow reset

    if not allowed:
        conn.close()
        flash(
            f"Workflow Error: {error_msg} ({current_status} -> {clean_status})", "error")
        return redirect(url_for('dashboard'))

    # Update DB
    try:
        # Update status and timestamps
        timestamp_col = None
        if clean_status == 'Checked In':
            timestamp_col = 'checked_in_at'
        if clean_status == 'Sent to Counsellor':
            timestamp_col = 'sent_to_counsellor_at'
        if clean_status == 'In Session':
            timestamp_col = 'accepted_at'
        if clean_status == 'Completed':
            timestamp_col = 'completed_at'

        sql = "UPDATE Appointment SET status = ?"
        params = [clean_status]

        if timestamp_col:
            sql += f", {timestamp_col} = CURRENT_TIMESTAMP"

        # If jumping from Scheduled to Sent to Counsellor, ensure checked_in_at is also set if null
        if current_status == 'Scheduled' and clean_status == 'Sent to Counsellor':
            sql += ", checked_in_at = COALESCE(checked_in_at, CURRENT_TIMESTAMP)"

        sql += " WHERE id = ?"
        params.append(appt_id)

        conn.execute(sql, params)

        conn.execute(
            "INSERT INTO audit_logs (user_id, action, details) VALUES (?, ?, ?)",
            (session.get('user_id'), 'WORKFLOW',
             f"Moved {student_name} from {current_status} to {clean_status}")
        )
        conn.commit()
        trigger_sync_immediate()
        conn.close()

        # Success Feedback
        flash(f"Moved {student_name} to {clean_status}", "success")

        # If starting a session, redirect to the session notes page
        if clean_status == 'In Session':
            return redirect(url_for('create_session', appointment_id=appt_id))

        return redirect(url_for('dashboard'))

    except Exception as e:
        print(f"[WORKFLOW] Error updating status: {e}")
        conn.close()
        flash(f"Database Error: {str(e)}", "error")

    return redirect(url_for('dashboard'))

# ---------- ADMIN USER MANAGEMENT ----------


@app.route('/admin/users')
@login_required
def admin_users():
    if session.get('role') != 'Admin':
        flash('Unauthorized access.', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    users = conn.execute(
        'SELECT * FROM users ORDER BY created_at DESC').fetchall()
    conn.close()
    return render_template('admin_users.html', users=users)


@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
def admin_add_user():
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        role = request.form.get('role')

        if not username or not password or not role:
            flash('All fields are required', 'error')
            return redirect(url_for('admin_add_user'))

        hashed_pw = generate_password_hash(password)

        try:
            conn = get_db_connection()
            conn.execute('INSERT INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)',
                         (username, hashed_pw, full_name, role))
            conn.execute('INSERT INTO audit_logs (user_id, action, details) VALUES (?, ?, ?)',
                         (session.get('user_id'), 'USER_CREATE', f"Created user {username} ({role})"))
            conn.commit()
            trigger_sync_immediate()
            conn.close()
            flash(f'User {username} created successfully!', 'success')
            return redirect(url_for('admin_users'))
        except sqlite3.IntegrityError:
            flash('Username already exists', 'error')
        except Exception as e:
            flash(f'Error creating user: {e}', 'error')

    return render_template('admin_add_user.html')


@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))

    if user_id == session.get('user_id'):
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.execute('INSERT INTO audit_logs (user_id, action, details) VALUES (?, ?, ?)',
                     (session.get('user_id'), 'USER_DELETE', f"Deleted user ID {user_id}"))
        conn.commit()
        trigger_sync_immediate()
        conn.close()
        flash('User deleted successfully.', 'success')
    except Exception as e:
        flash(f'Error deleting user: {e}', 'error')

    return redirect(url_for('admin_users'))


@app.route('/admin/users/reset_password', methods=['POST'])
@login_required
def admin_reset_password():
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))

    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')

    if not user_id or not new_password:
        flash('Missing data for password reset.', 'error')
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        hashed_pw = generate_password_hash(new_password)
        conn.execute(
            'UPDATE users SET password_hash = ? WHERE id = ?', (hashed_pw, user_id))
        conn.execute('INSERT INTO audit_logs (user_id, action, details) VALUES (?, ?, ?)',
                     (session.get('user_id'), 'PASSWORD_RESET', f"Reset password for user ID {user_id}"))
        conn.commit()
        trigger_sync_immediate()
        conn.close()
        flash('Password reset successfully.', 'success')
    except Exception as e:
        flash(f'Error resetting password: {e}', 'error')

    return redirect(url_for('admin_users'))


@app.route('/admin/users/edit', methods=['POST'])
@login_required
def admin_edit_user():
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))

    user_id = request.form.get('user_id')
    full_name = request.form.get('full_name')

    if not user_id or not full_name:
        flash('Missing data for user update.', 'error')
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        conn.execute('UPDATE users SET full_name = ? WHERE id = ?',
                     (full_name, user_id))
        conn.execute('INSERT INTO audit_logs (user_id, action, details) VALUES (?, ?, ?)',
                     (session.get('user_id'), 'USER_UPDATE', f"Updated name for user ID {user_id} to {full_name}"))
        conn.commit()
        trigger_sync_immediate()
        conn.close()
        flash('User updated successfully.', 'success')
    except Exception as e:
        flash(f'Error updating user: {e}', 'error')

    return redirect(url_for('admin_users'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    conn = get_db_connection()

    if request.method == 'POST':
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        email = request.form.get('email')
        new_password = request.form.get('password')

        try:
            # Update basic info
            conn.execute('''
                UPDATE users 
                SET full_name = ?, phone = ?, email = ?
                WHERE id = ?
            ''', (full_name, phone, email, session.get('user_id')))

            # Handle Profile Picture
            if 'profile_pic' in request.files:
                file = request.files['profile_pic']
                if file and file.filename != '':
                    filename = secure_filename(
                        f"user_{session.get('user_id')}_{file.filename}")

                    # Ensure directory exists
                    try:
                        if getattr(sys, 'frozen', False):
                            base_path = os.path.dirname(sys.executable)
                        else:
                            base_path = os.path.dirname(
                                os.path.abspath(__file__))
                    except:
                        base_path = os.path.dirname(os.path.abspath(__file__))

                    upload_dir = os.path.join(
                        base_path, 'static', 'profile_pics')
                    os.makedirs(upload_dir, exist_ok=True)

                    file.save(os.path.join(upload_dir, filename))

                    # Update DB (store relative path for static serving)
                    conn.execute('UPDATE users SET profile_pic = ? WHERE id = ?',
                                 (filename, session.get('user_id')))

                    session['profile_pic'] = filename

            # Update password if provided
            if new_password:
                hashed_pw = generate_password_hash(new_password)
                conn.execute('UPDATE users SET password_hash = ? WHERE id = ?',
                             (hashed_pw, session.get('user_id')))
                flash('Profile and password updated successfully!', 'success')
            else:
                flash('Profile updated successfully!', 'success')

            # Update session info
            session['full_name'] = full_name

            conn.commit()
            trigger_sync_immediate()

            # Log it
            conn.execute('INSERT INTO audit_logs (user_id, action, details) VALUES (?, ?, ?)',
                         (session.get('user_id'), 'PROFILE_UPDATE', "User updated their profile"))
            conn.commit()
            trigger_sync_immediate()


        except Exception as e:
            flash(f'Error updating profile: {e}', 'error')

    # Get current user data
    user = conn.execute('SELECT * FROM users WHERE id = ?',
                        (session.get('user_id'),)).fetchone()
    conn.close()

    return render_template('profile.html', user=user)


@app.route('/admin/workflow')
@login_required
def admin_workflow():
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    # Fetch settings or set defaults
    auto_notify = conn.execute(
        "SELECT setting_value FROM app_settings WHERE setting_name = 'workflow_auto_notify'").fetchone()
    lock_notes = conn.execute(
        "SELECT setting_value FROM app_settings WHERE setting_name = 'workflow_lock_notes'").fetchone()
    conn.close()

    settings = {
        'auto_notify': auto_notify['setting_value'] == 'true' if auto_notify else True,
        'lock_notes': lock_notes['setting_value'] == 'true' if lock_notes else True
    }

    return render_template('admin_workflow.html', settings=settings)


@app.route('/admin/settings')
@login_required
def admin_settings():
    # Access control: All roles can access for Node Config; Admin checks handled in template
    # if session.get('role') != 'Admin':
    #     return redirect(url_for('dashboard'))

    conn = get_db_connection()
    # Fetch all settings
    settings_rows = conn.execute(
        "SELECT setting_name, setting_value FROM app_settings").fetchall()
    conn.close()

    # Convert list of rows to dictionary
    settings = {row['setting_name']: row['setting_value']
                for row in settings_rows}

    return render_template('admin_settings.html', settings=settings)


@app.route('/admin/settings/update', methods=['POST'])
@login_required
def admin_update_settings():
    if session.get('role') != 'Admin':
        flash("Unauthorized", "error")
        return redirect(url_for('dashboard'))

    try:
        conn = get_db_connection()

        # Handle Logo File Upload from native device file picker
        logo_url_val = request.form.get('logo_url')
        if 'logo_file' in request.files:
            logo_file = request.files['logo_file']
            if logo_file and logo_file.filename != '':
                import base64
                file_bytes = logo_file.read()
                if len(file_bytes) > 0:
                    mime_type = logo_file.content_type or 'image/png'
                    encoded = base64.b64encode(file_bytes).decode('utf-8')
                    logo_url_val = f"data:{mime_type};base64,{encoded}"

        # Dictionary of settings to update
        setting_updates = {
            'system_name': request.form.get('system_name'),
            'logo_url': logo_url_val,
        }

        selected_theme = request.form.get('theme_color') or request.form.get('active_theme')
        if selected_theme and str(selected_theme).strip() != '':
            setting_updates['active_theme'] = selected_theme
            setting_updates['theme_color'] = selected_theme

        sys_id = str(uuid.uuid4())
        for key, val in setting_updates.items():
            if val is not None and str(val).strip() != '':
                conn.execute("""
                    INSERT INTO app_settings (setting_name, setting_value, global_id, updated_at) 
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP) 
                    ON CONFLICT(setting_name) DO UPDATE SET 
                        setting_value=excluded.setting_value, 
                        updated_at=CURRENT_TIMESTAMP
                """, (key, val, sys_id))

        conn.commit()
        conn.close()

        # Immediate sync push to cloud
        try:
            from sync_engine import sync_manager
            sync_manager.run_automated_sync()
        except Exception as sync_err:
            print(f"[SETTINGS] Sync trigger warning: {sync_err}")

        flash("System configuration and branding updated successfully.", "success")
    except Exception as e:
        flash(f"Error saving settings: {e}", "error")

    return redirect(url_for('admin_settings'))


@app.route('/admin/workflow/save', methods=['POST'])
@login_required
def save_workflow_settings():
    if session.get('role') != 'Admin':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()
        conn = get_db_connection()

        # Upsert logic (delete then insert is easier for simple KV)
        conn.execute(
            "DELETE FROM app_settings WHERE setting_name IN ('workflow_auto_notify', 'workflow_lock_notes')")

        conn.execute("INSERT INTO app_settings (setting_name, setting_value) VALUES (?, ?)",
                     ('workflow_auto_notify', 'true' if data.get('auto_notify') else 'false'))
        conn.execute("INSERT INTO app_settings (setting_name, setting_value) VALUES (?, ?)",
                     ('workflow_lock_notes', 'true' if data.get('lock_notes') else 'false'))

        conn.commit()
        trigger_sync_immediate()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/forms')
@login_required
def admin_forms():
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))
    return render_template('admin_forms.html')


@app.route('/admin/export/master')
@login_required
def admin_export_master():
    if session.get('role') != 'Admin':
        return redirect(url_for('dashboard'))

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill
        import io
    except ImportError:
        flash("Export library missing. Please contact support.", "error")
        return redirect(url_for('dashboard'))

    conn = get_db_connection()

    # 1. Fetch Datasets
    students = conn.execute("SELECT * FROM Student").fetchall()
    appointments = conn.execute("SELECT * FROM Appointment").fetchall()
    intake_forms = conn.execute("SELECT * FROM intake_forms").fetchall()
    users = conn.execute(
        "SELECT id, username, full_name, role, last_login, created_at FROM users").fetchall()

    conn.close()

    # 2. Create Workbook
    wb = openpyxl.Workbook()

    # Helper to write sheet
    def write_sheet(wb, sheet_name, data, columns=None):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        if not data:
            ws.append(["No Data Available"])
            return

        # Headers
        if not columns:
            columns = data[0].keys()

        # Style headers
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill(
            start_color="FF4F81BD", end_color="FF4F81BD", fill_type="solid")

        for col_num, col_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=str(col_title).upper())
            cell.font = header_font
            cell.fill = header_fill

        # Data
        for row_data in data:
            row_values = [row_data[col] for col in columns]
            ws.append(row_values)

        # Autosize columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

    # 3. Populate Sheets
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    write_sheet(wb, "Students", students)
    write_sheet(wb, "Appointments", appointments)
    write_sheet(wb, "Intake Records", intake_forms)
    write_sheet(wb, "System Users", users)

    # 4. Return File
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers[
        'Content-Disposition'] = f'attachment; filename=Master_Data_Export_{timestamp}.xlsx'
    return response


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.')
    return redirect(url_for('welcome'))


@app.route('/add_student', methods=['GET', 'POST'])
@login_required
def add_student():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('students'))

        if request.method == 'POST':
            edit_id = request.form.get('edit_id')

            # ── GTEC: Preserving full names as requested ──────────────
            raw_name_input = request.form.get('name', '').strip()
            name = name_to_initials(raw_name_input)
            # ────────────────────────────────────────────────────────────────

            age = request.form.get('age')
            gender = request.form.get('gender')
            index_number = request.form.get('index_number')
            department = request.form.get('department')
            programme_base = request.form.get('programme')
            programme_other = request.form.get('programme_other')
            programme = programme_other if programme_base == 'Other' else programme_base

            # ── Encrypt sensitive contact fields before storage ─────────────
            contact       = encrypt_field(request.form.get('contact'))
            parent_contact = encrypt_field(request.form.get('parent_contact'))
            # ────────────────────────────────────────────────────────────────
            hall_of_residence = request.form.get('hall_of_residence')
            faculty = request.form.get('faculty', '')

            try:
                if edit_id:
                    conn.execute('''
                        UPDATE Student 
                        SET name=?, age=?, gender=?, index_number=?, department=?, faculty=?, 
                            programme=?, contact=?, parent_contact=?, hall_of_residence=?
                        WHERE id=?
                    ''', (name, age if age else None, gender, index_number, department, faculty,
                          programme, contact, parent_contact, hall_of_residence, edit_id))
                    flash('Client record updated successfully!', 'success')
                else:
                    case_num = generate_case_number(conn, name)
                    conn.execute(
                        'INSERT INTO Student (name, case_number, age, gender, index_number, department, faculty, programme, contact, parent_contact, hall_of_residence) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                        (name, case_num, age if age else None, gender, index_number, department,
                         faculty, programme, contact, parent_contact, hall_of_residence)
                    )
                    flash(
                        f'Client registered! Case Number: {case_num}', 'success')
                conn.commit()
                trigger_sync_immediate()
                return redirect(url_for('students'))
            except sqlite3.IntegrityError:
                conn.rollback()
                flash('Error: Index number already exists.', 'error')
                if edit_id:
                    try:
                        student = conn.execute(
                            'SELECT * FROM Student WHERE id = ?', (edit_id,)).fetchone()
                        return render_template('add_student.html', student=student)
                    except Exception:
                        pass
            except Exception as e:
                conn.rollback()
                print(f"[ADD_STUDENT] Error saving student: {e}")
                import traceback
                traceback.print_exc()
                # Check if it's a table missing error and reinitialize
                if 'no such table' in str(e).lower() or 'Student' in str(e):
                    print(
                        "[ADD_STUDENT] Table missing error detected, reinitializing database...")
                    try:
                        ensure_database_initialized()
                        conn = get_db_connection()
                        flash('Database was reinitialized. Please try again.', 'info')
                    except Exception as init_error:
                        print(
                            f"[ADD_STUDENT] Error reinitializing: {init_error}")
                        flash(
                            f'Database error: {str(e)}. Please restart the application.', 'error')
                else:
                    flash(f'Error saving student: {str(e)}', 'error')
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        else:
            student = None
            edit_id = request.args.get('edit')
            if edit_id:
                try:
                    raw = conn.execute(
                        'SELECT * FROM Student WHERE id = ?', (edit_id,)).fetchone()
                    if raw:
                        # ── Decrypt sensitive fields for display in edit form ──
                        student = dict(raw)
                        student['contact']        = decrypt_field(student.get('contact'))
                        student['parent_contact']  = decrypt_field(student.get('parent_contact'))
                        # Convert Row back to a simple namespace for template access
                        from types import SimpleNamespace
                        student = SimpleNamespace(**student)
                    else:
                        student = None
                except Exception as e:
                    print(f"[ADD_STUDENT] Error loading student for edit: {e}")
                    import traceback
                    traceback.print_exc()
                    if 'no such table' in str(e).lower() or 'Student' in str(e):
                        print(
                            "[ADD_STUDENT] Table missing error detected in GET, reinitializing database...")
                        try:
                            ensure_database_initialized()
                            conn = get_db_connection()
                            raw = conn.execute(
                                'SELECT * FROM Student WHERE id = ?', (edit_id,)).fetchone()
                            if raw:
                                student = dict(raw)
                                student['contact']       = decrypt_field(student.get('contact'))
                                student['parent_contact'] = decrypt_field(student.get('parent_contact'))
                                from types import SimpleNamespace
                                student = SimpleNamespace(**student)
                        except Exception as init_error:
                            print(
                                f"[ADD_STUDENT] Error reinitializing: {init_error}")
                            student = None
                    else:
                        student = None
            try:
                conn.close()
            except Exception:
                pass
            return render_template('add_student.html', student=student, user_name=session.get('full_name'))

    except Exception as e:
        print(f"[ADD_STUDENT] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading student form. Please try again.', 'error')
        return redirect(url_for('students'))


@app.route('/sessions')
@login_required
def sessions_list():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        sessions_raw = []
        try:
            # Get sessions with full details including student, counsellor, and appointment info
            sessions_raw = conn.execute('''
                SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
                       s.id as student_db_id, s.name as student_name, s.case_number, s.created_at as student_created_at,
                       c.name as Counsellor_name,
                       a.date, a.time, a.status,
                       sess.appointment_id
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                ORDER BY sess.created_at DESC
            ''').fetchall()
        except Exception as e:
            print(f"[SESSIONS] Error getting sessions: {e}")
            sessions_raw = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Convert to list and add professional IDs
        sessions = []
        for sess in sessions_raw:
            sess_dict = dict(sess)
            
            # GTEC REQUIREMENT: Always mask name for display
            original_name = sess_dict.get('student_name')
            clinical_id = get_clinical_id(original_name, sess_dict.get('student_record_id'), sess_dict.get('student_created_at'))
            sess_dict['student_name'] = name_to_initials(original_name)
            sess_dict['student_clinical_id'] = clinical_id
            sess_dict['professional_id'] = sess_dict.get('case_number') or clinical_id
            # Clean display dates/times
            sess_dict['date'] = clean_date_string(sess_dict.get('date'))
            sess_dict['time'] = clean_time_string(sess_dict.get('time'))
            
            sessions.append(sess_dict)

        # Create a simple pagination object to prevent template errors
        class SimplePagination:
            has_prev = False
            has_next = False
            page = 1
            prev_num = None
            next_num = None

            def iter_pages(self):
                return []

        pagination = SimplePagination()
        return render_template('sessions.html', sessions=sessions, pagination=pagination)
    except Exception as e:
        print(f"[SESSIONS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading sessions. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/create_session', methods=['GET', 'POST'])
@login_required
def create_session():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            appointment_id = request.form.get('appointment_id')
            student_id = request.form.get('student_id')
            session_type = request.form.get('session_type')
            notes = request.form.get('notes')
            outcome = request.form.get('outcome', '')

            try:
                # If no appointment selected but student is selected, create a walk-in appointment
                if not appointment_id and student_id:
                    print(f"[CREATE_SESSION] Creating walk-in appointment for student {student_id}")
                    # Use the first counsellor as default for walk-ins
                    counsellor = conn.execute("SELECT id FROM Counsellor LIMIT 1").fetchone()
                    counsellor_id = counsellor['id'] if counsellor else 1
                    
                    cursor = conn.execute('''
                        INSERT INTO Appointment (student_id, Counsellor_id, date, time, purpose, status, urgency, referral_source, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (student_id, counsellor_id, datetime.now().strftime('%Y-%m-%d'), 
                          datetime.now().strftime('%H:%M'), 'Walk-in Session', 'Completed', 'Normal', 'None',
                          datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                    appointment_id = cursor.lastrowid
                    print(f"[CREATE_SESSION] Walk-in appointment created: ID {appointment_id}")

                if appointment_id:
                    # Get appointment details
                    appointment = conn.execute('''
                        SELECT student_id, status 
                        FROM Appointment 
                        WHERE id = ?
                    ''', (appointment_id,)).fetchone()

                    if appointment:
                        student_id = appointment['student_id']
                        appointment_status = appointment['status']

                        # Insert new session
                        import uuid as _uuid
                        sess_gid = str(_uuid.uuid4())
                        conn.execute('''
                            INSERT INTO session (appointment_id, session_type, notes, outcome, created_at, global_id)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', (appointment_id, session_type, notes, outcome, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sess_gid))

                        # Update appointment status
                        if appointment_status.lower() in ['scheduled', 'sent to counsellor']:
                            conn.execute('UPDATE Appointment SET status = ? WHERE id = ?',
                                         ('Completed', appointment_id))

                        conn.commit()
                        trigger_sync_immediate()
                        flash('Session created successfully!', 'success')

                        if request.form.get('schedule_followup'):
                            return redirect(url_for('appointment', student_id=student_id))

                        return redirect(url_for('sessions_list'))
                    else:
                        flash('Invalid appointment selected.', 'error')
                        return redirect(url_for('create_session'))
                else:
                    flash('Please select either an appointment or a student for a walk-in.', 'error')
                    return redirect(url_for('create_session'))
            except Exception as e:
                conn.rollback()
                print(f"[CREATE_SESSION] Error creating session: {e}")
                flash(f'Error creating session: {str(e)}', 'error')
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
            return redirect(url_for('create_session'))

        # GET: load appointments and students
        appointments = []
        students = []
        try:
            appointments = conn.execute('''
                SELECT a.id, a.date as date, a.time as time, a.status, s.name as student_name,
                       c.name as counsellor_name, a.urgency, a.referral_source, a.purpose
                FROM Appointment a
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE a.status IN ('scheduled', 'Scheduled', 'In Session', 'Completed', 'completed', 'Sent to Counsellor')
                ORDER BY a.date DESC, a.time DESC
            ''').fetchall()
            # Clean and format dates/times for the template
            appointments = [dict(a) for a in appointments]
            for a in appointments:
                a['date'] = clean_date_string(a['date'])
                a['time'] = clean_time_string(a['time'])
                
            students = conn.execute('SELECT id, name, index_number, case_number FROM Student ORDER BY name').fetchall()
        except Exception as e:
            print(f"[CREATE_SESSION] Error getting data: {e}")
            appointments = []
            students = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Check if specific appointment_id is passed in args (from Dashboard or Queue)
        selected_appt_id = request.args.get('appointment_id')
        selected_student_id = request.args.get('student_id')
        return render_template('create_session.html', 
                             appointments=appointments, 
                             students=students,
                             selected_appt_id=selected_appt_id,
                             selected_student_id=selected_student_id)
    except Exception as e:
        print(f"[CREATE_SESSION] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading session creation page. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/case_note', methods=['GET', 'POST'])
@login_required
def case_note():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            session_id = request.form.get('session_id')
            client_appearance = request.form.get('client_appearance', '')
            problems = request.form.get('problems', '')
            interventions = request.form.get('interventions', '')
            recommendations = request.form.get('recommendations', '')
            next_visit_date = request.form.get('next_visit_date') or None
            counsellor_signature = request.form.get('counsellor_signature', '')

            # Validate required fields
            if not session_id:
                flash('Please select a session', 'error')
                return redirect(url_for('case_note'))

            if not all([client_appearance, problems, interventions, recommendations]):
                flash('Please fill in all required fields', 'error')
                return redirect(url_for('case_note'))

            try:
                # Check if case management record already exists for this session
                existing = conn.execute(
                    'SELECT id FROM CaseManagement WHERE session_id = ?', (session_id,)).fetchone()

                if existing:
                    # Update existing record
                    conn.execute('''
                        UPDATE CaseManagement 
                        SET client_appearance = ?, problems = ?, interventions = ?, recommendations = ?,
                            next_visit_date = ?, counsellor_signature = ?
                        WHERE session_id = ?
                    ''', (client_appearance, problems, interventions, recommendations,
                          next_visit_date, counsellor_signature, session_id))
                else:
                    # Insert new record
                    import uuid as _uuid
                    cm_gid = str(_uuid.uuid4())
                    conn.execute('''
                        INSERT INTO CaseManagement 
                        (session_id, client_appearance, problems, interventions, recommendations, 
                         next_visit_date, counsellor_signature, created_at, global_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (session_id, client_appearance, problems, interventions, recommendations,
                          next_visit_date, counsellor_signature, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), cm_gid))

                conn.commit()
                trigger_sync_immediate()
                flash('Case notes saved successfully!', 'success')
                return redirect(url_for('sessions_list'))
            except Exception as e:
                conn.rollback()
                print(f"[CASE_NOTE] Error saving case notes: {e}")
                import traceback
                traceback.print_exc()
                flash(f'Error saving case notes: {str(e)}', 'error')
                return redirect(url_for('case_note'))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - show form
        sessions = []
        try:
            sessions_raw = conn.execute('''
                SELECT s.id, st.id as student_db_id, st.name as student_name, 
                       st.case_number, s.created_at, a.purpose as session_topic, st.global_id
                FROM session s
                JOIN Appointment a ON s.appointment_id = a.id
                JOIN Student st ON a.student_id = st.id
                ORDER BY s.created_at DESC
            ''').fetchall()

            sessions = []
            for sess in sessions_raw:
                sess_dict = dict(sess)
                s_date = (sess_dict.get('created_at') or '').split(' ')[0]
                full_name = sess_dict.get('student_name', 'N/A')
                # Only use FIRST name as requested to avoid clutter
                first_name = full_name.split(' ')[0]
                
                # Use stored case_number (GCC format)
                s_case = sess_dict.get('case_number') or f"GCC-{datetime.now().year}-{sess_dict.get('student_db_id',0):04d}"
                
                # Simplified format: [Date] — [First Name] | [Case ID] 
                sess_dict['display_name'] = f"{s_date} — {first_name} | {s_case}"
                sessions.append(sess_dict)
        except Exception as e:
            print(f"[CASE_NOTE] Error getting sessions: {e}")
            sessions = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('case_note.html', sessions=sessions, now=datetime.utcnow())
    except Exception as e:
        print(f"[CASE_NOTE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case notes page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

# ---------- Reports Routes ----------


@app.route('/reports')
@login_required
def reports_list():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        reports = []
        try:
            # Get all generated reports from the database
            reports = conn.execute('''
                SELECT id, title, date_generated, report_type, summary, file_path
                FROM reports
                ORDER BY date_generated DESC
            ''').fetchall()
        except Exception as e:
            print(f"[REPORTS] Error getting reports: {e}")
            reports = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('reports.html', reports=reports)
    except Exception as e:
        print(f"[REPORTS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading reports. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/generate_report_manual', methods=['POST'], endpoint='generate_report_manual')
@login_required
def generate_report_manual():
    """Generate a report manually with custom options"""
    report_type = request.form.get('report_type', 'manual')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    try:
        if report_type == 'custom' and start_date and end_date:
            # Custom date range - modify the generate_report function to accept dates
            manual_generate_report()  # For now, use manual generation
            flash(
                'Custom report generation is being prepared. Report generated successfully!', 'success')
        else:
            manual_generate_report()
            flash('Report generated successfully!', 'success')
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')

    return redirect(url_for('reports_list'))


@app.route('/toggle_auto_report', methods=['GET', 'POST'])
@login_required
def toggle_auto_report():
    """Toggle auto report generation on/off"""
    if request.method == 'POST':
        data = request.get_json()
        enable = data.get('enable', False)

        try:
            toggle_scheduler(enable)
            return jsonify({
                'status': 'success',
                'message': f'Auto report generation {"enabled" if enable else "disabled"} successfully.',
                'is_enabled': enable
            })
        except Exception as e:
            return jsonify({
                'status': 'error',
                'message': str(e)
            }), 500
    else:
        # GET request - return current status
        is_running = scheduler.running if scheduler else False
        return jsonify({
            'status': 'success',
            'is_enabled': is_running
        })


@app.route('/generate_report_now', methods=['POST'])
@login_required
def generate_report_now():
    """Manually trigger report generation"""
    try:
        manual_generate_report()
        return jsonify({
            'status': 'success',
            'message': 'Report generated successfully! Check the Reports page to view it.'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error generating report: {str(e)}'
        }), 500


@app.route('/my_cases')
@login_required
def my_cases():
    user_role = (session.get('role') or '').strip().lower()
    if user_role not in ['counsellor', 'counselor', 'admin']:
        flash(f"Access restricted to clinical staff (Current Role: {user_role.capitalize() or 'Unknown'}).", "error")
        return redirect(url_for('dashboard'))

    try:
        conn = get_db_connection()
        user_full_name = session.get('full_name') or session.get('username', '')
        username = session.get('username', '')

        # 1. Find Counsellor record — try full_name first, then username
        counsellor = conn.execute(
            "SELECT id FROM Counsellor WHERE name = ? OR name = ?",
            (user_full_name, username)).fetchone()

        if not counsellor:
            if user_role == 'admin':
                return redirect(url_for('students'))

            # Auto-heal: create Counsellor profile from session data
            if user_role in ['counsellor', 'counselor']:
                try:
                    conn.execute(
                        "INSERT INTO Counsellor (name, contact) VALUES (?, '')", (user_full_name,))
                    conn.commit()
                    trigger_sync_immediate()
                    counsellor = conn.execute(
                        "SELECT id FROM Counsellor WHERE name = ?", (user_full_name,)).fetchone()
                except Exception as e:
                    print(f"[MY_CASES] Auto-create counsellor record failed: {e}")

            if not counsellor:
                # Last resort — redirect to student registry (safer than rendering raw)
                flash(f"Professional profile not fully registered. Viewing all clients.", "info")
                return redirect(url_for('students'))

        counsellor_id = counsellor['id']

        # 2. Find students who have appointments with this counsellor (Past or Future)
        students_raw = conn.execute('''
            SELECT s.*, 
                   COUNT(DISTINCT sess.id) as session_count
            FROM Student s
            JOIN Appointment a ON s.id = a.student_id
            LEFT JOIN session sess ON a.id = sess.appointment_id
            WHERE a.Counsellor_id = ?
            GROUP BY s.id
            ORDER BY MAX(a.date) DESC
        ''', (counsellor_id,)).fetchall()

        conn.close()

        students = []
        for student in students_raw:
            student_dict = dict(student)
            student_db_id = student_dict.get('id', 0)
            
            # Generate professional_id
            s_name = student_dict.get('name', 'GCC')
            parts = [p.strip() for p in s_name.split() if p.strip()]
            if len(parts) >= 2:
                initials = (parts[0][0] + parts[-1][0]).upper()
            elif len(parts) == 1:
                initials = parts[0][:2].upper()
            else:
                initials = "GC"
                
            # Use case number as first priority for ID display
            if not student_dict.get('case_number'):
                student_dict['case_number'] = f"GCC-{datetime.now().year}-{student_db_id:04d}"
            student_dict['professional_id'] = student_dict['case_number']
                
            students.append(student_dict)

        return render_template('students.html', students=students, programs=[], page_title="My Cases")

    except Exception as e:
        print(f"[MY_CASES] Error: {e}")
        flash(f"Error accessing cases: {str(e)}", "error")
        return redirect(url_for('dashboard'))


@app.route('/download_report_file/<int:report_id>')
@login_required
def download_report_file(report_id):
    """Download the actual report file (DOCX)"""
    conn = get_db_connection()

    report = conn.execute(
        'SELECT * FROM reports WHERE id = ?', (report_id,)).fetchone()
    conn.close()

    if not report:
        flash('Report not found', 'error')
        return redirect(url_for('reports_list'))

    # Convert Row to dict for easier access
    report_dict = dict(report)
    file_path = report_dict.get('file_path')

    if not file_path:
        flash('Report file path not found', 'error')
        return redirect(url_for('reports_list'))

    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
    else:
        flash('Report file not found on disk', 'error')
        return redirect(url_for('reports_list'))


@app.route('/students')
@login_required
def students():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        students_raw = []
        program_rows = []
        try:
            # Get all students and their information with session counts
            students_raw = conn.execute('''
                SELECT s.id, s.name, s.index_number, s.gender, s.age, s.faculty, s.department, s.programme,
                       s.hall_of_residence, s.contact, s.email, s.parent_contact, s.created_at,
                       s.case_number, s.global_id, s.updated_at, s.last_synced_at,
                       COUNT(DISTINCT sess.id) as session_count
                FROM Student s
                LEFT JOIN Appointment a ON s.id = a.student_id
                LEFT JOIN session sess ON a.id = sess.appointment_id
                WHERE s.is_deleted = 0
                GROUP BY s.id
                ORDER BY s.id DESC
            ''').fetchall()

            # Get unique programs for the filter
            program_rows = conn.execute(
                "SELECT DISTINCT programme FROM Student WHERE programme IS NOT NULL AND is_deleted = 0").fetchall()
        except Exception as e:
            print(f"[STUDENTS] Error getting students: {e}")
            students_raw = []
            program_rows = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Map to specific identifiers as per technical spec
        students = []
        for student in students_raw:
            student_dict = dict(student)
            
            # GTEC REQUIRED: Standardize identifiers while preserving full name
            clinical_id = get_clinical_id(student_dict.get('name'), student_dict.get('id'), student_dict.get('created_at'))
            student_dict['clinical_id'] = clinical_id
            student_dict['professional_id'] = clinical_id
            # Ensure name is full name (not masked)
            student_dict['name'] = name_to_initials(student_dict.get('name', 'User'))
            
            # 1. Case ID (GCC-2026-####)
            student_dict['case_id'] = student_dict.get('case_number') or f"GCC-{datetime.now().year}-{student_dict.get('id', 0):04d}"

            # Decrypt sensitive fields
            for field in STUDENT_SENSITIVE_FIELDS:
                if field in student_dict and student_dict[field]:
                    student_dict[field] = decrypt_field(student_dict[field])

            students.append(student_dict)

        # Convert Row objects to strings
        programs = [row['programme']
                    for row in program_rows] if program_rows else []

        return render_template('students.html', students=students, programs=programs)
    except Exception as e:
        print(f"[STUDENTS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading students. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/import_students', methods=['POST'])
@login_required
def import_students():
    user_role = (session.get('role') or '').strip().lower()
    if user_role not in ['admin', 'counsellor', 'counselor', 'secretary']:
        flash("You do not have permission to import data.", "error")
        return redirect(url_for('admin_settings'))

    if 'file' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('admin_settings'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('admin_settings'))

    if file:
        filename = secure_filename(file.filename)
        extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        try:
            conn = get_db_connection()
            import_count = 0
            
            if extension == 'csv':
                stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
                csv_input = csv.DictReader(stream)
                for row in csv_input:
                    # Robust header helper for DictReader
                    def get_csv_val(possible_keys):
                        if isinstance(possible_keys, str):
                            possible_keys = [possible_keys]
                        for pk in possible_keys:
                            # Try exact, lowercase, and title case
                            for variant in [pk, pk.lower(), pk.capitalize(), pk.upper(), pk.title()]:
                                val = row.get(variant)
                                if val is not None:
                                    return str(val).strip()
                        return ''

                    name = get_csv_val(['Name', 'Student Name', 'Full Name', 'Client Name', 'student_name'])
                    index_number = get_csv_val(['Index Number', 'Index_Number', 'Index No', 'IndexNo', 'Student ID', 'ID'])
                    department = get_csv_val(['Department', 'Dept', 'Department Name'])
                    programme = get_csv_val(['Programme', 'Program', 'Course', 'Study Programme'])
                    
                    if not name or not index_number:
                        continue
                        
                    case_number = generate_case_number(conn, name)
                    
                    conn.execute('''
                        INSERT INTO Student (name, case_number, age, gender, contact, index_number, department, faculty, programme)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        name,
                        case_number,
                        get_csv_val(['Age', 'Student Age']),
                        get_csv_val(['Gender', 'Sex']),
                        get_csv_val(['Contact', 'Phone', 'Phone Number', 'Mobile']),
                        index_number,
                        department,
                        get_csv_val(['Faculty', 'School']),
                        programme
                    ))
                    import_count += 1
                    
            elif extension in ['xlsx', 'xls']:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                # Assume first row is header
                headers = [cell.value for cell in sheet[1]]
                # Map headers to column indices
                header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    def get_val(possible_keys):
                        if isinstance(possible_keys, str):
                            possible_keys = [possible_keys]
                        for pk in possible_keys:
                            idx = header_map.get(pk.lower())
                            if idx is not None and idx < len(row) and row[idx] is not None:
                                return str(row[idx]).strip()
                        return ''

                    name = get_val(['name', 'student name', 'full name', 'client name', 'student_name'])
                    index_number = get_val(['index number', 'index_number', 'index no', 'indexno', 'student id', 'id'])
                    department = get_val(['department', 'dept', 'department name'])
                    programme = get_val(['programme', 'program', 'course', 'study programme'])
                    
                    if not name or not index_number:
                        continue
                        
                    case_number = generate_case_number(conn, name)
                    
                    conn.execute('''
                        INSERT INTO Student (name, case_number, age, gender, contact, index_number, department, faculty, programme)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        name,
                        case_number,
                        get_val(['age', 'student age']),
                        get_val(['gender', 'sex']),
                        get_val(['contact', 'phone', 'phone number', 'mobile']),
                        index_number,
                        department,
                        get_val(['faculty', 'school']),
                        programme
                    ))
                    import_count += 1
            else:
                flash('Unsupported file format. Please use CSV or Excel (.xlsx).', 'error')
                return redirect(url_for('admin_settings'))

            conn.commit()
            trigger_sync_immediate()
            conn.close()
            flash(f'Successfully imported {import_count} student records.', 'success')
            
        except Exception as e:
            flash(f'Import failed: {str(e)}', 'error')
            print(f"[IMPORT] Error: {e}")
            
        return redirect(url_for('admin_settings'))


@app.route('/student_profile/<int:id>')
@login_required
def student_profile(id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        # Get student information
        student = None
        sessions = []
        referrals = []
        try:
            student = conn.execute(
                'SELECT * FROM Student WHERE id = ?', (id,)).fetchone()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting student: {e}")

        if not student:
            try:
                conn.close()
            except Exception:
                pass
            flash('Student not found', 'error')
            return redirect(url_for('students'))

        # Standardize name for GTEC privacy at runtime (Detect existing initials)
        student = dict(student)
        if "(" not in student.get('name', ''):
            student['name'] = name_to_initials(student['name'])

        # Decrypt sensitive fields
        for field in STUDENT_SENSITIVE_FIELDS:
            if field in student and student[field]:
                student[field] = decrypt_field(student[field])

        # Get all sessions for this student
        sessions_raw = []
        try:
            sessions_raw = conn.execute('''
                SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
                       s.name as student_name,
                       c.name as Counsellor_name,
                       a.date, a.time, a.status
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE s.id = ?
                ORDER BY sess.created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting sessions: {e}")
            sessions_raw = []

        # Convert sessions to list and initialize student names
        sessions = []
        for s in sessions_raw:
            s_dict = dict(s)
            s_dict['student_name'] = name_to_initials(s_dict.get('student_name', 'Client'))
            sessions.append(s_dict)

        # Get all referrals for this student
        try:
            referrals = conn.execute('''
                SELECT r.id, r.referred_by, r.contact, r.reasons, r.action_taken, r.outcome, r.created_at
                FROM Referral r
                JOIN Session sess ON r.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                WHERE a.student_id = ?
                ORDER BY r.created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting referrals: {e}")
            referrals = []

        # Get DASS-21 scores if table exists
        dass21_scores = []
        try:
            dass21_scores = conn.execute('''
                SELECT depression_score, anxiety_score, stress_score, completion_date, created_at
                FROM DASS21
                WHERE student_id = ?
                ORDER BY created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting DASS21 scores: {e}")
            dass21_scores = []

        # Get OQ-45.2 scores if table exists
        oq_scores = []
        try:
            oq_scores = conn.execute('''
                SELECT total_score, completion_date, created_at,
                       (SELECT created_at FROM session WHERE id = OutcomeQuestionnaire.session_id) as session_date
                FROM OutcomeQuestionnaire
                WHERE student_id = ?
                ORDER BY created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting OQ scores: {e}")
            oq_scores = []

        try:
            conn.close()
        except Exception:
            pass

        return render_template('student_profile.html',
                               student=student,
                               sessions=sessions,
                               referrals=referrals,
                               dass21_scores=dass21_scores,
                               oq_scores=oq_scores)
    except Exception as e:
        print(f"[STUDENT_PROFILE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading student profile. Please try again.', 'error')
        return redirect(url_for('students'))


@app.route('/export_students')
@login_required
def export_students():
    import csv
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Get format parameter (default to csv for backward compatibility)
    export_format = request.args.get('format', 'csv').lower()

    conn = get_db_connection()

    # Get all students data with professional ID
    students_raw = conn.execute('''
        SELECT s.*, 
               COUNT(DISTINCT sess.id) as session_count
        FROM Student s
        LEFT JOIN Appointment a ON s.id = a.student_id
        LEFT JOIN Session sess ON a.id = sess.appointment_id
        GROUP BY s.id
        ORDER BY s.name
    ''').fetchall()

    conn.close()

    # Generate professional IDs (same logic as the students page)
    students = []
    for student in students_raw:
        student_dict = dict(student)
        student_db_id = student_dict.get('id', 0)
        s_name = student_dict.get('name', 'GCC')
        parts = [p.strip() for p in s_name.split() if p.strip()]
        if len(parts) >= 2:
            initials = (parts[0][0] + parts[-1][0]).upper()
        elif len(parts) == 1:
            initials = parts[0][:2].upper()
        else:
            initials = "C"
        initials = "".join([c for c in initials if c.isalpha()])
        if not initials:
            initials = "C"
        # Standardize professional ID to use the new case_number format
        student_dict['professional_id'] = student_dict.get('case_number') or f"GCC-{datetime.now().year}-{student_db_id:04d}"
        students.append(student_dict)

    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Define header style
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)

        # Write headers (matching the app's column labels)
        headers = ['Case Number', 'Client ID', 'Name', 'Index Number', 'Age', 'Gender',
                   'Email', 'Phone', 'Program', 'Department', 'Sessions', 'Registered Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=student.get('case_number') or 'N/A')
            ws.cell(row=row_num, column=2, value=student['professional_id'])
            ws.cell(row=row_num, column=3, value=student['name'])
            ws.cell(row=row_num, column=4,
                    value=student['index_number'] or 'N/A')
            ws.cell(row=row_num, column=5, value=student.get('age') or 'N/A')
            ws.cell(row=row_num, column=6,
                    value=student.get('gender') or 'N/A')
            ws.cell(row=row_num, column=7, value=student.get('email') or 'N/A')
            ws.cell(row=row_num, column=8, value=student['contact'] or 'N/A')
            ws.cell(row=row_num, column=9, value=student.get(
                'programme') or student.get('program') or 'N/A')
            ws.cell(row=row_num, column=10,
                    value=student.get('department') or 'N/A')
            ws.cell(row=row_num, column=11, value=student['session_count'])
            ws.cell(row=row_num, column=12, value=student['created_at'])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=students_export.xlsx'
        return response
    else:
        # Create CSV (default)
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header (matching the app's column labels)
        writer.writerow(['Case Number', 'Client ID', 'Name', 'Index Number',
                        'Email', 'Phone', 'Program', 'Sessions', 'Registered Date'])

        # Write data rows
        for student in students:
            writer.writerow([
                student.get('case_number') or 'N/A',
                student['professional_id'],
                student['name'],
                student['index_number'] or 'N/A',
                student.get('email') or 'N/A',
                student['contact'] or 'N/A',
                student.get('programme') or student.get('program') or 'N/A',
                student['session_count'],
                student['created_at']
            ])

        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=students_export.csv'

        return response


@app.route('/api/check_appointment/<int:student_id>')
@login_required
def check_appointment(student_id):
    """
    Checks for any pending or scheduled appointments for a student.
    Handles legacy, ISO, and Supabase timestamp string formats flawlessly.
    """
    try:
        conn = get_db_connection()
        today_iso = datetime.now().strftime('%Y-%m-%d')
        
        # Check for any scheduled or in-progress appointments
        appt_rows = conn.execute('''
            SELECT id, date, time FROM Appointment 
            WHERE student_id = ? 
            AND (status IN ('scheduled', 'Scheduled', 'Sent to Counsellor', 'In Session'))
        ''', (student_id,)).fetchall()
        conn.close()
        
        if not appt_rows:
            return jsonify({'status': 'not_found'})

        best_appt = None
        best_appt_date = None
        is_today = False
        
        # Sort and pick: preferring Today first, then the nearest future
        processed_appts = []
        for row in appt_rows:
            clean_d = clean_date_string(row['date'])
            processed_appts.append({
                'row': row,
                'clean_date': clean_d,
                'is_today': (clean_d == today_iso)
            })
            
        # Priority 1: Today
        today_match = next((a for a in processed_appts if a['is_today']), None)
        if today_match:
            best_appt = today_match['row']
            is_today = True
            best_appt_date = today_match['clean_date']
        else:
            # Priority 2: Future (Nearest)
            future_appts = [a for a in processed_appts if a['clean_date'] > today_iso]
            future_appts.sort(key=lambda x: x['clean_date'])
            if future_appts:
                best_appt = future_appts[0]['row']
                best_appt_date = future_appts[0]['clean_date']

        if best_appt:
            # Format nicely for the UI
            d_obj = datetime.strptime(best_appt_date, '%Y-%m-%d')
            fmt_date = d_obj.strftime('%a, %b %d, %Y')
            fmt_time = clean_time_string(best_appt['time'])
            
            # Check if it's overdue (only for today's appointments)
            is_overdue = False
            if is_today:
                try:
                    # Parse simplified time for comparison
                    raw_t = str(best_appt['time']).strip().upper()
                    if 'PM' in raw_t or 'AM' in raw_t:
                        t_parts = raw_t.replace('AM', '').replace('PM', '').strip().split(':')
                        h = int(t_parts[0])
                        m = int(t_parts[1]) if len(t_parts) > 1 else 0
                        if 'PM' in raw_t and h < 12: h += 12
                        if 'AM' in raw_t and h == 12: h = 0
                    else:
                        t_parts = raw_t.split(':')
                        h = int(t_parts[0])
                        m = int(t_parts[1]) if len(t_parts) > 1 else 0
                    
                    appt_time_today = datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
                    # Overdue if it's been more than 30 minutes past the scheduled time
                    if (datetime.now() - appt_time_today).total_seconds() > 1800:
                        is_overdue = True
                except:
                    pass

            return jsonify({
                'status': 'found',
                'id': best_appt['id'], 
                'date': fmt_date,
                'time': fmt_time,
                'is_today': is_today,
                'is_overdue': is_overdue
            })
        
        return jsonify({'status': 'not_found'})
    except Exception as e:
        print(f"[CHECK_APPOINTMENT_ERROR] {e}")
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/export_sessions')
@login_required
def export_sessions():
    import csv
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Get format parameter (default to csv for backward compatibility)
    export_format = request.args.get('format', 'csv').lower()

    conn = get_db_connection()

    # Get all sessions with full details
    sessions = conn.execute('''
        SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
               s.name as student_name, s.case_number, s.id as student_db_id,
               c.name as Counsellor_name,
               a.date, a.time, a.status as appointment_status
        FROM session sess
        LEFT JOIN Appointment a ON sess.appointment_id = a.id
        LEFT JOIN Student s ON a.student_id = s.id
        LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
        ORDER BY sess.created_at DESC
    ''').fetchall()

    conn.close()

    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sessions"

        # Define header style
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)

        # Write headers
        headers = ['ID', 'Date', 'Time', 'Student Name', 'Student ID',
                   'Counsellor', 'Session Type', 'Status', 'Notes', 'Created At']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_num, session in enumerate(sessions, 2):
            student_db_id = session.get(
                'student_db_id', 0) if session.get('student_db_id') else 0
            # Use case number for student ID in export
            professional_id = session.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')

            ws.cell(row=row_num, column=1, value=session['id'])
            ws.cell(row=row_num, column=2, value=session['date'] or 'N/A')
            ws.cell(row=row_num, column=3, value=session['time'] or 'N/A')
            ws.cell(row=row_num, column=4,
                    value=session['student_name'] or 'N/A')
            ws.cell(row=row_num, column=5, value=professional_id)
            ws.cell(row=row_num, column=6,
                    value=session['Counsellor_name'] or 'N/A')
            ws.cell(row=row_num, column=7,
                    value=session['session_type'] or 'N/A')
            ws.cell(row=row_num, column=8,
                    value=session['appointment_status'] or 'N/A')
            ws.cell(row=row_num, column=9, value=session['notes'] or 'N/A')
            ws.cell(row=row_num, column=10, value=session['created_at'])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=sessions_export.xlsx'
        return response
    else:
        # Create CSV (default)
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['ID', 'Date', 'Time', 'Student Name', 'Student ID',
                        'Counsellor', 'Session Type', 'Status', 'Notes', 'Created At'])

        # Write data rows
        for session in sessions:
            student_db_id = session.get(
                'student_db_id', 0) if session.get('student_db_id') else 0
            # Use case number for student ID in export
            professional_id = session.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')

            writer.writerow([
                session['id'],
                session['date'] or 'N/A',
                session['time'] or 'N/A',
                session['student_name'] or 'N/A',
                professional_id,
                session['Counsellor_name'] or 'N/A',
                session['session_type'] or 'N/A',
                session['appointment_status'] or 'N/A',
                (session['notes'] or '').replace('\n', ' ')[:200],
                session['created_at']
            ])

        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=sessions_export.csv'

    return response


@app.route('/referral', methods=['GET', 'POST'])
@login_required
def referral():
    try:
        ensure_database_initialized()

        if request.method == 'POST':
            # Get form data
            session_id = request.form.get('session_id')
            referred_by = request.form.get('referred_by')
            contact = request.form.get('contact')
            action_taken = request.form.get('action_taken', '')
            outcome = request.form.get('outcome', '')

            # Get selected referral reasons (checkboxes)
            selected_reasons = request.form.getlist('referral_reasons')
            other_reason_text = request.form.get(
                'other_reason_text', '').strip()

            # Combine reasons
            reasons_list = selected_reasons.copy()
            if 'Something Else' in selected_reasons and other_reason_text:
                # Replace "Something Else" with the actual text
                reasons_list.remove('Something Else')
                reasons_list.append(f'Something Else: {other_reason_text}')

            # Join reasons with commas, or use existing textarea value if checkboxes weren't used
            if reasons_list:
                reasons = ', '.join(reasons_list)
            else:
                # Fallback to old textarea
                reasons = request.form.get('reasons', '')

            # Validate required fields
            if not all([session_id, referred_by, contact]) or not reasons:
                flash(
                    'Please fill in all required fields and select at least one reason', 'error')
                return redirect(url_for('referral'))

            # Insert referral into database
            conn = get_db_connection()
            if conn is None:
                flash(
                    'Database connection failed. Please restart the application.', 'error')
                return redirect(url_for('dashboard'))

            try:
                import uuid as _uuid
                ref_gid = str(_uuid.uuid4())
                conn.execute('''
                    INSERT INTO Referral (session_id, referred_by, contact, reasons, action_taken, outcome, created_at, global_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (session_id, referred_by, contact, reasons, action_taken, outcome, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ref_gid))
                conn.commit()
                trigger_sync_immediate()
                flash('Referral created successfully!', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                conn.rollback()
                print(f"[REFERRAL] Error creating referral: {e}")
                flash(f'Error creating referral: {str(e)}', 'error')
                return redirect(url_for('referral'))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - display referral form
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        sessions = []
        try:
            sessions_raw = conn.execute('''
                SELECT s.id, s.created_at, st.name as student_name, st.case_number, st.index_number, c.name as Counsellor_name
                FROM session s
                LEFT JOIN Appointment a ON s.appointment_id = a.id
                LEFT JOIN Student st ON a.student_id = st.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                ORDER BY s.created_at DESC
            ''').fetchall()
            
            # Convert to list of dicts to avoid 'sqlite3.Row' has no attribute 'get' error
            sessions = [dict(row) for row in sessions_raw]
            
        except Exception as e:
            print(f"[REFERRAL] Error getting sessions: {e}")
            sessions = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('referral.html', sessions=sessions)
    except Exception as e:
        print(f"[REFERRAL] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading referral page. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/all_referrals')
@login_required
def all_referrals():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        referrals = []
        try:
            referrals_raw = conn.execute('''
                SELECT r.id, r.session_id, r.referred_by, r.contact, r.reasons, r.created_at,
                                st.id as student_db_id, st.name as student_name, st.case_number, st.contact as student_contact, st.index_number
                FROM Referral r
                LEFT JOIN session sess ON r.session_id = sess.id
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student st ON a.student_id = st.id
                ORDER BY r.created_at DESC
            ''').fetchall()

            # Convert to list and add professional ID
            referrals = []
            for ref in referrals_raw:
                ref_dict = dict(ref)
                ref_dict = dict(ref)
                # Use standardized case number format
                student_db_id = ref_dict.get('student_db_id', 0)
                ref_dict['professional_id'] = ref_dict.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')
                referrals.append(ref_dict)
        except Exception as e:
            print(f"[REFERRALS] Error getting referrals: {e}")
            referrals = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('referrals.html', referrals=referrals)
    except Exception as e:
        print(f"[REFERRALS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading referrals. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/export_referrals')
@login_required
def export_referrals():
    import csv
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Get format parameter (default to csv for backward compatibility)
    export_format = request.args.get('format', 'csv').lower()

    conn = get_db_connection()

    # Get all referrals with full details
    referrals_raw = conn.execute('''
        SELECT r.id, r.session_id, r.referred_by, r.contact, r.reasons, 
               r.action_taken, r.outcome, r.created_at,
               st.id as student_db_id, st.name as student_name, st.case_number, st.contact as student_contact
        FROM Referral r
        JOIN session sess ON r.session_id = sess.id
        JOIN Appointment a ON sess.appointment_id = a.id
        JOIN Student st ON a.student_id = st.id
        ORDER BY r.created_at DESC
    ''').fetchall()

    conn.close()

    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Referrals"

        # Define header style
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)

        # Write headers
        headers = ['ID', 'Date', 'Student Name', 'Student ID',
                   'Referred By', 'Contact', 'Reasons', 'Action Taken', 'Outcome']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows with professional IDs
        for row_num, referral in enumerate(referrals_raw, 2):
            ref_dict = dict(referral)
            student_db_id = ref_dict.get('student_db_id', 0)
            # Use standardized case number format
            professional_id = ref_dict.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')

            ws.cell(row=row_num, column=1, value=referral['id'])
            ws.cell(row=row_num, column=2, value=referral['created_at'])
            ws.cell(row=row_num, column=3,
                    value=referral['student_name'] or 'N/A')
            ws.cell(row=row_num, column=4, value=professional_id)
            ws.cell(row=row_num, column=5,
                    value=referral['referred_by'] or 'N/A')
            ws.cell(row=row_num, column=6, value=referral['contact'] or 'N/A')
            ws.cell(row=row_num, column=7, value=referral['reasons'] or 'N/A')
            ws.cell(row=row_num, column=8,
                    value=referral['action_taken'] or 'N/A')
            ws.cell(row=row_num, column=9, value=referral['outcome'] or 'N/A')

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=referrals_export.xlsx'
        return response
    else:
        # Create CSV (default)
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header - using "Student ID" instead of "Student Contact"
        writer.writerow(['ID', 'Date', 'Student Name', 'Student ID',
                        'Referred By', 'Contact', 'Reasons', 'Action Taken', 'Outcome'])

        # Write data rows with professional IDs
        for referral in referrals_raw:
            ref_dict = dict(referral)
            # Use standardized case number format
            student_db_id = ref_dict.get('student_db_id', 0)
            professional_id = ref_dict.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')

            writer.writerow([
                referral['id'],
                referral['created_at'],
                referral['student_name'] or 'N/A',
                professional_id,  # Use professional ID instead of phone number
                referral['referred_by'] or 'N/A',
                referral['contact'] or 'N/A',
                (referral['reasons'] or '').replace('\n', ' ')[
                    :200],  # Limit length and replace newlines
                (referral['action_taken'] or '').replace('\n', ' ')[:200],
                (referral['outcome'] or '').replace('\n', ' ')[:200]
            ])

        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=referrals_export.csv'

    return response


@app.route('/outcome_questionnaire', methods=['GET', 'POST'])
@login_required
def outcome_questionnaire():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            # Get form data
            student_id = request.form.get('student_id')
            session_id = request.form.get('session_id')
            age = request.form.get('age')
            sex = request.form.get('sex')

            # Get all 25 item scores
            item_scores = []
            try:
                for i in range(1, 26):
                    score = request.form.get(f'item{i}')
                    if not score or score == '':
                        flash('Please fill in all item scores', 'error')
                        return redirect(url_for('outcome_questionnaire'))
                    item_scores.append(int(score))
            except ValueError:
                flash('Invalid score values. Please enter numbers only.', 'error')
                return redirect(url_for('outcome_questionnaire'))

            # Calculate total score
            total_score = sum(item_scores)

            # Insert questionnaire data into database
            try:
                import uuid as _uuid
                oq_gid = str(_uuid.uuid4())
                # Use correct table name and column names (item1, item2, etc. without underscores)
                conn.execute('''
                    INSERT INTO OutcomeQuestionnaire 
                    (student_id, session_id, age, sex, item1, item2, item3, item4, item5,
                     item6, item7, item8, item9, item10, item11, item12, item13, item14, item15,
                     item16, item17, item18, item19, item20, item21, item22, item23, item24, item25,
                     total_score, created_at, global_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (student_id, session_id, age if age else None, sex if sex else None, *item_scores, total_score, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), oq_gid))
                conn.commit()
                from sync_engine import trigger_sync_immediate
                trigger_sync_immediate()
                flash('Outcome questionnaire submitted successfully!', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                conn.rollback()
                print(f"[OUTCOME_QUESTIONNAIRE] Error saving: {e}")
                flash(f'Error submitting questionnaire: {str(e)}', 'error')
                return redirect(url_for('outcome_questionnaire'))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - display the form
        students_raw = []
        sessions_raw = []
        try:
            students_raw = conn.execute(
                'SELECT id, name, case_number, programme FROM Student ORDER BY name').fetchall()
            sessions_raw = conn.execute('''
                SELECT sess.id, sess.created_at, a.student_id, s.id as student_db_id, s.name as student_name, s.case_number
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                WHERE a.student_id IS NOT NULL
                ORDER BY sess.created_at DESC
            ''').fetchall()
        except Exception as e:
            print(f"[OUTCOME_QUESTIONNAIRE] Error getting dropdown data: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Convert students to list
        students = []
        for student in students_raw:
            students.append(dict(student))

        # Convert sessions to list and format dates
        sessions = []
        for sess in sessions_raw:
            sess_dict = dict(sess)
            # Format created_at for display
            if sess_dict.get('created_at'):
                try:
                    dt = datetime.strptime(
                        sess_dict['created_at'], '%Y-%m-%d %H:%M:%S')
                    sess_dict['created_at_formatted'] = dt.strftime(
                        '%Y-%m-%d %H:%M')
                except:
                    sess_dict['created_at_formatted'] = sess_dict['created_at']
            else:
                sess_dict['created_at_formatted'] = 'N/A'

            # Add consistent professional ID (case_number)
            student_db_id = sess_dict.get('student_db_id', 0)
            sess_dict['professional_id'] = sess_dict.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')
            sessions.append(sess_dict)

        return render_template('outcome_questionnaire.html', students=students, sessions=sessions)
    except Exception as e:
        print(f"[OUTCOME_QUESTIONNAIRE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading outcome questionnaire page. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/dass21', methods=['GET', 'POST'])
@login_required
def dass21():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            # Get form data
            student_id = request.form.get('student_id', '').strip()
            depression_score = request.form.get('depression_score', '0')
            anxiety_score = request.form.get('anxiety_score', '0')
            stress_score = request.form.get('stress_score', '0')

            # Validate required fields
            if not student_id or student_id == '':
                flash('Please select a student', 'error')
                try:
                    students_raw = conn.execute(
                        'SELECT id, name, case_number, programme FROM Student ORDER BY name').fetchall()
                    students = [dict(s) for s in students_raw]
                except Exception as e:
                    print(f"[DASS21] Error getting students: {e}")
                    students = []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                return render_template('dass21.html', students=students)

            try:
                depression_score = float(depression_score)
                anxiety_score = float(anxiety_score)
                stress_score = float(stress_score)
            except ValueError:
                flash('Please enter valid numeric scores', 'error')
                try:
                    students_raw = conn.execute(
                        'SELECT id, name, case_number, programme FROM Student ORDER BY name').fetchall()
                    students = [dict(s) for s in students_raw]
                except Exception as e:
                    print(f"[DASS21] Error getting students: {e}")
                    students = []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                return render_template('dass21.html', students=students)

            # Calculate final scores (multiply by 2)
            final_depression = depression_score * 2
            final_anxiety = anxiety_score * 2
            final_stress = stress_score * 2

            # Insert DASS-21 scores into database
            try:
                import uuid as _uuid
                dass_gid = str(_uuid.uuid4())
                conn.execute('''
                    INSERT INTO DASS21 
                    (student_id, depression_score, anxiety_score, stress_score, created_at, global_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (student_id, depression_score, anxiety_score, stress_score,
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S'), dass_gid))
                conn.commit()
                from sync_engine import trigger_sync_immediate
                trigger_sync_immediate()
                trigger_sync_immediate()
                flash('DASS-21 scores saved successfully!', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                conn.rollback()
                print(f"[DASS21] Error saving scores: {e}")
                import traceback
                traceback.print_exc()
                flash(f'Error saving DASS-21 scores: {str(e)}', 'error')
                try:
                    students_raw = conn.execute(
                        'SELECT id, name, case_number, programme FROM Student ORDER BY name').fetchall()
                    students = [dict(s) for s in students_raw]
                except Exception as e:
                    print(f"[DASS21] Error getting students: {e}")
                    students = []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                return render_template('dass21.html', students=students)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - display the form
        students_raw = []
        try:
            students_raw = conn.execute(
                'SELECT id, name, case_number, programme FROM Student ORDER BY name').fetchall()
        except Exception as e:
            print(f"[DASS21] Error getting students: {e}")
            students_raw = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Convert to list
        students = []
        for student in students_raw:
            students.append(dict(student))

        return render_template('dass21.html', students=students)
    except Exception as e:
        print(f"[DASS21] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading DASS-21 page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/import_template/<import_type>')
@login_required
def import_template(import_type):
    """Download a blank CSV template for the chosen import type."""
    import io
    from flask import Response
    output = io.StringIO()
    if import_type == 'students':
        output.write("name,index_number,email,phone,department,programme,parent_contact,gender,age,hall_of_residence\n")
        output.write("Kofi Mensah,22334455,kofi@example.com,0244123456,Computer Science,BSc CS,0244000000,Male,21,Akuafo Hall\n")
    elif import_type == 'appointments':
        output.write("student_name,date,time,counsellor,purpose,status\n")
        output.write("Kofi Mensah,2026-04-01,10:00,Mrs. Gertrude Efa,Academic Stress,Scheduled\n")
    else:
        flash("Unknown template type", "error")
        return redirect(url_for('import_csv'))
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={"Content-Disposition": f"attachment; filename={import_type}_template.csv"}
    )


@app.route('/import_csv', methods=['GET', 'POST'])
@login_required
def import_csv():
    """
    General Import Utility for Students and Appointments.
    Complies with GTEC privacy (initials) and local encryption.
    """
    if request.method == 'POST':
        if 'confirm' in request.form:
            # ── Process confirmed import ──
            import_type = request.form.get('import_type')
            confirmed_data = request.form.get('confirmed_data')

            if not confirmed_data:
                flash('No data to import', 'error')
                return redirect(url_for('import_csv'))

            try:
                import traceback as _tb
                data = json.loads(confirmed_data)
                conn = get_db_connection()
                imported = 0
                row_errors = []

                if import_type == 'students':
                    for idx, row in enumerate(data, start=1):
                        try:
                            # ── GTEC Privacy & Encryption ──
                            raw_name = row.get('name', '').strip()
                            if not raw_name:
                                row_errors.append(f'Row {idx}: Missing name')
                                continue
                            name = raw_name  # Disable initials masking
                            
                            # Encrypt sensitive contact fields
                            contact        = encrypt_field(row.get('phone') or row.get('contact'))
                            parent_contact = encrypt_field(row.get('parent_contact'))
                            email          = encrypt_field(row.get('email'))

                            # Generate a case number for the new student
                            case_number = generate_case_number(conn, name)
                            import uuid
                            g_id = str(uuid.uuid4())

                            conn.execute('''
                                INSERT INTO Student
                                (name, index_number, email, contact, department, programme,
                                 parent_contact, gender, age, hall_of_residence, case_number, global_id)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                name,
                                row.get('index_number', ''),
                                email,
                                contact,
                                row.get('department', ''),
                                row.get('programme', ''),
                                parent_contact,
                                row.get('gender', ''),
                                row.get('age', ''),
                                row.get('hall_of_residence', ''),
                                case_number,
                                g_id
                            ))
                            imported += 1
                        except Exception as row_err:
                            row_errors.append(f'Row {idx} ({row.get("name","?")}): {row_err}')
                            print(f"[IMPORT] Skipping student row {idx}: {row_err}")
                            _tb.print_exc()

                elif import_type == 'appointments':
                    for row in data:
                        # Try to find student via index number (most reliable) then name
                        student = conn.execute(
                            'SELECT id FROM Student WHERE index_number = ? OR name = ?',
                            (row.get('index_number', ''), row.get('student_name', ''))
                        ).fetchone()
                        
                        if student:
                            conn.execute('''
                                INSERT INTO Appointment
                                (student_id, date, time, purpose, status)
                                VALUES (?, ?, ?, ?, ?)
                            ''', (
                                student['id'],
                                row.get('date', ''),
                                row.get('time', ''),
                                row.get('purpose', ''),
                                row.get('status', 'Scheduled')
                            ))
                            imported += 1

                conn.commit()
                trigger_sync_immediate()
                conn.close()
                if imported == 0 and row_errors:
                    flash(f'⚠️ Imported 0 records. Errors: {"; ".join(row_errors[:5])}', 'error')
                elif row_errors:
                    flash(f'✅ Imported {imported} {import_type}. Skipped {len(row_errors)} rows with errors.', 'warning')
                else:
                    flash(f'✅ Successfully imported {imported} {import_type}!', 'success')
                return redirect(url_for('students') if import_type == 'students' else url_for('manage_appointments'))

            except Exception as e:
                import traceback
                traceback.print_exc()
                flash(f'Error importing data: {str(e)}', 'error')
                return redirect(url_for('import_csv'))

        else:
            # ── Handle file upload and preview ──
            import_type = request.form.get('import_type')
            upload_file = request.files.get('csv_file')

            if not import_type or not upload_file:
                flash('Please select import type and a file', 'error')
                return redirect(url_for('import_csv'))

            filename = upload_file.filename.lower()
            preview_data = []
            errors = []

            try:
                # Support both CSV and Excel
                if filename.endswith('.xlsx') or filename.endswith('.xls'):
                    import openpyxl
                    wb = openpyxl.load_workbook(upload_file, data_only=True, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        flash('Excel file is empty', 'error')
                        return redirect(url_for('import_csv'))
                    
                    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
                    data_rows = [dict(zip(headers, [str(c).strip() if c is not None else '' for c in r])) for r in rows[1:]]
                else:
                    # CSV
                    content = upload_file.read()
                    try:
                        text = content.decode('utf-8')
                    except UnicodeDecodeError:
                        text = content.decode('latin-1')
                    import csv as csv_module
                    data_rows = list(csv_module.DictReader(text.splitlines()))

                for row_num, row in enumerate(data_rows, start=2):
                    # Normalise keys to lowercase
                    row = {k.lower().strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}

                    if import_type == 'students':
                        name = row.get('name') or row.get('full name') or row.get('student name') or ''
                        index = row.get('index_number') or row.get('index number') or row.get('id') or row.get('student id') or ''
                        if not name:
                            errors.append(f'Row {row_num}: Missing name — skipped')
                        else:
                            preview_data.append({
                                'name': name,
                                'index_number': index,
                                'email': row.get('email', ''),
                                'phone': row.get('phone') or row.get('contact') or row.get('phone number') or '',
                                'department': row.get('department', ''),
                                'programme': row.get('programme') or row.get('program') or row.get('course') or '',
                                'gender': row.get('gender', ''),
                                'age': row.get('age', ''),
                                'hall_of_residence': row.get('hall_of_residence') or row.get('hall') or '',
                                'parent_contact': row.get('parent_contact') or row.get('parent number') or ''
                            })

                    elif import_type == 'appointments':
                        sname = row.get('student_name') or row.get('student name') or row.get('name') or ''
                        index = row.get('index_number') or row.get('index number') or ''
                        date = row.get('date', '')
                        time = row.get('time', '')
                        if not sname or not date or not time:
                            errors.append(f'Row {row_num}: Missing student_name, date, or time — skipped')
                        else:
                            preview_data.append({
                                'student_name': sname,
                                'index_number': index,
                                'date': date,
                                'time': time,
                                'purpose': row.get('purpose', ''),
                                'status': row.get('status') or 'Scheduled'
                            })

                if not preview_data:
                    flash('No valid data rows found. Check the file format and column headers.', 'error')
                    return redirect(url_for('import_csv'))

                return render_template('import_csv.html',
                                       preview_data=preview_data,
                                       headers=list(preview_data[0].keys()) if preview_data else [],
                                       errors=errors,
                                       import_type=import_type,
                                       confirmed_data=json.dumps(preview_data))

            except Exception as e:
                flash(f'Error reading file: {str(e)}', 'error')
                return redirect(url_for('import_csv'))

    return render_template('import_csv.html')


@app.route('/intake', methods=['GET', 'POST'])
@login_required
def intake():
    """
    Phase 1: Secretary Intake Flow
    - Registers new student (if needed)
    - Creates Appointment
    - Records Urgency, Purpose, and Referral Source
    - Sets status to 'Scheduled' (Pending Handover)
    """
    if request.method == 'POST':
        conn = get_db_connection()
        try:
            # 1. Extract Student Info
            name = request.form.get('name')
            age = request.form.get('age')
            gender = request.form.get('gender')
            index_number = request.form.get('index_number')
            
            dept_base = request.form.get('department')
            dept_other = request.form.get('department_other')
            department = dept_other if dept_base == 'OTHER' else dept_base
            
            faculty = request.form.get('faculty')
            
            programme_base = request.form.get('programme')
            programme_other = request.form.get('programme_other')
            programme = programme_other if programme_base == 'Other' else programme_base
            
            contact = request.form.get('contact')
            parent_contact = request.form.get('parent_contact')
            hall = request.form.get('hall_of_residence')

            # 2. Extract Appointment/Intake Info
            appt_date = request.form.get('appointment_date')
            appt_time = request.form.get('appointment_time')
            purpose = request.form.get('purpose')
            urgency = request.form.get('urgency')
            referral = request.form.get('referral_source')

            # 3. Create/Check Student
            # Determine logic: assume New Client per form design, but check index_number to avoid dupes
            existing_student = conn.execute(
                "SELECT id FROM Student WHERE index_number = ?", (index_number,)).fetchone()

            if existing_student:
                student_id = existing_student['id']
                # Optional: Update contact info if changed
            else:
                case_num = generate_case_number(conn, name)
                import uuid as _uuid
                g_id = str(_uuid.uuid4())
                cursor = conn.execute('''
                    INSERT INTO Student (name, case_number, age, gender, index_number, department, 
                    faculty, programme, contact, parent_contact, hall_of_residence, global_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (name, case_num, age, gender, index_number, department, faculty, programme, contact, parent_contact, hall, g_id))
                student_id = cursor.lastrowid
                conn.commit()

            # 4. Create Appointment (The "Intake Record")
            import uuid as _uuid
            appt_gid = str(_uuid.uuid4())
            conn.execute('''
                INSERT INTO Appointment (student_id, date, time, purpose, status, urgency, referral_source, global_id)
                VALUES (?, ?, ?, ?, 'Scheduled', ?, ?, ?)
            ''', (student_id, appt_date, appt_time, purpose, urgency, referral, appt_gid))

            conn.commit()
            from sync_engine import trigger_sync_immediate
            trigger_sync_immediate()
            flash(
                'Student intake registered and appointment scheduled successfully.', 'success')
            return redirect(url_for('dashboard'))

        except Exception as e:
            conn.rollback()
            flash(f'Error processing intake: {str(e)}', 'danger')
            return redirect(url_for('intake'))
        finally:
            conn.close()

    return render_template('intake.html', now=datetime.now())


@app.route('/appointment', methods=['GET', 'POST'])
@login_required
def appointment():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            # Get form data
            student_id = request.form.get('student_id')
            appointment_date = request.form.get('date')
            appointment_time = request.form.get('time')
            purpose = request.form.get('purpose')
            counselor_id = request.form.get('counselor_id')
            urgency = request.form.get('urgency', 'Normal')
            referral = request.form.get('referral_source', 'Self')

            students = []
            counsellors = []
            try:
                students = conn.execute(
                    'SELECT id, name, case_number, programme, created_at FROM Student ORDER BY name').fetchall()
                counsellors = conn.execute(
                    'SELECT id, name FROM Counsellor ORDER BY name').fetchall()
            except Exception as e:
                print(f"[APPOINTMENT] Error getting dropdown data: {e}")

            # Validate data
            if not student_id or not appointment_date or not appointment_time or not counselor_id:
                try:
                    conn.close()
                except Exception:
                    pass
                flash('Please fill in all required fields', 'danger')
                return render_template('appointment.html', students=students, Counsellors=counsellors)

            # Check if student exists
            try:
                student = conn.execute(
                    'SELECT * FROM Student WHERE id = ?', (student_id,)).fetchone()
            except Exception as e:
                print(f"[APPOINTMENT] Error checking student: {e}")
                student = None

            if not student:
                try:
                    conn.close()
                except Exception:
                    pass
                flash(
                    'Student not found. Please check the ID or add the student first.', 'danger')
                return render_template('appointment.html', students=students, Counsellors=counsellors)

            # Save appointment to database
            try:
                import uuid as _uuid
                g_id = str(_uuid.uuid4())
                conn.execute('''
                        INSERT INTO Appointment (student_id, date, time, purpose, Counsellor_id, status, urgency, referral_source, global_id)
                    VALUES (?, ?, ?, ?, ?, 'Scheduled', ?, ?, ?)
                ''', (student_id, appointment_date, appointment_time, purpose, counselor_id, urgency, referral, g_id))
                conn.commit()
                from sync_engine import trigger_sync_immediate
                trigger_sync_immediate()
                flash('Appointment scheduled successfully!', 'success')
                return redirect(url_for('manage_appointments'))
            except Exception as e:
                conn.rollback()
                print(f"[APPOINTMENT] Error saving appointment: {e}")
                flash(f'Error scheduling appointment: {str(e)}', 'danger')
                return render_template('appointment.html', students=students, Counsellors=counsellors)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - display the form
        students = []
        counsellors = []
        try:
            students = conn.execute(
                'SELECT id, name, case_number, programme, created_at FROM Student ORDER BY name').fetchall()
            counsellors = conn.execute(
                'SELECT id, name FROM Counsellor ORDER BY name').fetchall()
        except Exception as e:
            print(f"[APPOINTMENT] Error getting dropdown data: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('appointment.html',
                               students=students,
                               Counsellors=counsellors,
                               selected_student_id=request.args.get('student_id'))
    except Exception as e:
        print(f"[APPOINTMENT] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading appointment page. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/manage_appointments')
@login_required
def manage_appointments():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        appointments = []
        try:
            # Get all appointments with student and counsellor names
            appointments = conn.execute('''
                SELECT a.*, s.name as student_name, s.id as student_record_id, s.case_number, s.index_number, s.created_at as student_created_at, c.name as Counsellor_name
                FROM Appointment a
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                ORDER BY a.date DESC, a.time DESC
            ''').fetchall()
            # Clean and format dates/times for the template
            appointments = [dict(a) for a in appointments]
            for a in appointments:
                a['date'] = clean_date_string(a['date'])
                a['time'] = clean_time_string(a['time'])
                
        except Exception as e:
            print(f"[APPOINTMENTS] Error getting appointments: {e}")
            appointments = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('appointments.html', appointments=appointments)
    except Exception as e:
        print(f"[APPOINTMENTS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading appointments. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/update_appointment_status/<int:appointment_id>', methods=['POST'])
@login_required
def update_appointment_status(appointment_id):
    """Update the status of an appointment"""
    new_status = request.form.get('status')

    if not new_status:
        flash('Status is required', 'error')
        return redirect(url_for('manage_appointments'))

    # Valid statuses
    valid_statuses = ['Scheduled', 'Sent to Counsellor',
                      'In Session', 'Completed', 'Cancelled', 'Postponed']
    if new_status not in valid_statuses:
        flash('Invalid status', 'error')
        return redirect(url_for('manage_appointments'))

    conn = get_db_connection()
    try:
        # Check if appointment exists
        appointment = conn.execute(
            'SELECT id FROM Appointment WHERE id = ?', (appointment_id,)).fetchone()

        if not appointment:
            flash('Appointment not found', 'error')
            return redirect(url_for('manage_appointments'))

        # Update the status
        conn.execute('''
            UPDATE Appointment 
            SET status = ? 
            WHERE id = ?
        ''', (new_status, appointment_id))
        conn.commit()

        flash(
            f'Appointment status updated to {new_status} successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error updating appointment status: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('manage_appointments'))

# ---------- Print Routes ----------


@app.route('/view_report/<int:report_id>')
@login_required
def view_report(report_id):
    conn = get_db_connection()

    # Get report details
    report = conn.execute('''
        SELECT * FROM reports WHERE id = ?
    ''', (report_id,)).fetchone()

    if not report:
        conn.close()
        flash('Report not found', 'error')
        return redirect(url_for('reports_list'))

    conn.close()
    # Convert Row to dict for easier template access
    report_dict = dict(report)
    return render_template('view_report.html', report=report_dict, now=datetime.utcnow())


@app.route('/download_report/<int:report_id>')
@login_required
def download_report(report_id):
    conn = get_db_connection()

    # Get report details
    report = conn.execute('''
        SELECT * FROM reports WHERE id = ?
    ''', (report_id,)).fetchone()

    if not report:
        conn.close()
        flash('Report not found', 'error')
        return redirect(url_for('reports_list'))

    conn.close()

    # For now, redirect to print report as download functionality
    # This can be enhanced to generate actual downloadable files
    return redirect(url_for('print_report', report_id=report_id))


@app.route('/delete_report/<int:report_id>', methods=['POST'])
@login_required
def delete_report(report_id):
    conn = get_db_connection()

    # Delete report from database
    try:
        result = conn.execute(
            'DELETE FROM reports WHERE id = ?', (report_id,)).rowcount
        conn.commit()

        if result > 0:
            flash('Report deleted successfully!', 'success')
        else:
            flash('Report not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting report: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('reports_list'))


@app.route('/delete_student/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    """Delete a student and all related records"""
    conn = get_db_connection()

    try:
        # Check if student exists
        student = conn.execute(
            'SELECT * FROM Student WHERE id = ?', (student_id,)).fetchone()
        if not student:
            flash('Student not found', 'error')
            return redirect(url_for('students'))

        # Delete related records first (due to foreign keys)
        # Delete referrals (through sessions through appointments)
        conn.execute('''
            DELETE FROM Referral 
            WHERE session_id IN (
                SELECT s.id FROM session s
                JOIN Appointment a ON s.appointment_id = a.id
                WHERE a.student_id = ?
            )
        ''', (student_id,))

        # Delete sessions
        conn.execute('''
            DELETE FROM session
            WHERE appointment_id IN (
                SELECT id FROM Appointment WHERE student_id = ?
            )
        ''', (student_id,))

        # Delete appointments
        conn.execute(
            'DELETE FROM Appointment WHERE student_id = ?', (student_id,))

        # Delete assessments
        conn.execute('DELETE FROM DASS21 WHERE student_id = ?', (student_id,))
        conn.execute(
            'DELETE FROM OutcomeQuestionnaire WHERE student_id = ?', (student_id,))

        # Finally delete the student
        result = conn.execute(
            'DELETE FROM Student WHERE id = ?', (student_id,)).rowcount
        conn.commit()

        if result > 0:
            flash('Student and all related records deleted successfully!', 'success')
        else:
            flash('Student not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting student: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('students'))


@app.route('/delete_appointment/<int:appointment_id>', methods=['POST'])
@login_required
def delete_appointment(appointment_id):
    """Delete an appointment and related sessions"""
    conn = get_db_connection()

    try:
        # Check if appointment exists
        appointment = conn.execute(
            'SELECT * FROM Appointment WHERE id = ?', (appointment_id,)).fetchone()
        if not appointment:
            flash('Appointment not found', 'error')
            return redirect(url_for('manage_appointments'))

        # Delete related sessions first
        conn.execute(
            'DELETE FROM session WHERE appointment_id = ?', (appointment_id,))

        # Delete the appointment
        result = conn.execute(
            'DELETE FROM Appointment WHERE id = ?', (appointment_id,)).rowcount
        conn.commit()

        if result > 0:
            flash('Appointment deleted successfully!', 'success')
        else:
            flash('Appointment not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting appointment: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('manage_appointments'))


@app.route('/delete_session/<int:session_id>', methods=['POST'])
@login_required
def delete_session(session_id):
    """Delete a session and related records"""
    conn = get_db_connection()

    try:
        # Check if session exists
        session_record = conn.execute(
            'SELECT * FROM session WHERE id = ?', (session_id,)).fetchone()
        if not session_record:
            flash('Session not found', 'error')
            return redirect(url_for('sessions_list'))

        # Delete related referrals
        conn.execute(
            'DELETE FROM Referral WHERE session_id = ?', (session_id,))

        # Delete case management records
        conn.execute(
            'DELETE FROM CaseManagement WHERE session_id = ?', (session_id,))

        # Delete the session
        result = conn.execute(
            'DELETE FROM session WHERE id = ?', (session_id,)).rowcount
        conn.commit()

        if result > 0:
            flash('Session deleted successfully!', 'success')
        else:
            flash('Session not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting session: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('sessions_list'))


@app.route('/delete_referral/<int:referral_id>', methods=['POST'])
@login_required
def delete_referral(referral_id):
    """Delete a referral"""
    conn = get_db_connection()

    try:
        result = conn.execute(
            'DELETE FROM Referral WHERE id = ?', (referral_id,)).rowcount
        conn.commit()

        if result > 0:
            flash('Referral deleted successfully!', 'success')
        else:
            flash('Referral not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting referral: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('all_referrals'))


@app.route('/get_session/<int:session_id>')
@login_required
def get_session(session_id):
    """Get session details as JSON for the modal"""
    conn = get_db_connection()

    try:
        session_data = conn.execute('''
            SELECT sess.id, sess.session_type, sess.notes, sess.outcome, sess.created_at,
                   s.name as student_name, s.id as student_record_id, s.case_number, s.index_number, s.department, s.programme, s.created_at as student_created_at,
                   c.name as Counsellor_name,
                   a.date as appt_date, a.time as appt_time, a.status, a.purpose
            FROM session sess
            LEFT JOIN Appointment a ON sess.appointment_id = a.id
            LEFT JOIN Student s ON a.student_id = s.id
            LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
            WHERE sess.id = ?
        ''', (session_id,)).fetchone()

        if not session_data:
            conn.close()
            return jsonify({'error': 'Session not found'}), 404

        # Get clinical ID for masking
        clinical_id = get_clinical_id(session_data['student_name'], 
                                       session_data['student_record_id'], 
                                       session_data['student_created_at'])

        # Convert to dict for JSON serialization (Using correct aliased column names)
        session_dict = {
            'id': session_data['id'],
            'student_name': clinical_id,
            'Counsellor_name': session_data['Counsellor_name'] or 'N/A',
            'date': clean_date_string(session_data['appt_date']),
            'time': clean_time_string(session_data['appt_time']),
            'session_type': session_data['session_type'] or 'N/A',
            'appointment_status': session_data['status'] if 'status' in session_data.keys() else 'N/A',
            'notes': session_data['notes'] or 'No notes provided',
            'outcome': session_data['outcome'] or 'N/A',
            'created_at': clean_date_string(session_data['created_at'])
        }

        conn.close()
        return jsonify(session_dict)
    except Exception as e:
        import traceback
        print(f"[GET_SESSION_ERROR] {e}\n{traceback.format_exc()}")
        try: conn.close()
        except: pass
        return jsonify({'error': f"Internal Error: {str(e)}"}), 500


@app.route('/print_session/<int:session_id>')
@login_required
def print_session(session_id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

    # Get session details with student information
        session_data = None
        try:
            session_data = conn.execute('''
                SELECT sess.*, s.name as student_name, s.id as student_record_id, s.created_at as student_created_at,
                       s.index_number, s.programme, s.contact, s.department,
                       c.name as Counsellor_name,
                       a.date, a.time, a.status as appointment_status
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE sess.id = ?
            ''', (session_id,)).fetchone()
        except Exception as e:
            print(f"[PRINT_SESSION] Error getting session: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if not session_data:
            flash('Session not found', 'error')
            return redirect(url_for('dashboard'))

        return render_template('print_session.html', session_data=session_data)
    except Exception as e:
        print(f"[PRINT_SESSION] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading session for printing.', 'error')
        return redirect(url_for('sessions_list'))


@app.route('/print_referral/<int:id>')
@login_required
def print_referral(id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

    # Get referral details with student information
        referral = None
        student_info = None
        try:
            referral = conn.execute('''
                SELECT r.*, s.id as student_db_id, s.name as student_name, 
                       s.index_number, s.contact as student_contact, s.department as student_department,
                       sess.created_at as session_date,
                       a.date as appointment_date, a.time as appointment_time
                FROM Referral r
                JOIN session sess ON r.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                JOIN Student s ON a.student_id = s.id
        WHERE r.id = ?
    ''', (id,)).fetchone()

            referral_dict = dict(referral) if referral else None

            # Parse reasons from comma-separated string
            reasons_str = referral['reasons'] or ''
            referral_reasons_list = [r.strip()
                                     for r in reasons_str.split(',') if r.strip()]

            # Check for "Something Else" and extract text
            other_reason_text = None
            if referral_reasons_list:
                for idx, reason in enumerate(referral_reasons_list):
                    if reason.startswith('Something Else:'):
                        other_reason_text = reason.replace(
                            'Something Else:', '').strip()
                        # Replace with just the checkbox name
                        referral_reasons_list[idx] = 'Something Else'
                        break

            # Use consistent case_number as professional ID
            if referral and referral_dict.get('case_number'):
                referral_dict['professional_id'] = referral_dict['case_number']
            elif referral and referral_dict.get('student_db_id'):
                student_db_id = referral_dict.get('student_db_id', 0)
                referral_dict['professional_id'] = f"GCC-{datetime.now().year}-{student_db_id:04d}"
            else:
                referral_dict['professional_id'] = 'N/A'
            referral_dict['referral_reasons_list'] = referral_reasons_list
            referral_dict['other_reason_text'] = other_reason_text
            student_department = referral_dict.get(
                'student_department', 'N/A') if referral else 'N/A'

        except Exception as e:
            print(f"[PRINT_REFERRAL] Error getting referral: {e}")
            import traceback
            traceback.print_exc()
            referral = None
            referral_dict = None
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if not referral or not referral_dict:
            # Only flash error if it's a legitimate "not found" case (not a database error)
            if referral is None:
                flash('Referral not found', 'error')
            else:
                flash('Error loading referral details. Please try again.', 'error')
            return redirect(url_for('all_referrals'))

        return render_template('print_referral.html',
                               referral=referral_dict,
                               referral_reasons_list=referral_reasons_list,
                               other_reason_text=other_reason_text,
                               student_department=student_department)
    except Exception as e:
        print(f"[PRINT_REFERRAL] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading referral for printing.', 'error')
        return redirect(url_for('all_referrals'))


@app.route('/print_case/<int:case_id>')
@login_required
def print_case(case_id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

    # Get case details with student information
        case = None
        try:
            case = conn.execute('''
                SELECT cm.*, s.name as student_name, 
                       s.index_number, s.programme, s.contact,
                       sess.created_at as session_date
                FROM CaseManagement cm
                JOIN session sess ON cm.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                JOIN Student s ON a.student_id = s.id
                WHERE cm.id = ?
            ''', (case_id,)).fetchone()
        except Exception as e:
            print(f"[PRINT_CASE] Error getting case: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if not case:
            flash('Case not found', 'error')
            return redirect(url_for('dashboard'))

        return render_template('print_case.html', case=case)
    except Exception as e:
        print(f"[PRINT_CASE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case for printing.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/case_notes_list')
@login_required
def case_notes_list():
    """Display all case notes linked to students"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        case_notes = []
        try:
            case_notes_raw = conn.execute('''
                SELECT cm.id, cm.session_id, cm.client_appearance, cm.problems, cm.interventions,
                       cm.recommendations, cm.next_visit_date, cm.counsellor_signature, cm.created_at,
                       s.id as student_db_id, s.name as student_name, s.index_number, s.programme, s.case_number,
                       sess.created_at as session_date
                FROM CaseManagement cm
                JOIN session sess ON cm.session_id = sess.id
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                ORDER BY cm.created_at DESC
            ''').fetchall()

            # Convert to list and add professional IDs
            for cn in case_notes_raw:
                cn_dict = dict(cn)
                student_db_id = cn_dict.get('student_db_id', 0)
                # Use consistent case_number as professional ID
                cn_dict['professional_id'] = cn_dict.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')
                case_notes.append(cn_dict)
        except Exception as e:
            print(f"[CASE_NOTES_LIST] Error getting case notes: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('case_notes_list.html', case_notes=case_notes)
    except Exception as e:
        print(f"[CASE_NOTES_LIST] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case notes. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/dass21_list')
@login_required
def dass21_list():
    """Display all DASS-21 records linked to students"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        dass21_records = []
        try:
            dass21_raw = conn.execute('''
                SELECT d.id, d.student_id, d.depression_score, d.anxiety_score, d.stress_score,
                       d.created_at, d.completion_date,
                       s.id as student_db_id, s.name as student_name, s.index_number, s.programme, s.case_number
                FROM DASS21 d
                LEFT JOIN Student s ON d.student_id = s.id
                ORDER BY d.created_at DESC
            ''').fetchall()

            # Convert to list and add professional IDs and final scores
            for d in dass21_raw:
                d_dict = dict(d)
                student_db_id = d_dict.get('student_db_id', 0)
                # Use consistent case_number as professional ID
                d_dict['professional_id'] = d_dict.get('case_number') or (f"GCC-{datetime.now().year}-{student_db_id:04d}" if student_db_id else 'N/A')
                # Calculate final scores (x2)
                d_dict['final_depression'] = (
                    d_dict.get('depression_score', 0) or 0) * 2
                d_dict['final_anxiety'] = (
                    d_dict.get('anxiety_score', 0) or 0) * 2
                d_dict['final_stress'] = (
                    d_dict.get('stress_score', 0) or 0) * 2
                dass21_records.append(d_dict)
        except Exception as e:
            print(f"[DASS21_LIST] Error getting DASS-21 records: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('dass21_list.html', dass21_records=dass21_records)
    except Exception as e:
        print(f"[DASS21_LIST] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading DASS-21 records. Please try again.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/print_case_note/<int:case_id>')
@login_required
def print_case_note(case_id):
    """Print case note"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        case = None
        try:
            case = conn.execute('''
                SELECT cm.*, s.id as student_db_id, s.name as student_name, 
                       s.index_number, s.programme, s.contact, s.department, s.case_number,
                       sess.created_at as session_date
                FROM CaseManagement cm
                JOIN session sess ON cm.session_id = sess.id
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                WHERE cm.id = ?
            ''', (case_id,)).fetchone()

            if case:
                case_dict = dict(case)
                # Use consistent case_number as professional ID
                case_dict['professional_id'] = case_dict.get('case_number') or (f"GCC-{datetime.now().year}-{case_dict.get('student_db_id', 0):04d}" if case_dict.get('student_db_id') else 'N/A')
                case = case_dict
        except Exception as e:
            print(f"[PRINT_CASE_NOTE] Error getting case: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if not case:
            flash('Case note not found', 'error')
            return redirect(url_for('case_notes_list'))

        return render_template('print_case_note.html', case=case)
    except Exception as e:
        print(f"[PRINT_CASE_NOTE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case note for printing.', 'error')
        return redirect(url_for('case_notes_list'))


@app.route('/print_dass21/<int:dass21_id>')
@login_required
def print_dass21(dass21_id):
    """Print DASS-21 record"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        dass21_record = None
        try:
            dass21_raw = conn.execute('''
                SELECT d.*, s.id as student_db_id, s.name as student_name, 
                       s.index_number, s.programme, s.contact, s.department, s.case_number
                FROM DASS21 d
                LEFT JOIN Student s ON d.student_id = s.id
                WHERE d.id = ?
            ''', (dass21_id,)).fetchone()

            if dass21_raw:
                dass21_dict = dict(dass21_raw)
                # Use consistent case_number as professional ID
                dass21_dict['professional_id'] = dass21_dict.get('case_number') or (f"GCC-{datetime.now().year}-{dass21_dict.get('student_db_id', 0):04d}" if dass21_dict.get('student_db_id') else 'N/A')
                # Calculate final scores (x2)
                dass21_dict['final_depression'] = (
                    dass21_dict.get('depression_score', 0) or 0) * 2
                dass21_dict['final_anxiety'] = (
                    dass21_dict.get('anxiety_score', 0) or 0) * 2
                dass21_dict['final_stress'] = (
                    dass21_dict.get('stress_score', 0) or 0) * 2
                dass21_record = dass21_dict
        except Exception as e:
            print(f"[PRINT_DASS21] Error getting DASS-21 record: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if not dass21_record:
            flash('DASS-21 record not found', 'error')
            return redirect(url_for('dass21_list'))

        return render_template('print_dass21.html', dass21=dass21_record)
    except Exception as e:
        print(f"[PRINT_DASS21] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading DASS-21 record for printing.', 'error')
        return redirect(url_for('dass21_list'))


@app.route('/print_report/<int:report_id>')
@login_required
def print_report(report_id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

    # Get report details
        report = None
        try:
            report = conn.execute('''
                SELECT * FROM reports WHERE id = ?
            ''', (report_id,)).fetchone()
        except Exception as e:
            print(f"[PRINT_REPORT] Error getting report: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass

        if not report:
            flash('Report not found', 'error')
            return redirect(url_for('reports_list'))

        # Convert Row to dict for easier template access
        report_dict = dict(report) if report else {}
        return render_template('print_report.html', report=report_dict, now=datetime.utcnow())
    except Exception as e:
        print(f"[PRINT_REPORT] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading report for printing.', 'error')
        return redirect(url_for('reports_list'))


@app.route('/statistics')
@login_required
def statistics():
    """Display comprehensive statistics with charts"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        # Initialize with defaults
        total_students = 0
        total_appointments = 0
        total_sessions = 0
        total_referrals = 0
        gender_stats = []
        programme_stats = []
        level_stats = []
        appointment_status_stats = []

        try:
            # Overall Statistics
            result = conn.execute(
                'SELECT COUNT(*) as count FROM Student').fetchone()
            total_students = result['count'] if result else 0

            result = conn.execute(
                'SELECT COUNT(*) as count FROM Appointment').fetchone()
            total_appointments = result['count'] if result else 0

            result = conn.execute(
                'SELECT COUNT(*) as count FROM session').fetchone()
            total_sessions = result['count'] if result else 0

            try:
                result = conn.execute(
                    'SELECT COUNT(*) as count FROM Referral').fetchone()
                total_referrals = result['count'] if result else 0
            except:
                total_referrals = 0

            # Students by Gender
            try:
                gender_stats = conn.execute('''
                    SELECT COALESCE(gender, 'Not Specified') as gender, COUNT(*) as count 
                    FROM Student 
                    GROUP BY gender
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting gender stats: {e}")
                gender_stats = []

            # Students by Programme
            try:
                programme_stats = conn.execute('''
                    SELECT COALESCE(programme, 'Not Specified') as programme, COUNT(*) as count 
                    FROM Student 
                    GROUP BY programme
                    ORDER BY count DESC
                    LIMIT 10
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting programme stats: {e}")
                programme_stats = []

            # Students by Department
            try:
                department_stats = conn.execute('''
                    SELECT COALESCE(department, 'Not Specified') as department, COUNT(*) as count 
                    FROM Student 
                    GROUP BY department
                    ORDER BY count DESC
                    LIMIT 10
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting department stats: {e}")
                department_stats = []

            # Appointments by Status
            try:
                appointment_status_stats = conn.execute('''
                    SELECT COALESCE(status, 'Not Specified') as status, COUNT(*) as count 
                    FROM Appointment 
                    GROUP BY status
                ''').fetchall()
            except Exception as e:
                print(
                    f"[STATISTICS] Error getting appointment status stats: {e}")
                appointment_status_stats = []

            # Appointments Over Time (Last 6 Months)
            try:
                appointment_timeline = conn.execute('''
                    SELECT strftime('%Y-%m', date) as month, COUNT(*) as count 
                    FROM Appointment 
                    WHERE date >= date('now', '-6 months')
                    GROUP BY month
                    ORDER BY month
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting appointment timeline: {e}")
                appointment_timeline = []

            # Sessions Over Time (Last 6 Months)
            try:
                session_timeline = conn.execute('''
                    SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
                    FROM session 
                    WHERE created_at >= date('now', '-6 months')
                    GROUP BY month
                    ORDER BY month
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting session timeline: {e}")
                session_timeline = []

            # Sessions by Type
            try:
                session_type_stats = conn.execute('''
                    SELECT COALESCE(session_type, 'Not Specified') as session_type, COUNT(*) as count 
                    FROM session 
                    GROUP BY session_type
                    ORDER BY count DESC
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting session type stats: {e}")
                session_type_stats = []

            # Age Distribution
            try:
                age_distribution = conn.execute('''
                    SELECT 
                        CASE 
                            WHEN age < 18 THEN 'Under 18'
                            WHEN age BETWEEN 18 AND 20 THEN '18-20'
                            WHEN age BETWEEN 21 AND 25 THEN '21-25'
                            WHEN age BETWEEN 26 AND 30 THEN '26-30'
                            WHEN age > 30 THEN 'Over 30'
                            ELSE 'Not Specified'
                        END as age_group,
                        COUNT(*) as count
                    FROM Student
                    GROUP BY age_group
                    ORDER BY 
                        CASE age_group
                            WHEN 'Under 18' THEN 1
                            WHEN '18-20' THEN 2
                            WHEN '21-25' THEN 3
                            WHEN '26-30' THEN 4
                            WHEN 'Over 30' THEN 5
                            ELSE 6
                        END
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting age distribution: {e}")
                age_distribution = []

            # Top Counsellors by Appointments
            try:
                top_counsellors = conn.execute('''
                    SELECT c.name, COUNT(a.id) as appointment_count
                    FROM Counsellor c
                    LEFT JOIN Appointment a ON c.id = a.Counsellor_id
                    GROUP BY c.id, c.name
                    ORDER BY appointment_count DESC
                    LIMIT 5
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting top counsellors: {e}")
                top_counsellors = []

            # Monthly New Students (Last 6 Months)
            try:
                new_students_timeline = conn.execute('''
                    SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
                    FROM Student 
                    WHERE created_at >= date('now', '-6 months')
                    GROUP BY month
                    ORDER BY month
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting new students timeline: {e}")
                new_students_timeline = []

        except Exception as e:
            print(f"[STATISTICS] Error in statistics queries: {e}")
        finally:
            conn.close()

        # Convert to dictionaries for JSON serialization (use safe defaults if missing)
        stats_data = {
            'overall': {
                'total_students': total_students,
                'total_appointments': total_appointments,
                'total_sessions': total_sessions,
                'total_referrals': total_referrals
            },
            'gender': [{'gender': row['gender'], 'count': row['count']} for row in gender_stats] if gender_stats else [],
            'programme': [{'programme': row['programme'], 'count': row['count']} for row in programme_stats] if programme_stats else [],
            'department': [{'department': row['department'], 'count': row['count']} for row in department_stats] if department_stats else [],
            'appointment_status': [{'status': row['status'], 'count': row['count']} for row in appointment_status_stats] if appointment_status_stats else [],
            'appointment_timeline': [{'month': row['month'], 'count': row['count']} for row in appointment_timeline] if appointment_timeline else [],
            'session_timeline': [{'month': row['month'], 'count': row['count']} for row in session_timeline] if session_timeline else [],
            'session_type': [{'session_type': row['session_type'], 'count': row['count']} for row in session_type_stats] if session_type_stats else [],
            'age_distribution': [{'age_group': row['age_group'], 'count': row['count']} for row in age_distribution] if age_distribution else [],
            'top_counsellors': [{'name': row['name'], 'count': row['appointment_count']} for row in top_counsellors] if top_counsellors else [],
            'new_students_timeline': [{'month': row['month'], 'count': row['count']} for row in new_students_timeline] if new_students_timeline else []
        }

        return render_template('statistics.html', stats_data=stats_data)

    except Exception as e:
        print(f"[STATISTICS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading statistics: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

# Add error handler for 500 errors - display them properly


@app.errorhandler(500)
def internal_error(error):
    import traceback
    error_trace = traceback.format_exc()
    print(f"[ERROR 500] {error}")
    print(f"[ERROR 500] Traceback:\n{error_trace}")

    # Try to log to file if running as EXE
    try:
        if getattr(sys, 'frozen', False):
            try:
                base_path = os.path.dirname(sys.executable)
            except:
                base_path = os.path.dirname(os.path.abspath(__file__))
            error_log_path = os.path.join(base_path, 'error_log.txt')
            with open(error_log_path, 'a') as f:
                f.write(f"\n=== ERROR {datetime.now()} ===\n")
                f.write(f"{error}\n")
                f.write(f"{error_trace}\n")
                f.write("=" * 50 + "\n")
            print(f"[ERROR] Details logged to: {error_log_path}")
    except:
        pass

    return '''
    <html>
    <head><title>Internal Server Error</title></head>
    <body style="font-family: sans-serif; padding: 2rem; line-height: 1.6;">
        <h1 style="color: #dc3545;">Internal Server Error</h1>
        <p>The server encountered an error while processing your request.</p>
        <div style="background: #f8f9fa; padding: 1rem; border-radius: 4px; border: 1px solid #dee2e6; margin: 1rem 0;">
            <code>''' + str(error) + '''</code>
        </div>
        <p><a href="/welcome" style="display: inline-block; background: #0d6efd; color: white; padding: 0.5rem 1rem; text-decoration: none; border-radius: 4px;">Return to Login</a></p>
    </body>
    </html>
    ''', 500


@app.errorhandler(Exception)
def handle_exception(e):
    # Pass through HTTP errors (like 404, 405, etc.) so Flask handles them normally
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e

    # For actual unhandled code exceptions, log and return 500
    import traceback
    error_trace = traceback.format_exc()
    print(f"[UNHANDLED ERROR] {e}")
    print(f"[UNHANDLED ERROR] Traceback:\n{error_trace}")
    return internal_error(e)


@app.route('/admin/bookings/<int:booking_id>/register')
def register_portal_booking(booking_id):
    """Manually register a portal booking: Create Student and Appointment."""
    is_ajax = request.args.get('ajax') == '1'
    if session.get('role') not in ['Secretary', 'Admin', 'Counsellor', 'Counselor']:
        if is_ajax: return jsonify({'status': 'error', 'message': 'Access restricted.'}), 403
        flash('Access restricted.', 'error')
        return redirect(url_for('dashboard'))
        
    try:
        conn = get_db_connection()
        conn.row_factory = sqlite3.Row
        
        # 1. Fetch booking
        booking_row = conn.execute("SELECT * FROM BookingRequest WHERE id = ?", (booking_id,)).fetchone()
        if not booking_row:
            if is_ajax: return jsonify({'status': 'error', 'message': 'Booking not found.'}), 404
            flash('Booking not found.', 'error')
            conn.close()
            return redirect(url_for('admin_bookings'))
        
        booking = dict(booking_row)
            
        if booking['status'] == 'Accepted' and conn.execute("SELECT 1 FROM Appointment WHERE booking_ref = ?", (booking['reference'],)).fetchone():
             if is_ajax: return jsonify({'status': 'info', 'message': 'Already registered.'}), 200
             flash('This booking is already registered.', 'info')
             conn.close()
             return redirect(url_for('admin_bookings'))

        # 2. Assign a counsellor (default to first one if none specified)
        counsellor = conn.execute("SELECT id FROM Counsellor LIMIT 1").fetchone()
        counsellor_id = counsellor['id'] if counsellor else None

        # 3. Check/Create student
        # First check by Index Number
        student = conn.execute("SELECT id FROM Student WHERE index_number = ?", (booking['index_number'],)).fetchone()
        
        if not student:
            # Then check by Name because the table erroneously has a UNIQUE constraint on 'name'
            student = conn.execute("SELECT id FROM Student WHERE name = ?", (booking['full_name'],)).fetchone()

        if not student:
            # Generate both identifiers as per spec
            case_num = generate_case_number(conn, booking['full_name'])
            import uuid as _uuid
            prof_id = generate_professional_id(conn, booking['full_name'])
            g_id = str(_uuid.uuid4())
            
            cursor = conn.cursor()
            cursor.execute(
                '''INSERT INTO Student (name, case_number, global_id, index_number, department, programme, program, faculty, hall_of_residence, contact, email, age, gender)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (booking['full_name'], case_num, g_id, booking['index_number'],
                 booking['department'], booking['programme'], booking['programme'], 
                 booking.get('faculty', 'General'), booking.get('hall_of_residence'), 
                 booking['phone'], booking.get('email'), booking.get('age'), booking.get('gender'))
            )
            student_id = cursor.lastrowid
        else:
            student_id = student['id']
            # Optionally update student info if it changed? For now, just use existing ID

        # 4. Create appointment (Cleaning date/time strings first)
        clean_d = clean_date_string(booking['preferred_date'] or datetime.now().strftime('%Y-%m-%d'))
        clean_t = clean_time_string(booking['preferred_time'] or '09:00')
        
        import uuid as _uuid
        appt_gid = str(_uuid.uuid4())
        conn.execute(
            '''INSERT INTO Appointment (student_id, Counsellor_id, date, time, purpose, status, booking_ref, global_id)
               VALUES (?, ?, ?, ?, ?, 'Scheduled', ?, ?)''',
            (student_id, counsellor_id, clean_d, clean_t,
             f"[Portal] {booking['reason'] or 'Session request'}",
             booking['reference'], appt_gid)
        )
        
        # 5. Update booking status
        conn.execute("UPDATE BookingRequest SET status = 'Accepted', accepted_at = CURRENT_TIMESTAMP WHERE id = ?", (booking_id,))
        
        conn.commit()
        from sync_engine import trigger_sync_immediate
        trigger_sync_immediate()
        conn.close()
        msg = f"Successfully registered {booking['full_name']} and scheduled appointment."
        if is_ajax: return jsonify({'status': 'success', 'message': msg}), 200
        flash(msg, 'success')
        
    except Exception as e:
        print(f"[REGISTRATION ERROR] {e}")
        if is_ajax: return jsonify({'status': 'error', 'message': str(e)}), 500
        flash(f"Error during registration: {str(e)}", 'error')
        
    return redirect(request.referrer or url_for('admin_bookings'))

def run_auto_sync_loop():
    """Background thread to auto-sync every 10 seconds"""
    print("--- Auto-Sync Service Started ---")
    while True:
        try:
            # Check if sync is enabled and peer IP is set
            config = node_config.load_config()
            peer_ip = config.get('peer_ip')
            if peer_ip:
                # Trigger sync silently
                # We use a slight delay or check to avoid spamming if offline
                trigger_sync_immediate()

        except Exception as e:
            print(f"[AUTO-SYNC] Error: {e}")

        # Sleep for 10 seconds before next sync attempt
        time.sleep(10)


# ==========================================
# STUDENT BOOKING PORTAL (Public Routes)
# ==========================================

@app.route('/api/submit_booking', methods=['POST', 'OPTIONS'])
def api_submit_booking():
    """API endpoint for standalone booking form."""
    if request.method == 'OPTIONS':
        # Provide CORS headers
        response = make_response()
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "POST")
        return response

    if request.method == 'POST':
        try:
            conn = get_db_connection()
            data = request.get_json()
            if not data:
                return jsonify({'status': 'error', 'message': 'Invalid data format'}), 400

            full_name = data.get('full_name', '').strip()
            index_number = data.get('index_number', '').strip()
            department = data.get('department', '').strip()
            programme = data.get('programme', '').strip()
            phone = data.get('phone', '').strip()
            email = data.get('email', '').strip()
            hall_of_residence = data.get('hall_of_residence', '').strip()
            preferred_date = data.get('preferred_date', '').strip()
            preferred_time = data.get('preferred_time', 'Any')
            reason = data.get('reason', '').strip()

            # Server-side validation
            missing_fields = []
            if not full_name: missing_fields.append("Full Name")
            if not index_number: missing_fields.append("Index Number")
            if not phone: missing_fields.append("Phone Number")
            if not email: missing_fields.append("Email")
            if not department: missing_fields.append("Department")
            if not programme: missing_fields.append("Programme")
            if not hall_of_residence: missing_fields.append("Hall of Residence")
            if not preferred_time or preferred_time == 'Click to set time': missing_fields.append("Preferred Time")

            if missing_fields:
                response = jsonify({'status': 'error', 'message': f"The following fields are required: {', '.join(missing_fields)}."})
                response.headers.add("Access-Control-Allow-Origin", "*")
                return response, 400

            ref = generate_booking_ref(conn)
            import uuid as _uuid
            booking_gid = str(_uuid.uuid4())
            conn.execute(
                '''INSERT INTO BookingRequest
                   (reference, full_name, index_number, department, programme, phone,
                    preferred_date, preferred_time, reason, status, email, hall_of_residence, global_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?, ?, ?)''',
                (ref, full_name, index_number, department, programme, phone,
                 preferred_date, preferred_time, reason, email, hall_of_residence, booking_gid)
            )
            conn.commit()
            from sync_engine import trigger_sync_immediate
            trigger_sync_immediate()

            # Fire in-app notification to all Desk Admins and Counsellors
            try:
                staff = conn.execute(
                    "SELECT id FROM users WHERE role IN ('Secretary', 'Admin', 'Counsellor', 'Counselor')"
                ).fetchall()
                for s in staff:
                    conn.execute(
                        '''INSERT INTO Notification (user_id, message, type, link, is_read)
                           VALUES (?, ?, 'in_app', '/admin/bookings', 0)''',
                        (s['id'],
                         f"New booking request {ref} from {full_name} ({index_number})")
                    )
                conn.commit()
            except Exception as notif_err:
                print(f"[API_BOOKING] Notification error: {notif_err}")

            conn.close()
            
            response = jsonify({'status': 'success', 'reference': ref, 'message': 'Booking submitted successfully!'})
            response.headers.add("Access-Control-Allow-Origin", "*")
            return response

        except Exception as e:
            print(f"[API_BOOKING] Error: {e}")
            response = jsonify({'status': 'error', 'message': 'Something went wrong. Please try again.'})
            response.headers.add("Access-Control-Allow-Origin", "*")
            return response, 500

@app.route('/booking', methods=['GET', 'POST'])
def booking_portal():
    """Public booking form — auto-accepted on submission, no staff approval needed."""
    if request.method == 'POST':
        try:
            conn = get_db_connection()
            full_name = request.form.get('full_name', '').strip()
            index_number = request.form.get('index_number', '').strip()
            programme_base = request.form.get('programme', '').strip()
            programme_other = request.form.get('programme_other', '').strip()
            programme = programme_other if programme_base == 'Other' else programme_base
            
            department = request.form.get('department', '').strip()
            phone = request.form.get('phone', '').strip()
            hall_of_residence = request.form.get('hall_of_residence', '').strip()
            preferred_date = request.form.get('preferred_date', '').strip()
            preferred_time = request.form.get('preferred_time', 'Any')
            reason = request.form.get('reason', '').strip()

            # Server-side validation
            missing_fields = []
            if not full_name:
                missing_fields.append("Full Name")
            if not index_number:
                missing_fields.append("Index Number")
            if not phone:
                missing_fields.append("Phone Number")
            if not department:
                missing_fields.append("Department")
            if not programme:
                missing_fields.append("Programme")
            if not preferred_time or preferred_time == 'Click to set time':
                missing_fields.append("Preferred Time")

            if missing_fields:
                return render_template('booking_portal.html',
                                       error=f"The following fields are required: {', '.join(missing_fields)}.")

            ref = generate_booking_ref(conn)
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            import uuid as _uuid
            booking_gid = str(_uuid.uuid4())
            # ── AUTO-ACCEPT: Insert as Accepted immediately ──
            try:
                conn.execute(
                    '''INSERT INTO BookingRequest
                       (reference, full_name, index_number, department, programme, phone,
                        preferred_date, preferred_time, reason, status, accepted_at, hall_of_residence, global_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Accepted', ?, ?, ?)''',
                    (ref, full_name, index_number, department, programme, phone,
                     preferred_date, preferred_time, reason, now_str, hall_of_residence, booking_gid)
                )
            except Exception:
                # Fallback if accepted_at column not yet present in older DBs
                conn.execute(
                    '''INSERT INTO BookingRequest
                       (reference, full_name, index_number, department, programme, phone,
                        preferred_date, preferred_time, reason, status, hall_of_residence, global_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Accepted', ?, ?)''',
                    (ref, full_name, index_number, department, programme, phone,
                     preferred_date, preferred_time, reason, hall_of_residence, booking_gid)
                )

            # Find or create the Student record
            student = conn.execute(
                "SELECT id FROM Student WHERE index_number = ?", (index_number,)
            ).fetchone()

            if not student:
                case_num = generate_case_number(conn, full_name)
                import uuid as _uuid
                stud_gid = str(_uuid.uuid4())
                conn.execute(
                    '''INSERT INTO Student (name, case_number, index_number, department,
                       programme, contact, created_at, hall_of_residence, global_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (full_name, case_num, index_number, department, programme, phone, now_str, hall_of_residence, stud_gid)
                )
                conn.commit()
                student = conn.execute(
                    "SELECT id FROM Student WHERE index_number = ?", (index_number,)
                ).fetchone()

            # Assign to first available counsellor if any
            counsellor = conn.execute("SELECT id FROM Counsellor LIMIT 1").fetchone()
            counsellor_id = counsellor['id'] if counsellor else None

            import uuid as _uuid
            appt_gid = str(_uuid.uuid4())
            conn.execute(
                '''INSERT INTO Appointment (student_id, Counsellor_id, date, time, purpose,
                   status, booking_ref, global_id)
                   VALUES (?, ?, ?, ?, ?, 'Scheduled', ?, ?)''',
                (student['id'], counsellor_id,
                 preferred_date or datetime.now().strftime('%Y-%m-%d'),
                 preferred_time or '09:00',
                 f"[Booked via Portal] {reason or 'Counselling session'}",
                 ref, appt_gid)
            )
            conn.commit()
            from sync_engine import trigger_sync_immediate
            trigger_sync_immediate()

            # Notify all staff of the new auto-accepted booking
            try:
                staff = conn.execute(
                    "SELECT id FROM users WHERE role IN ('Secretary', 'Admin', 'Counsellor', 'Counselor')"
                ).fetchall()
                for s in staff:
                    conn.execute(
                        '''INSERT INTO Notification (user_id, message, type, link, is_read)
                           VALUES (?, ?, 'in_app', '/admin/bookings', 0)''',
                        (s['id'],
                         f"🔔 New booking: {ref} — {full_name} ({index_number}) auto-accepted & scheduled")
                    )
                conn.commit()
            except Exception as notif_err:
                print(f"[BOOKING] Notification error: {notif_err}")

            conn.close()
            return redirect(url_for('booking_confirm', ref=ref))
        except Exception as e:
            print(f"[BOOKING] Error: {e}")
            import traceback
            traceback.print_exc()
            return render_template('booking_portal.html',
                                   error="Something went wrong. Please try again.")

    return render_template('booking_portal.html')



@app.route('/booking/confirm/<ref>')
def booking_confirm(ref):
    """Public confirmation page after booking is submitted."""
    return render_template('booking_confirmation.html', ref=ref)


@app.route('/admin/bookings')
@login_required
def admin_bookings():
    """Desk Admin & Counsellor view: list of all booking requests."""
    if session.get('role') not in ['Secretary', 'Admin', 'Counsellor', 'Counselor']:
        flash('Access restricted.', 'error')
        return redirect(url_for('dashboard'))
    try:
        page = request.args.get('page', 1, type=int)
        tab = request.args.get('tab', 'recent')
        per_page = 15
        offset = (page - 1) * per_page

        conn = get_db_connection()
        
        # Base queries - Separating by status (Pending = Recent, everything else = History)
        if tab == 'recent':
            # Only Pending bookings (new requests)
            where_clause = "WHERE LOWER(status) = 'pending' AND reference NOT IN (SELECT COALESCE(booking_ref, '') FROM Appointment WHERE booking_ref IS NOT NULL)"
            order_clause = "ORDER BY created_at DESC"
        elif tab == 'history':
            # Already Accepted, Declined or older processed ones
            where_clause = "WHERE LOWER(status) != 'pending' OR reference IN (SELECT COALESCE(booking_ref, '') FROM Appointment WHERE booking_ref IS NOT NULL)"
            order_clause = "ORDER BY COALESCE(accepted_at, created_at) DESC"
        else: # 'all'
            where_clause = ""
            order_clause = "ORDER BY created_at DESC"

        total = conn.execute(
            f"SELECT COUNT(*) FROM BookingRequest {where_clause}"
        ).fetchone()[0]
            
        bookings_raw = conn.execute(
            f"SELECT * FROM BookingRequest {where_clause} {order_clause} LIMIT ? OFFSET ?",
            (per_page, offset)
        ).fetchall()
        
        # Mark registry state for global sync (Unified Status Logic)
        bookings = []
        for b in bookings_raw:
            b_dict = dict(b)
            original_full_name = b_dict.get('full_name')
            # GTEC REQUIRED: Identification while preserving namevisibility
            clinical_id = get_clinical_id(original_full_name, b_dict.get('id'), b_dict.get('created_at'))
            b_dict['masked_name'] = name_to_initials(original_full_name)
            b_dict['clinical_id'] = clinical_id
            
            # GTEC Compliance: Hide original full name for view
            b_dict['full_name'] = b_dict['masked_name']
            
            # Check by index, original name, or email for safety
            exists = conn.execute(
                "SELECT 1 FROM Student WHERE (index_number IS NOT NULL AND index_number != '' AND index_number = ?) OR name = ? OR (email IS NOT NULL AND email != '' AND email = ?)",
                (b_dict.get('index_number'), original_full_name, b_dict.get('email', ''))
            ).fetchone()
            b_dict['is_registered'] = True if exists else False
            bookings.append(b_dict)
        total_pages = (total + per_page - 1) // per_page
        
        pending_count = conn.execute(
            "SELECT COUNT(*) FROM BookingRequest WHERE status = 'Pending'"
        ).fetchone()[0]
        
        # Also need count for the tabs
        recent_count = conn.execute(
            "SELECT COUNT(*) FROM BookingRequest WHERE status = 'Pending' AND reference NOT IN (SELECT COALESCE(booking_ref, '') FROM Appointment WHERE booking_ref IS NOT NULL)"
        ).fetchone()[0]
        
        conn.close()
        return render_template('admin_bookings.html', 
                               bookings=bookings,
                               pending_count=pending_count,
                               recent_count=recent_count,
                               role=session.get('role'),
                               page=page,
                               total_pages=total_pages,
                               current_tab=tab)
    except Exception as e:
        print(f"[ADMIN_BOOKINGS] Error: {e}")
        flash('Could not load bookings.', 'error')
        return redirect(url_for('dashboard'))


@app.route('/admin/bookings/<int:booking_id>/accept', methods=['POST'])
@login_required
def accept_booking(booking_id):
    """Accept a booking: create an Appointment record and mark as Accepted."""
    if session.get('role') not in ['Secretary', 'Admin', 'Counsellor', 'Counselor']:
        flash('Permission denied.', 'error')
        return redirect(url_for('admin_bookings'))
    try:
        conn = get_db_connection()
        booking = conn.execute(
            "SELECT * FROM BookingRequest WHERE id = ?", (booking_id,)
        ).fetchone()

        if not booking:
            flash('Booking not found.', 'error')
            conn.close()
            return redirect(url_for('admin_bookings'))

        # Check if already accepted
        if booking['status'] == 'Accepted':
            flash('This booking has already been accepted.', 'info')
            conn.close()
            return redirect(url_for('admin_bookings'))

        # Check if student exists by index_number first, then by original name
        student = conn.execute(
            "SELECT id FROM Student WHERE index_number = ?",
            (booking['index_number'],)
        ).fetchone()

        if not student:
            case_num = generate_case_number(conn, booking['full_name'])
            s_gid = str(uuid.uuid4())
            conn.execute(
                '''INSERT INTO Student (name, case_number, global_id, index_number, department, programme, contact, email, hall_of_residence)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (booking['full_name'], case_num, s_gid, booking['index_number'],
                 booking['department'], booking['programme'], booking['phone'], booking['email'], booking.get('hall_of_residence'))
            )
            conn.commit()
            student = conn.execute(
                "SELECT id FROM Student WHERE index_number = ?",
                (booking['index_number'],)
            ).fetchone()

        counsellor = conn.execute(
            "SELECT id FROM Counsellor LIMIT 1").fetchone()
        counsellor_id = counsellor['id'] if counsellor else None

        appt_gid = str(uuid.uuid4())
        conn.execute(
            '''INSERT INTO Appointment (student_id, Counsellor_id, date, time, purpose, status, booking_ref, global_id)
               VALUES (?, ?, ?, ?, ?, 'Scheduled', ?, ?)''',
            (student['id'], counsellor_id,
             booking['preferred_date'] or datetime.now().strftime('%Y-%m-%d'),
             booking['preferred_time'] or '09:00',
             f"[Booked via Portal] {booking['reason'] or 'Counselling session'}",
             booking['reference'], appt_gid)
        )
        conn.execute(
            "UPDATE BookingRequest SET status = 'Accepted', accepted_at = CURRENT_TIMESTAMP WHERE id = ?",
            (booking_id,)
        )
        conn.commit()
        conn.close()

        # Trigger sync to push changes to cloud
        try:
            from sync_engine import trigger_sync_immediate
            trigger_sync_immediate()
        except Exception:
            pass

        flash(
            f'Booking {booking["reference"]} accepted and appointment created.', 'success')
    except Exception as e:
        print(f"[ACCEPT_BOOKING] Error: {e}")
        flash(f'Error accepting booking: {str(e)}', 'error')
    return redirect(url_for('admin_bookings'))


@app.route('/admin/bookings/<int:booking_id>/decline', methods=['POST'])
@login_required
def decline_booking(booking_id):
    """Decline a booking with an optional reason."""
    if session.get('role') not in ['Secretary', 'Admin', 'Counsellor']:
        flash('Permission denied.', 'error')
        return redirect(url_for('admin_bookings'))
    try:
        reason = request.form.get('decline_reason', '').strip()
        conn = get_db_connection()
        conn.execute(
            "UPDATE BookingRequest SET status = 'Declined', decline_reason = ? WHERE id = ?",
            (reason, booking_id)
        )
        conn.commit()
        conn.close()
        flash('Booking declined.', 'info')
    except Exception as e:
        print(f"[DECLINE_BOOKING] Error: {e}")
        flash(f'Error declining booking: {str(e)}', 'error')
    return redirect(url_for('admin_bookings'))


if __name__ == '__main__':
    # Initialize database FIRST before anything else
    print("Initializing database...")
    try:
        # Force database initialization at startup with serverless path awareness
        try:
            import db_setup as _db_setup_mod
            db_path = _db_setup_mod.get_db_path()
        except Exception:
            import tempfile
            if os.environ.get('VERCEL') or os.environ.get('AWS_LAMBDA_FUNCTION_NAME'):
                db_path = os.path.join(tempfile.gettempdir(), 'counseling.db')
            elif getattr(sys, 'frozen', False):
                db_path = os.path.join(os.path.dirname(sys.executable), 'counseling.db')
            else:
                db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'counseling.db')

        # Check and initialize
        if not os.path.exists(db_path):
            print(f"Database not found, creating at: {db_path}")
            import db_setup
            db_setup.init_db()
        else:
            # Verify Appointment table exists
            test_conn = sqlite3.connect(db_path)
            cursor = test_conn.cursor()
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='Appointment'")
            if not cursor.fetchone():
                print("Appointment table missing, reinitializing database...")
                test_conn.close()
                import db_setup
                db_setup.init_db()
            else:
                test_conn.close()
                print("Database check passed - Appointment table exists")
    except Exception as e:
        print(f"WARNING: Database initialization issue: {e}")
        print("Attempting to continue anyway...")

    # Check if port 5050 is already in use and kill the process if needed
    import socket
    import subprocess
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('127.0.0.1', 5050))
        sock.close()
    except OSError as e:
        # Port already in use
        if e.errno == 10048 or (hasattr(e, 'winerror') and e.winerror == 10048):
            print("WARNING: Port 5050 is already in use!")
            print("Attempting to kill the process using port 5050...")
            try:
                # Find process using port 5050
                result = subprocess.run(
                    ['netstat', '-ano'], capture_output=True, text=True, timeout=5)
                for line in result.stdout.split('\n'):
                    if ':5050' in line and 'LISTENING' in line:
                        parts = line.split()
                        if len(parts) > 4:
                            pid = parts[-1]
                            print(
                                f"Found process {pid} using port 5050. Killing it...")
                            subprocess.run(['taskkill', '/F', '/PID', pid],
                                           capture_output=True, timeout=5)
                            import time
                            time.sleep(1)  # Wait a bit for port to be released
                # Try binding again
                try:
                    sock2 = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock2.bind(('127.0.0.1', 5050))
                    sock2.close()
                    print("Port 5050 is now available!")
                except:
                    print("ERROR: Could not free port 5050.")
                    print(
                        "Please manually close other instances or restart your computer.")
                    is_exe = getattr(sys, 'frozen', False)
                    if not is_exe:
                        input("Press Enter to exit...")
                    sys.exit(1)
            except Exception as kill_error:
                print(f"Could not kill process: {kill_error}")
                print("Please manually close other instances.")
                is_exe = getattr(sys, 'frozen', False)
                if not is_exe:
                    input("Press Enter to exit...")
                    sys.exit(1)

    # Log available routes for debugging
    print('=' * 60)
    print('USTED Counselling Management System')
    print('=' * 60)
    print('Starting server on http://127.0.0.1:5050')
    print('Registered routes:')
    for rule in app.url_map.iter_rules():
        if rule.endpoint not in ['static']:
            print(f"  - {rule.endpoint}: {rule}")
    print('=' * 60)
    print()

    # For EXE, don't use debug mode and open browser automatically
    import webbrowser
    import threading

    def open_browser():
        import time
        time.sleep(2.5)  # Wait longer for server to fully start
        try:
            # Try to open browser
            webbrowser.open('http://127.0.0.1:5050')
            print("Browser opened automatically!")
        except Exception as e:
            print(f"Could not open browser automatically: {e}")
            print("Please manually open: http://localhost:5050")

    # Open browser automatically (only if not in debug mode or forced via Env)
    is_exe = getattr(sys, 'frozen', False) or os.environ.get(
        'USTED_AUTO_OPEN_BROWSER') == '1'
    if is_exe:
        browser_thread = threading.Thread(target=open_browser)
        browser_thread.daemon = True
        browser_thread.start()
        print("Server starting... Browser will open automatically.")
    else:
        print("Starting in development mode...")
        print("Local access: http://localhost:5050")
        print("Network access: http://<your-ip-address>:5050")

    print()

    try:
        # Run app (debug=False for production EXE)
        app.run(debug=not getattr(sys, 'frozen', False),
                host='0.0.0.0', port=5050, use_reloader=False)
    except Exception as e:
        print(f"Error starting server: {e}")
        input("Press Enter to exit...")
        sys.exit(1)
