from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Response, send_file, make_response
from functools import wraps
import sqlite3
import csv
import io
import json
from datetime import datetime, timedelta
import os
import sys
from werkzeug.security import check_password_hash
from auto_report_writer import scheduler, toggle_scheduler, manual_generate_report

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'

# Global flag to prevent concurrent initialization
_db_initialization_lock = False
_db_initialized = False

def ensure_database_initialized():
    """Ensure database exists and has all required tables - atomic, single-call initialization"""
    global _db_initialization_lock, _db_initialized
    
    # If already initialized successfully, skip
    if _db_initialized:
        return
    
    # If currently initializing, wait (prevent concurrent calls)
    if _db_initialization_lock:
        import time
        wait_count = 0
        while _db_initialization_lock and wait_count < 50:  # Wait up to 5 seconds
            time.sleep(0.1)
            wait_count += 1
        if _db_initialized:
            return
    
    # Set lock to prevent concurrent initialization
    _db_initialization_lock = True
    
    try:
        # Determine database path
        try:
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
        except:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        db_path = os.path.join(base_path, 'counseling.db')
        
        # Required tables for the application
        required_tables = ['Appointment', 'Student', 'Counsellor', 'session', 'app_settings', 'Referral']
        
        # Check if database exists and has all required tables
        needs_init = False
        
        if not os.path.exists(db_path):
            print(f"[STARTUP] Database not found at: {db_path}")
            needs_init = True
        else:
            # Check for all required tables in a single connection
            try:
                check_conn = sqlite3.connect(db_path, timeout=5.0)
                cursor = check_conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                existing_tables = {row[0].lower() for row in cursor.fetchall()}
                check_conn.close()
                
                # Check if all required tables exist (case-insensitive)
                missing_tables = []
                for req_table in required_tables:
                    if req_table.lower() not in existing_tables:
                        missing_tables.append(req_table)
                
                if missing_tables:
                    print(f"[STARTUP] Missing tables: {missing_tables}")
                    needs_init = True
                else:
                    print(f"[STARTUP] Database check passed - all required tables exist")
                    _db_initialized = True
                    _db_initialization_lock = False
                    return
            except Exception as e:
                print(f"[STARTUP] Error checking database: {e}")
                needs_init = True
        
        # Initialize database if needed (only once, atomically)
        if needs_init:
            print(f"[STARTUP] Initializing database at: {db_path}")
            try:
                # Import and run initialization (uses same path logic)
                import db_setup
                db_setup.init_db()
                
                # Verify initialization succeeded - check for Appointment table
                verify_conn = sqlite3.connect(db_path, timeout=5.0)
                verify_cursor = verify_conn.cursor()
                verify_cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                verify_tables = {row[0].lower() for row in verify_cursor.fetchall()}
                verify_conn.close()
                
                # Check if all required tables now exist (case-insensitive check)
                missing_after_init = []
                for req_table in required_tables:
                    if req_table.lower() not in verify_tables:
                        missing_after_init.append(req_table)
                
                if missing_after_init:
                    print(f"[STARTUP] ERROR: Initialization incomplete! Missing tables: {missing_after_init}")
                    print(f"[STARTUP] Available tables: {verify_tables}")
                    raise Exception(f"Database initialization incomplete. Missing tables: {missing_after_init}")
                else:
                    print("[STARTUP] Database initialized successfully - all required tables verified")
                    _db_initialized = True
                    
            except Exception as e:
                print(f"[STARTUP] ERROR initializing database: {e}")
                import traceback
                traceback.print_exc()
                # Don't set _db_initialized = True so it can retry
                raise
    
    finally:
        _db_initialization_lock = False

# Initialize database when module is loaded
ensure_database_initialized()

@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()}

# Add custom Jinja2 filters
@app.template_filter('nl2br')
def nl2br(value):
    if value:
        return value.replace('\n', '<br>')
    return ''

# ---------- Helper Functions ----------
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

def get_db_connection():
    """Get database connection - works in both dev and EXE mode"""
    # Ensure database is initialized first (only once)
    ensure_database_initialized()
    
    # Get database path
    try:
        if getattr(sys, 'frozen', False):
            # Running as compiled EXE
            base_path = os.path.dirname(sys.executable)
        else:
            # Running as script
            base_path = os.path.dirname(os.path.abspath(__file__))
    except:
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    db_path = os.path.join(base_path, 'counseling.db')
    
    # Connect with timeout to prevent locking issues
    conn = sqlite3.connect(db_path, timeout=10.0)
    conn.row_factory = sqlite3.Row
    
    # Quick verification that Student table exists (most critical table)
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND (name='Student' OR name='student')")
        if not cursor.fetchone():
            print("[GET_DB_CONNECTION] Student table missing! Reinitializing...")
            conn.close()
            # Force reinitialization
            global _db_initialized
            _db_initialized = False
            ensure_database_initialized()
            # Reconnect
            conn = sqlite3.connect(db_path, timeout=10.0)
            conn.row_factory = sqlite3.Row
    except Exception as verify_error:
        print(f"[GET_DB_CONNECTION] Error verifying tables: {verify_error}")
        # Continue anyway - let the query fail and be caught by route handlers
    
    return conn

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('welcome'))
        return f(*args, **kwargs)
    return decorated_function

# ---------- Routes ----------
@app.route('/')
def home():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('welcome'))

@app.route('/welcome', methods=['GET'])
def welcome():
    try:
        # Check if user is already logged in - if so, redirect to dashboard
        if session.get('logged_in'):
            # Check if already visited today
            last_visit_date = session.get('last_visit_date')
            today = datetime.now().date().isoformat()
            
            # If already visited today, go to dashboard
            if last_visit_date == today:
                return redirect(url_for('dashboard'))
        
        # Ensure database is initialized before showing welcome page
        try:
            ensure_database_initialized()
        except Exception as e:
            print(f"[WELCOME] Database init error: {e}")
            # Still show welcome page, but log the error
        
        # Show welcome screen (for login page)
        return render_template('welcome.html')
    except Exception as e:
        print(f"[WELCOME] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        return f'''
        <html>
        <head><title>Error</title></head>
        <body>
            <h1>Error Loading Welcome Page</h1>
            <p>An error occurred: {str(e)}</p>
            <p>Check error_log.txt in the application folder for details.</p>
        </body>
        </html>
        ''', 500

@app.route('/login', methods=['POST'])
def login():
    try:
        password = request.form.get('password', '')
        
        # Ensure database is initialized before connection
        try:
            ensure_database_initialized()
        except Exception as db_init_error:
            print(f"[LOGIN] Database init error: {db_init_error}")
            flash('Database initialization error. Please restart the application.', 'error')
            return redirect(url_for('welcome'))
        
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('welcome'))
        
        stored_hash = None
        try:
            # Check if app_settings table exists first
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='app_settings'")
            if not cursor.fetchone():
                # Table doesn't exist - initialize database
                print("[LOGIN] app_settings table missing, reinitializing database...")
                conn.close()
                ensure_database_initialized()
                conn = get_db_connection()
            
            row = conn.execute(
                "SELECT setting_value FROM app_settings WHERE setting_name = 'password_hash'"
            ).fetchone()
            stored_hash = row['setting_value'] if row else None
            if stored_hash is None:
                # Auto-create default password hash if missing
                from werkzeug.security import generate_password_hash
                default_hash = generate_password_hash('Counsellor123')
                try:
                    conn.execute(
                        "INSERT INTO app_settings (setting_name, setting_value) VALUES (?, ?)",
                        ('password_hash', default_hash)
                    )
                    conn.commit()
                    stored_hash = default_hash
                except Exception as insert_error:
                    print(f"[LOGIN] Error creating password hash: {insert_error}")
                    # Try to rollback and continue
                    conn.rollback()
                    stored_hash = default_hash
        except Exception as db_error:
            print(f"[LOGIN] Database query error: {db_error}")
            import traceback
            traceback.print_exc()
            try:
                conn.close()
            except:
                pass
            # Try one more time to initialize database
            try:
                ensure_database_initialized()
                conn = get_db_connection()
                row = conn.execute(
                    "SELECT setting_value FROM app_settings WHERE setting_name = 'password_hash'"
                ).fetchone()
                stored_hash = row['setting_value'] if row else None
            except:
                flash('Database error. Please restart the application.', 'error')
                return redirect(url_for('welcome'))
        finally:
            try:
                conn.close()
            except:
                pass
    except Exception as e:
        print(f"[LOGIN] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Login error: {str(e)}. Please check console for details.', 'error')
        return redirect(url_for('welcome'))

    # Debug: Print password attempt (without revealing it)
    print(f"[LOGIN] Password attempt received. Hash exists: {stored_hash is not None}")
    
    if stored_hash:
        password_valid = check_password_hash(stored_hash, password)
        print(f"[LOGIN] Password check result: {password_valid}")
        
        if password_valid:
            session['logged_in'] = True
            session.permanent = True
            
            # Check if this is first visit today BEFORE setting the date
            today = datetime.now().date().isoformat()
            last_visit = session.get('last_visit_date')
            is_first_visit_today = (last_visit != today)
            
            # Mark today's visit (this will make next check show False)
            session['last_visit_date'] = today
            session['first_visit_today'] = is_first_visit_today
            
            return redirect(url_for('dashboard'))
        else:
            # Password was wrong
            print(f"[LOGIN] Invalid password - hash exists but doesn't match")
            flash('Invalid password. Default is: Counsellor123 (Capital C, double L).', 'error')
    else:
        # No password hash in database - this shouldn't happen
        print(f"[LOGIN] ERROR: No password hash found in database!")
        flash('Database error: No password configured. Please restart the application.', 'error')
    
    return redirect(url_for('welcome'))

@app.route('/dashboard')
@login_required
def dashboard():
    try:
        # Check if this is first visit today using the flag set during login
        # The login() function sets this flag correctly before setting last_visit_date
        if 'first_visit_today' in session:
            # Use the flag set during login
            show_welcome_message = session.pop('first_visit_today', False)
            # Ensure last_visit_date is set (in case it wasn't)
            if 'last_visit_date' not in session or session.get('last_visit_date') != datetime.now().date().isoformat():
                session['last_visit_date'] = datetime.now().date().isoformat()
        else:
            # Fallback: Check date if flag wasn't set (shouldn't happen normally)
            today = datetime.now().date().isoformat()
            last_visit = session.get('last_visit_date')
            if last_visit != today:
                session['last_visit_date'] = today
                show_welcome_message = True
            else:
                show_welcome_message = False
        
        # Ensure database initialized
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('welcome'))
        
        # Initialize all variables with defaults
        counsellor_name = 'Mrs. Gertrude Effeh Brew'
        total_students = 0
        total_appointments = 0
        today_appointments = 0
        total_sessions = 0
        this_month_sessions = 0
        pending_appointments = 0
        completed_this_month = 0
        total_referrals = 0
        updated_count = 0
        today_appts = []
        appointments = []
        sessions = []
        referrals = []
        
        # Get counsellor name (Mrs. Gertrude Effeh Brew)
        try:
            counsellor = conn.execute('SELECT name FROM Counsellor WHERE id = 1').fetchone()
            if counsellor:
                counsellor_name = counsellor['name']
        except Exception as e:
            print(f"[DASHBOARD] Error getting counsellor: {e}")
    
    # Statistics: Total Students
        try:
            result = conn.execute('SELECT COUNT(*) as count FROM Student').fetchone()
            total_students = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting total_students: {e}")
            total_students = 0
    
    # Statistics: Total Appointments
        try:
            result = conn.execute('SELECT COUNT(*) as count FROM Appointment').fetchone()
            total_appointments = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting total_appointments: {e}")
            total_appointments = 0
    
    # Statistics: Today's Appointments
        try:
            result = conn.execute('''
        SELECT COUNT(*) as count 
        FROM Appointment 
        WHERE date = DATE('now')
            ''').fetchone()
            today_appointments = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting today_appointments: {e}")
            today_appointments = 0
    
    # Statistics: Total Sessions
        try:
            result = conn.execute('SELECT COUNT(*) as count FROM session').fetchone()
            total_sessions = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting total_sessions: {e}")
            total_sessions = 0
    
    # Statistics: This Month's Sessions
        try:
            result = conn.execute('''
        SELECT COUNT(*) as count 
        FROM session 
                WHERE strftime('%Y-%m', created_at)  = strftime('%Y-%m', 'now')
            ''').fetchone()
            this_month_sessions = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting this_month_sessions: {e}")
            this_month_sessions = 0
    
    # Statistics: Pending Appointments
        try:
            result = conn.execute('''
        SELECT COUNT(*) as count 
        FROM Appointment 
        WHERE status IN ('scheduled', 'Scheduled') 
        AND date >= DATE('now')
            ''').fetchone()
            pending_appointments = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting pending_appointments: {e}")
            pending_appointments = 0
    
    # Statistics: Completed Sessions This Month
        try:
            result = conn.execute('''
        SELECT COUNT(*) as count 
        FROM Appointment 
        WHERE status IN ('Completed', 'completed')
        AND date >= date('now', 'start of month')
            ''').fetchone()
            completed_this_month = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting completed_this_month: {e}")
            completed_this_month = 0
    
        # Statistics: Total Referrals
        try:
            result = conn.execute('SELECT COUNT(*) as count FROM Referral').fetchone()
            total_referrals = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting total_referrals: {e}")
            total_referrals = 0
        
        # Get updated_count for auto-update banner
        try:
            result = conn.execute('''
                SELECT COUNT(*) as count 
                FROM Appointment 
                WHERE status = 'In Session' 
                AND date = DATE('now')
            ''').fetchone()
            updated_count = result['count'] if result else 0
        except Exception as e:
            print(f"[DASHBOARD] Error getting updated_count: {e}")
            updated_count = 0
    
        # Get today's appointments
        try:
            today_appts_raw = conn.execute('''
                SELECT a.id, a.date, a.time, a.purpose, a.status,
                       s.id as student_db_id, s.name as student_name,
                       c.name as counsellor_name
                FROM Appointment a
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE a.date = DATE('now')
                ORDER BY a.time
            ''').fetchall()
            # Convert to list and add professional IDs
            today_appts = []
            for apt in today_appts_raw:
                apt_dict = dict(apt)
                student_db_id = apt_dict.get('student_db_id', 0)
                apt_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                today_appts.append(apt_dict)
        except Exception as e:
            print(f"[DASHBOARD] Error getting today_appts: {e}")
            today_appts = []
    
        # Get upcoming appointments (next 7 days)
        try:
            appointments_raw = conn.execute('''
                SELECT a.id, a.date, a.time, a.purpose, a.status,
                       s.id as student_db_id, s.name as student_name,
                       c.name as counsellor_name
                FROM Appointment a
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE a.date >= DATE('now') AND a.date <= DATE('now', '+7 days')
                ORDER BY a.date, a.time
                LIMIT 10
            ''').fetchall()
            # Convert to list and add professional IDs
            appointments = []
            for apt in appointments_raw:
                apt_dict = dict(apt)
                student_db_id = apt_dict.get('student_db_id', 0)
                apt_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                appointments.append(apt_dict)
        except Exception as e:
            print(f"[DASHBOARD] Error getting appointments: {e}")
            appointments = []
    
        # Get recent sessions
        try:
            sessions_raw = conn.execute('''
                SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
                       s.id as student_db_id, s.name as student_name
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                ORDER BY sess.created_at DESC
                LIMIT 5
            ''').fetchall()
            # Convert to list and add professional IDs
            sessions = []
            for sess in sessions_raw:
                sess_dict = dict(sess)
                student_db_id = sess_dict.get('student_db_id', 0)
                sess_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                sessions.append(sess_dict)
        except Exception as e:
            print(f"[DASHBOARD] Error getting sessions: {e}")
            sessions = []
    
        # Get recent referrals - handle missing table gracefully
        try:
            referrals_raw = conn.execute('''
            SELECT r.id, r.referred_by, r.reasons, r.created_at,
                   s.id as student_db_id, s.name as student_name
            FROM Referral r
            JOIN session sess ON r.session_id = sess.id
            JOIN Appointment a ON sess.appointment_id = a.id
            JOIN Student s ON a.student_id = s.id
            ORDER BY r.created_at DESC
            LIMIT 5
        ''').fetchall()
            # Convert to list and add professional IDs
            referrals = []
            for ref in referrals_raw:
                ref_dict = dict(ref)
                student_db_id = ref_dict.get('student_db_id', 0)
                ref_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                referrals.append(ref_dict)
        except Exception as e:
            print(f"[DASHBOARD] Error getting referrals: {e}")
            referrals = []
    
        # Get appointment statistics by status for chart - with error handling
        try:
            status_stats = conn.execute('''
                SELECT 
                    COUNT(CASE WHEN status IN ('scheduled', 'Scheduled') THEN 1 END) as scheduled_count,
                    COUNT(CASE WHEN status IN ('In Session', 'in session') THEN 1 END) as in_session_count,
                    COUNT(CASE WHEN status IN ('Completed', 'completed') THEN 1 END) as completed_count,
                    COUNT(CASE WHEN status IN ('Cancelled', 'cancelled') THEN 1 END) as cancelled_count
                FROM Appointment
                WHERE date >= DATE('now', '-30 days')
            ''').fetchone()
            # Ensure status_stats has default values
            if not status_stats:
                status_stats = {
                    'scheduled_count': 0,
                    'in_session_count': 0,
                    'completed_count': 0,
                    'cancelled_count': 0
                }
        except Exception as e:
            print(f"[DASHBOARD] Error getting status stats: {e}")
            status_stats = {
                'scheduled_count': 0,
                'in_session_count': 0,
                'completed_count': 0,
                'cancelled_count': 0
            }

        # Generate greeting based on time and visit status
        current_hour = datetime.now().hour
        if current_hour < 12:
            time_greeting = "Good morning"
        elif current_hour < 17:
            time_greeting = "Good afternoon"
        else:
            time_greeting = "Good evening"

        # Create personalized greeting
        if show_welcome_message:
            greeting = f"{time_greeting}, {counsellor_name}, Welcome Back Counsellor! 👋"
        else:
            greeting = f"{time_greeting}, {counsellor_name}"

        # Ensure connection is closed before rendering
        try:
            conn.close()
        except:
            pass

        # Pass all template variables
        return render_template('dashboard.html', 
                             updated_count=updated_count,
                             appointments=appointments,
                             today_appts=today_appts,
                             sessions=sessions,
                             referrals=referrals,
                             greeting=greeting,
                             counsellor_name=counsellor_name,
                             show_welcome_message=show_welcome_message,
                             total_students=total_students,
                             total_appointments=total_appointments,
                             today_appointments=today_appointments,
                             total_sessions=total_sessions,
                             this_month_sessions=this_month_sessions,
                             pending_appointments=pending_appointments,
                             completed_this_month=completed_this_month,
                             total_referrals=total_referrals,
                             status_stats=status_stats)
    
    except Exception as e:
        print(f"[DASHBOARD] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        try:
            conn.close()
        except:
            pass
        flash('Error loading dashboard. Please try refreshing the page.', 'error')
        return redirect(url_for('welcome'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.')
    return redirect(url_for('welcome'))

@app.route('/add_student', methods=['GET', 'POST'])
@login_required
def add_student():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('students'))

        if request.method == 'POST':
            edit_id = request.form.get('edit_id')

            name = request.form.get('name')
            age = request.form.get('age')
            gender = request.form.get('gender')
            index_number = request.form.get('index_number')
            department = request.form.get('department')
            programme = request.form.get('programme')
            contact = request.form.get('contact')
            parent_contact = request.form.get('parent_contact')
            hall_of_residence = request.form.get('hall_of_residence')
            faculty = request.form.get('faculty', '')

            try:
                if edit_id:
                    conn.execute('''
                        UPDATE Student 
                        SET name=?, age=?, gender=?, index_number=?, department=?, faculty=?, 
                            programme=?, contact=?, parent_contact=?, hall_of_residence=?
                        WHERE id=?
                    ''', (name, age if age else None, gender, index_number, department, faculty, 
                          programme, contact, parent_contact, hall_of_residence, edit_id))
                    flash('Student updated successfully!', 'success')
                else:
                    conn.execute(
                        'INSERT INTO Student (name, age, gender, index_number, department, faculty, programme, contact, parent_contact, hall_of_residence) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                        (name, age if age else None, gender, index_number, department, faculty, programme, contact, parent_contact, hall_of_residence)
                    )
                    flash('Student added successfully!', 'success')
                conn.commit()
                return redirect(url_for('students'))
            except sqlite3.IntegrityError:
                conn.rollback()
                flash('Error: Index number already exists.', 'error')
                if edit_id:
                    try:
                        student = conn.execute('SELECT * FROM Student WHERE id = ?', (edit_id,)).fetchone()
                        return render_template('add_student.html', student=student)
                    except Exception:
                        pass
            except Exception as e:
                conn.rollback()
                print(f"[ADD_STUDENT] Error saving student: {e}")
                import traceback
                traceback.print_exc()
                # Check if it's a table missing error and reinitialize
                if 'no such table' in str(e).lower() or 'Student' in str(e):
                    print("[ADD_STUDENT] Table missing error detected, reinitializing database...")
                    try:
                        ensure_database_initialized()
                        conn = get_db_connection()
                        flash('Database was reinitialized. Please try again.', 'info')
                    except Exception as init_error:
                        print(f"[ADD_STUDENT] Error reinitializing: {init_error}")
                        flash(f'Database error: {str(e)}. Please restart the application.', 'error')
                else:
                    flash(f'Error saving student: {str(e)}', 'error')
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        else:
            student = None
            edit_id = request.args.get('edit')
            if edit_id:
                try:
                    student = conn.execute('SELECT * FROM Student WHERE id = ?', (edit_id,)).fetchone()
                except Exception as e:
                    print(f"[ADD_STUDENT] Error loading student for edit: {e}")
                    import traceback
                    traceback.print_exc()
                    # Check if it's a table missing error and reinitialize
                    if 'no such table' in str(e).lower() or 'Student' in str(e):
                        print("[ADD_STUDENT] Table missing error detected in GET, reinitializing database...")
                        try:
                            ensure_database_initialized()
                            conn = get_db_connection()
                            student = conn.execute('SELECT * FROM Student WHERE id = ?', (edit_id,)).fetchone()
                        except Exception as init_error:
                            print(f"[ADD_STUDENT] Error reinitializing: {init_error}")
                            student = None
                    else:
                        student = None
            try:
                conn.close()
            except Exception:
                pass
            return render_template('add_student.html', student=student)

    except Exception as e:
        print(f"[ADD_STUDENT] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading student form. Please try again.', 'error')
        return redirect(url_for('students'))

@app.route('/sessions')
@login_required
def sessions_list():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        sessions_raw = []
        try:
            # Get sessions with full details including student, counsellor, and appointment info
            sessions_raw = conn.execute('''
                SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
                       s.id as student_db_id, s.name as student_name,
                       c.name as Counsellor_name,
                       a.date, a.time, a.status,
                       sess.appointment_id
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                ORDER BY sess.created_at DESC
            ''').fetchall()
        except Exception as e:
            print(f"[SESSIONS] Error getting sessions: {e}")
            sessions_raw = []
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        # Convert to list and add professional IDs
        sessions = []
        for sess in sessions_raw:
            sess_dict = dict(sess)
            student_db_id = sess_dict.get('student_db_id', 0)
            sess_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
            sessions.append(sess_dict)
        
        # Create a simple pagination object to prevent template errors
        class SimplePagination:
            has_prev = False
            has_next = False
            page = 1
            prev_num = None
            next_num = None
            def iter_pages(self):
                return []
        
        pagination = SimplePagination()
        return render_template('sessions.html', sessions=sessions, pagination=pagination)
    except Exception as e:
        print(f"[SESSIONS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading sessions. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/create_session', methods=['GET', 'POST'])
@login_required
def create_session():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            appointment_id = request.form.get('appointment_id')
            session_type = request.form.get('session_type')
            notes = request.form.get('notes')
            outcome = request.form.get('outcome', '')

            try:
                # Get appointment details
                appointment = conn.execute('''
                    SELECT student_id, status 
                    FROM Appointment 
                    WHERE id = ?
                ''', (appointment_id,)).fetchone()

                if appointment:
                    student_id = appointment['student_id']
                    appointment_status = appointment['status']

                    # Insert new session - use appointment_id as the foreign key
                    conn.execute('''
                        INSERT INTO session (appointment_id, session_type, notes, outcome, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (appointment_id, session_type, notes, outcome, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

                    # Update appointment status to completed only if it's currently scheduled
                    if appointment_status == 'scheduled':
                        conn.execute('UPDATE Appointment SET status = ? WHERE id = ?', 
                                     ('Completed', appointment_id))

                    conn.commit()
                    flash('Session created successfully!')
                    return redirect(url_for('sessions_list'))
                else:
                    flash('Invalid appointment selected.')
                    return redirect(url_for('create_session'))
            except Exception as e:
                conn.rollback()
                print(f"[CREATE_SESSION] Error creating session: {e}")
                flash(f'Error creating session: {str(e)}', 'error')
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
            return redirect(url_for('create_session'))

        # GET: load appointments for dropdown
        appointments = []
        try:
            appointments = conn.execute('''
                SELECT a.id, a.date as date, a.time as time, a.status, s.name as student_name,
                       c.name as counsellor_name
                FROM Appointment a
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE a.status IN ('scheduled', 'Scheduled', 'Completed', 'completed')
                ORDER BY a.date DESC, a.time DESC
            ''').fetchall()
        except Exception as e:
            print(f"[CREATE_SESSION] Error getting appointments: {e}")
            appointments = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('create_session.html', appointments=appointments)
    except Exception as e:
        print(f"[CREATE_SESSION] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading session creation page. Please try again.', 'error')
        return redirect(url_for('dashboard'))
@app.route('/case_note', methods=['GET', 'POST'])
@login_required
def case_note():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            session_id = request.form.get('session_id')
            client_appearance = request.form.get('client_appearance', '')
            problems = request.form.get('problems', '')
            interventions = request.form.get('interventions', '')
            recommendations = request.form.get('recommendations', '')
            next_visit_date = request.form.get('next_visit_date') or None
            counsellor_signature = request.form.get('counsellor_signature', '')

            # Validate required fields
            if not session_id:
                flash('Please select a session', 'error')
                return redirect(url_for('case_note'))
            
            if not all([client_appearance, problems, interventions, recommendations]):
                flash('Please fill in all required fields', 'error')
                return redirect(url_for('case_note'))

            try:
                # Check if case management record already exists for this session
                existing = conn.execute('SELECT id FROM CaseManagement WHERE session_id = ?', (session_id,)).fetchone()
                
                if existing:
                    # Update existing record
                    conn.execute('''
                        UPDATE CaseManagement 
                        SET client_appearance = ?, problems = ?, interventions = ?, recommendations = ?,
                            next_visit_date = ?, counsellor_signature = ?
                        WHERE session_id = ?
                    ''', (client_appearance, problems, interventions, recommendations, 
                          next_visit_date, counsellor_signature, session_id))
                else:
                    # Insert new record
                    conn.execute('''
                        INSERT INTO CaseManagement 
                        (session_id, client_appearance, problems, interventions, recommendations, 
                         next_visit_date, counsellor_signature, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (session_id, client_appearance, problems, interventions, recommendations,
                          next_visit_date, counsellor_signature, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                
                conn.commit()
                flash('Case notes saved successfully!', 'success')
                return redirect(url_for('sessions_list'))
            except Exception as e:
                conn.rollback()
                print(f"[CASE_NOTE] Error saving case notes: {e}")
                import traceback
                traceback.print_exc()
                flash(f'Error saving case notes: {str(e)}', 'error')
                return redirect(url_for('case_note'))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - show form
        sessions = []
        try:
            sessions_raw = conn.execute('''
                SELECT s.id, st.id as student_db_id, st.name as student_name, s.created_at
                FROM session s
                JOIN Appointment a ON s.appointment_id = a.id
                JOIN Student st ON a.student_id = st.id
                ORDER BY s.created_at DESC
            ''').fetchall()
            # Convert to list and add professional IDs
            sessions = []
            for sess in sessions_raw:
                sess_dict = dict(sess)
                student_db_id = sess_dict.get('student_db_id', 0)
                sess_dict['professional_id'] = f"C{student_db_id:03d}"
                sessions.append(sess_dict)
        except Exception as e:
            print(f"[CASE_NOTE] Error getting sessions: {e}")
            sessions = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('case_note.html', sessions=sessions, now=datetime.utcnow())
    except Exception as e:
        print(f"[CASE_NOTE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case notes page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

# ---------- Reports Routes ----------
@app.route('/reports')
@login_required
def reports_list():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        reports = []
        try:
            # Get all generated reports from the database
            reports = conn.execute('''
                SELECT id, title, date_generated, report_type, summary, file_path
                FROM reports
                ORDER BY date_generated DESC
            ''').fetchall()
        except Exception as e:
            print(f"[REPORTS] Error getting reports: {e}")
            reports = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('reports.html', reports=reports)
    except Exception as e:
        print(f"[REPORTS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading reports. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/generate_report_manual', methods=['POST'], endpoint='generate_report_manual')
@login_required
def generate_report_manual():
    """Generate a report manually with custom options"""
    report_type = request.form.get('report_type', 'manual')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    
    try:
        if report_type == 'custom' and start_date and end_date:
            # Custom date range - modify the generate_report function to accept dates
            manual_generate_report()  # For now, use manual generation
            flash('Custom report generation is being prepared. Report generated successfully!', 'success')
        else:
            manual_generate_report()
            flash('Report generated successfully!', 'success')
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'error')
    
    return redirect(url_for('reports_list'))

@app.route('/toggle_auto_report', methods=['GET', 'POST'])
@login_required
def toggle_auto_report():
    """Toggle auto report generation on/off"""
    if request.method == 'POST':
        data = request.get_json()
        enable = data.get('enable', False)
        
        try:
            toggle_scheduler(enable)
            return jsonify({
                'status': 'success',
                'message': f'Auto report generation {"enabled" if enable else "disabled"} successfully.',
                'is_enabled': enable
            })
        except Exception as e:
            return jsonify({
                'status': 'error',
                'message': str(e)
            }), 500
    else:
        # GET request - return current status
        is_running = scheduler.running if scheduler else False
        return jsonify({
            'status': 'success',
            'is_enabled': is_running
        })

@app.route('/generate_report_now', methods=['POST'])
@login_required
def generate_report_now():
    """Manually trigger report generation"""
    try:
        manual_generate_report()
        return jsonify({
            'status': 'success',
            'message': 'Report generated successfully! Check the Reports page to view it.'
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error generating report: {str(e)}'
        }), 500

@app.route('/download_report_file/<int:report_id>')
@login_required
def download_report_file(report_id):
    """Download the actual report file (DOCX)"""
    conn = get_db_connection()
    
    report = conn.execute('SELECT * FROM reports WHERE id = ?', (report_id,)).fetchone()
    conn.close()
    
    if not report:
        flash('Report not found', 'error')
        return redirect(url_for('reports_list'))
    
    # Convert Row to dict for easier access
    report_dict = dict(report)
    file_path = report_dict.get('file_path')
    
    if not file_path:
        flash('Report file path not found', 'error')
        return redirect(url_for('reports_list'))
    
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
    else:
        flash('Report file not found on disk', 'error')
        return redirect(url_for('reports_list'))

@app.route('/students')
@login_required
def students():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        students_raw = []
        program_rows = []
        try:
            # Get all students and their information
            students_raw = conn.execute('''
                SELECT s.*, 
                       COUNT(DISTINCT sess.id) as session_count
                FROM Student s
                LEFT JOIN Appointment a ON s.id = a.student_id
                LEFT JOIN session sess ON a.id = sess.appointment_id
                GROUP BY s.id
                ORDER BY s.name
            ''').fetchall()

            # Get all unique programs for the filter dropdown (extract as strings, not Row objects)
            program_rows = conn.execute("SELECT DISTINCT programme FROM Student WHERE programme IS NOT NULL AND programme != '' ORDER BY programme").fetchall()
        except Exception as e:
            print(f"[STUDENTS] Error getting students: {e}")
            students_raw = []
            program_rows = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Convert to list and add professional IDs
        students = []
        for student in students_raw:
            student_dict = dict(student)
            student_db_id = student_dict.get('id', 0)
            student_dict['professional_id'] = f"C{student_db_id:03d}"
            students.append(student_dict)

        programs = [row['programme'] for row in program_rows] if program_rows else []  # Convert Row objects to strings
        
        return render_template('students.html', students=students, programs=programs)
    except Exception as e:
        print(f"[STUDENTS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading students. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/student_profile/<int:id>')
@login_required
def student_profile(id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        # Get student information
        student = None
        sessions = []
        referrals = []
        try:
            student = conn.execute('SELECT * FROM Student WHERE id = ?', (id,)).fetchone()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting student: {e}")

        if not student:
            try:
                conn.close()
            except Exception:
                pass
            flash('Student not found', 'error')
            return redirect(url_for('students'))

        # Get all sessions for this student
        try:
            sessions = conn.execute('''
                SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
                       s.name as student_name,
                       c.name as Counsellor_name,
                       a.date, a.time, a.status
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE s.id = ?
                ORDER BY sess.created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting sessions: {e}")
            sessions = []

        # Get all referrals for this student
        try:
            referrals = conn.execute('''
                SELECT r.id, r.referred_by, r.contact, r.reasons, r.action_taken, r.outcome, r.created_at
                FROM Referral r
                JOIN Session sess ON r.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                WHERE a.student_id = ?
                ORDER BY r.created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting referrals: {e}")
            referrals = []

        # Get DASS-21 scores if table exists
        dass21_scores = []
        try:
            dass21_scores = conn.execute('''
                SELECT depression_score, anxiety_score, stress_score, completion_date, created_at
                FROM DASS21
                WHERE student_id = ?
                ORDER BY created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting DASS21 scores: {e}")
            dass21_scores = []

        # Get OQ-45.2 scores if table exists
        oq_scores = []
        try:
            oq_scores = conn.execute('''
                SELECT total_score, completion_date, created_at,
                       (SELECT created_at FROM session WHERE id = OutcomeQuestionnaire.session_id) as session_date
                FROM OutcomeQuestionnaire
                WHERE student_id = ?
                ORDER BY created_at DESC
            ''', (id,)).fetchall()
        except Exception as e:
            print(f"[STUDENT_PROFILE] Error getting OQ scores: {e}")
            oq_scores = []

        try:
            conn.close()
        except Exception:
            pass

        return render_template('student_profile.html', 
                             student=student, 
                             session=sessions,
                             referrals=referrals,
                             dass21_scores=dass21_scores,
                             oq_scores=oq_scores)
    except Exception as e:
        print(f"[STUDENT_PROFILE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading student profile. Please try again.', 'error')
        return redirect(url_for('students'))

@app.route('/export_students')
@login_required
def export_students():
    import csv
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Get format parameter (default to csv for backward compatibility)
    export_format = request.args.get('format', 'csv').lower()
    
    conn = get_db_connection()
    
    # Get all students data with professional ID
    students_raw = conn.execute('''
        SELECT s.*, 
               COUNT(DISTINCT sess.id) as session_count
        FROM Student s
        LEFT JOIN Appointment a ON s.id = a.student_id
        LEFT JOIN Session sess ON a.id = sess.appointment_id
        GROUP BY s.id
        ORDER BY s.name
    ''').fetchall()
    
    conn.close()
    
    # Generate professional IDs
    students = []
    for student in students_raw:
        student_dict = dict(student)
        student_db_id = student_dict.get('id', 0)
        student_dict['professional_id'] = f"C{student_db_id:03d}"
        students.append(student_dict)
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Define header style
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # Write headers
        headers = ['ID', 'Professional ID', 'Name', 'Index Number', 'Age', 'Gender', 'Email', 'Phone', 'Program', 'Department', 'Session Count', 'Created Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=student['id'])
            ws.cell(row=row_num, column=2, value=student['professional_id'])
            ws.cell(row=row_num, column=3, value=student['name'])
            ws.cell(row=row_num, column=4, value=student['index_number'] or 'N/A')
            ws.cell(row=row_num, column=5, value=student.get('age') or 'N/A')
            ws.cell(row=row_num, column=6, value=student.get('gender') or 'N/A')
            ws.cell(row=row_num, column=7, value=student.get('email') or 'N/A')
            ws.cell(row=row_num, column=8, value=student['contact'] or 'N/A')
            ws.cell(row=row_num, column=9, value=student.get('programme') or student.get('program') or 'N/A')
            ws.cell(row=row_num, column=10, value=student.get('department') or 'N/A')
            ws.cell(row=row_num, column=11, value=student['session_count'])
            ws.cell(row=row_num, column=12, value=student['created_at'])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=students_export.xlsx'
        return response
    else:
        # Create CSV (default)
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['ID', 'Professional ID', 'Name', 'Index Number', 'Email', 'Phone', 'Program', 'Session Count', 'Created Date'])

        # Write data rows
        for student in students:
            writer.writerow([
                student['id'],
                student['professional_id'],
                student['name'],
                student['index_number'] or 'N/A',
                student.get('email') or 'N/A',
                student['contact'] or 'N/A',
                student.get('programme') or student.get('program') or 'N/A',
                student['session_count'],
                student['created_at']
            ])

        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=students_export.csv'

        return response

@app.route('/export_sessions')
@login_required
def export_sessions():
    import csv
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Get format parameter (default to csv for backward compatibility)
    export_format = request.args.get('format', 'csv').lower()
    
    conn = get_db_connection()
    
    # Get all sessions with full details
    sessions = conn.execute('''
        SELECT sess.id, sess.session_type, sess.notes, sess.created_at,
               s.name as student_name, s.id as student_db_id,
               c.name as Counsellor_name,
               a.date, a.time, a.status as appointment_status
        FROM session sess
        LEFT JOIN Appointment a ON sess.appointment_id = a.id
        LEFT JOIN Student s ON a.student_id = s.id
        LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
        ORDER BY sess.created_at DESC
    ''').fetchall()
    
    conn.close()
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sessions"
        
        # Define header style
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # Write headers
        headers = ['ID', 'Date', 'Time', 'Student Name', 'Student ID', 'Counsellor', 'Session Type', 'Status', 'Notes', 'Created At']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for row_num, session in enumerate(sessions, 2):
            student_db_id = session.get('student_db_id', 0) if session.get('student_db_id') else 0
            professional_id = f"C{student_db_id:03d}" if student_db_id else 'N/A'
            
            ws.cell(row=row_num, column=1, value=session['id'])
            ws.cell(row=row_num, column=2, value=session['date'] or 'N/A')
            ws.cell(row=row_num, column=3, value=session['time'] or 'N/A')
            ws.cell(row=row_num, column=4, value=session['student_name'] or 'N/A')
            ws.cell(row=row_num, column=5, value=professional_id)
            ws.cell(row=row_num, column=6, value=session['Counsellor_name'] or 'N/A')
            ws.cell(row=row_num, column=7, value=session['session_type'] or 'N/A')
            ws.cell(row=row_num, column=8, value=session['appointment_status'] or 'N/A')
            ws.cell(row=row_num, column=9, value=session['notes'] or 'N/A')
            ws.cell(row=row_num, column=10, value=session['created_at'])
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=sessions_export.xlsx'
        return response
    else:
        # Create CSV (default)
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['ID', 'Date', 'Time', 'Student Name', 'Student ID', 'Counsellor', 'Session Type', 'Status', 'Notes', 'Created At'])

        # Write data rows
        for session in sessions:
            student_db_id = session.get('student_db_id', 0) if session.get('student_db_id') else 0
            professional_id = f"C{student_db_id:03d}" if student_db_id else 'N/A'

            writer.writerow([
                session['id'],
                session['date'] or 'N/A',
                session['time'] or 'N/A',
                session['student_name'] or 'N/A',
                professional_id,
                session['Counsellor_name'] or 'N/A',
                session['session_type'] or 'N/A',
                session['appointment_status'] or 'N/A',
                (session['notes'] or '').replace('\n', ' ')[:200],
                session['created_at']
            ])

        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=sessions_export.csv'
    
    return response

@app.route('/referral', methods=['GET', 'POST'])
@login_required
def referral():
    try:
        ensure_database_initialized()
        
        if request.method == 'POST':
            # Get form data
            session_id = request.form.get('session_id')
            referred_by = request.form.get('referred_by')
            contact = request.form.get('contact')
            action_taken = request.form.get('action_taken', '')
            outcome = request.form.get('outcome', '')

            # Get selected referral reasons (checkboxes)
            selected_reasons = request.form.getlist('referral_reasons')
            other_reason_text = request.form.get('other_reason_text', '').strip()

            # Combine reasons
            reasons_list = selected_reasons.copy()
            if 'Something Else' in selected_reasons and other_reason_text:
                # Replace "Something Else" with the actual text
                reasons_list.remove('Something Else')
                reasons_list.append(f'Something Else: {other_reason_text}')

            # Join reasons with commas, or use existing textarea value if checkboxes weren't used
            if reasons_list:
                reasons = ', '.join(reasons_list)
            else:
                reasons = request.form.get('reasons', '')  # Fallback to old textarea

            # Validate required fields
            if not all([session_id, referred_by, contact]) or not reasons:
                flash('Please fill in all required fields and select at least one reason', 'error')
                return redirect(url_for('referral'))

            # Insert referral into database
            conn = get_db_connection()
            if conn is None:
                flash('Database connection failed. Please restart the application.', 'error')
                return redirect(url_for('dashboard'))

            try:
                conn.execute('''
                    INSERT INTO Referral (session_id, referred_by, contact, reasons, action_taken, outcome, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (session_id, referred_by, contact, reasons, action_taken, outcome, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                conn.commit()
                flash('Referral created successfully!', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                conn.rollback()
                print(f"[REFERRAL] Error creating referral: {e}")
                flash(f'Error creating referral: {str(e)}', 'error')
                return redirect(url_for('referral'))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
    
        # GET request - display referral form
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))

        sessions = []
        try:
            sessions = conn.execute('''
                SELECT s.id, s.created_at, st.name as student_name, st.index_number, c.name as Counsellor_name
                FROM session s
                JOIN Appointment a ON s.appointment_id = a.id
                JOIN Student st ON a.student_id = st.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                ORDER BY s.created_at DESC
            ''').fetchall()
        except Exception as e:
            print(f"[REFERRAL] Error getting sessions: {e}")
            sessions = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('referral.html', sessions=sessions)
    except Exception as e:
        print(f"[REFERRAL] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading referral page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/all_referrals')
@login_required
def all_referrals():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        referrals = []
        try:
            # Get all referrals with student information
            referrals_raw = conn.execute('''
                SELECT r.id, r.session_id, r.referred_by, r.contact, r.reasons, r.created_at,
                               st.id as student_db_id, st.name as student_name, st.contact as student_contact, st.index_number
                FROM Referral r
                JOIN session sess ON r.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                JOIN Student st ON a.student_id = st.id
                ORDER BY r.created_at DESC
            ''').fetchall()
            
            # Convert to list and add professional ID
            referrals = []
            for ref in referrals_raw:
                ref_dict = dict(ref)
                # Generate professional ID: C001, C002, etc. based on student database ID
                student_db_id = ref_dict.get('student_db_id', 0)
                professional_id = f"C{student_db_id:03d}"  # C001, C002, etc.
                ref_dict['professional_id'] = professional_id
                referrals.append(ref_dict)
        except Exception as e:
            print(f"[REFERRALS] Error getting referrals: {e}")
            referrals = []
        finally:
            try:
                conn.close()
            except Exception:
                pass

        return render_template('referrals.html', referrals=referrals)
    except Exception as e:
        print(f"[REFERRALS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading referrals. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/export_referrals')
@login_required
def export_referrals():
    import csv
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Get format parameter (default to csv for backward compatibility)
    export_format = request.args.get('format', 'csv').lower()
    
    conn = get_db_connection()
    
    # Get all referrals with full details
    referrals_raw = conn.execute('''
        SELECT r.id, r.session_id, r.referred_by, r.contact, r.reasons, 
               r.action_taken, r.outcome, r.created_at,
               st.id as student_db_id, st.name as student_name, st.contact as student_contact
        FROM Referral r
        JOIN session sess ON r.session_id = sess.id
        JOIN Appointment a ON sess.appointment_id = a.id
        JOIN Student st ON a.student_id = st.id
        ORDER BY r.created_at DESC
    ''').fetchall()
    
    conn.close()
    
    if export_format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Referrals"
        
        # Define header style
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # Write headers
        headers = ['ID', 'Date', 'Student Name', 'Student ID', 'Referred By', 'Contact', 'Reasons', 'Action Taken', 'Outcome']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows with professional IDs
        for row_num, referral in enumerate(referrals_raw, 2):
            ref_dict = dict(referral)
            student_db_id = ref_dict.get('student_db_id', 0)
            professional_id = f"C{student_db_id:03d}"
            
            ws.cell(row=row_num, column=1, value=referral['id'])
            ws.cell(row=row_num, column=2, value=referral['created_at'])
            ws.cell(row=row_num, column=3, value=referral['student_name'] or 'N/A')
            ws.cell(row=row_num, column=4, value=professional_id)
            ws.cell(row=row_num, column=5, value=referral['referred_by'] or 'N/A')
            ws.cell(row=row_num, column=6, value=referral['contact'] or 'N/A')
            ws.cell(row=row_num, column=7, value=referral['reasons'] or 'N/A')
            ws.cell(row=row_num, column=8, value=referral['action_taken'] or 'N/A')
            ws.cell(row=row_num, column=9, value=referral['outcome'] or 'N/A')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=referrals_export.xlsx'
        return response
    else:
        # Create CSV (default)
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header - using "Student ID" instead of "Student Contact"
        writer.writerow(['ID', 'Date', 'Student Name', 'Student ID', 'Referred By', 'Contact', 'Reasons', 'Action Taken', 'Outcome'])

        # Write data rows with professional IDs
        for referral in referrals_raw:
            ref_dict = dict(referral)
            # Generate professional ID: C001, C002, etc.
            student_db_id = ref_dict.get('student_db_id', 0)
            professional_id = f"C{student_db_id:03d}"

            writer.writerow([
                referral['id'],
                referral['created_at'],
                referral['student_name'] or 'N/A',
                professional_id,  # Use professional ID instead of phone number
                referral['referred_by'] or 'N/A',
                referral['contact'] or 'N/A',
                (referral['reasons'] or '').replace('\n', ' ')[:200],  # Limit length and replace newlines
                (referral['action_taken'] or '').replace('\n', ' ')[:200],
                (referral['outcome'] or '').replace('\n', ' ')[:200]
            ])

        # Prepare response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=referrals_export.csv'

        return response

@app.route('/outcome_questionnaire', methods=['GET', 'POST'])
@login_required
def outcome_questionnaire():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            # Get form data
            student_id = request.form.get('student_id')
            session_id = request.form.get('session_id')
            age = request.form.get('age')
            sex = request.form.get('sex')

            # Get all 25 item scores
            item_scores = []
            try:
                for i in range(1, 26):
                    score = request.form.get(f'item{i}')
                    if not score or score == '':
                        flash('Please fill in all item scores', 'error')
                        return redirect(url_for('outcome_questionnaire'))
                    item_scores.append(int(score))
            except ValueError:
                flash('Invalid score values. Please enter numbers only.', 'error')
                return redirect(url_for('outcome_questionnaire'))

            # Calculate total score
            total_score = sum(item_scores)

            # Insert questionnaire data into database
            try:
                # Use correct table name and column names (item1, item2, etc. without underscores)
                conn.execute('''
                    INSERT INTO OutcomeQuestionnaire 
                    (student_id, session_id, age, sex, item1, item2, item3, item4, item5,
                     item6, item7, item8, item9, item10, item11, item12, item13, item14, item15,
                     item16, item17, item18, item19, item20, item21, item22, item23, item24, item25,
                     total_score, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (student_id, session_id, age if age else None, sex if sex else None, *item_scores, total_score, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                conn.commit()
                flash('Outcome questionnaire submitted successfully!', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                conn.rollback()
                print(f"[OUTCOME_QUESTIONNAIRE] Error saving: {e}")
                flash(f'Error submitting questionnaire: {str(e)}', 'error')
                return redirect(url_for('outcome_questionnaire'))
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

        # GET request - display the form
        students_raw = []
        sessions_raw = []
        try:
            students_raw = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
            sessions_raw = conn.execute('''
                SELECT sess.id, sess.created_at, a.student_id, s.id as student_db_id, s.name as student_name
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                WHERE a.student_id IS NOT NULL
                ORDER BY sess.created_at DESC
            ''').fetchall()
        except Exception as e:
            print(f"[OUTCOME_QUESTIONNAIRE] Error getting dropdown data: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        # Convert students to list
        students = []
        for student in students_raw:
            students.append(dict(student))
        
        # Convert sessions to list and format dates
        sessions = []
        for sess in sessions_raw:
            sess_dict = dict(sess)
            # Format created_at for display
            if sess_dict.get('created_at'):
                try:
                    dt = datetime.strptime(sess_dict['created_at'], '%Y-%m-%d %H:%M:%S')
                    sess_dict['created_at_formatted'] = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    sess_dict['created_at_formatted'] = sess_dict['created_at']
            else:
                sess_dict['created_at_formatted'] = 'N/A'
            
            # Add professional ID
            student_db_id = sess_dict.get('student_db_id', 0)
            sess_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
            sessions.append(sess_dict)
        
        return render_template('outcome_questionnaire.html', students=students, sessions=sessions)
    except Exception as e:
        print(f"[OUTCOME_QUESTIONNAIRE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading outcome questionnaire page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/dass21', methods=['GET', 'POST'])
@login_required
def dass21():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            # Get form data
            student_id = request.form.get('student_id', '').strip()
            depression_score = request.form.get('depression_score', '0')
            anxiety_score = request.form.get('anxiety_score', '0')
            stress_score = request.form.get('stress_score', '0')
            
            # Validate required fields
            if not student_id or student_id == '':
                flash('Please select a student', 'error')
                try:
                    students_raw = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
                    students = [dict(s) for s in students_raw]
                except Exception as e:
                    print(f"[DASS21] Error getting students: {e}")
                    students = []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                return render_template('dass21.html', students=students)
            
            try:
                depression_score = float(depression_score)
                anxiety_score = float(anxiety_score)
                stress_score = float(stress_score)
            except ValueError:
                flash('Please enter valid numeric scores', 'error')
                try:
                    students_raw = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
                    students = [dict(s) for s in students_raw]
                except Exception as e:
                    print(f"[DASS21] Error getting students: {e}")
                    students = []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                return render_template('dass21.html', students=students)
            
            # Calculate final scores (multiply by 2)
            final_depression = depression_score * 2
            final_anxiety = anxiety_score * 2
            final_stress = stress_score * 2
            
            # Insert DASS-21 scores into database
            try:
                conn.execute('''
                    INSERT INTO DASS21 
                    (student_id, depression_score, anxiety_score, stress_score, created_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (student_id, depression_score, anxiety_score, stress_score,
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                conn.commit()
                flash('DASS-21 scores saved successfully!', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                conn.rollback()
                print(f"[DASS21] Error saving scores: {e}")
                import traceback
                traceback.print_exc()
                flash(f'Error saving DASS-21 scores: {str(e)}', 'error')
                try:
                    students_raw = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
                    students = [dict(s) for s in students_raw]
                except Exception as e:
                    print(f"[DASS21] Error getting students: {e}")
                    students = []
                finally:
                    try:
                        conn.close()
                    except Exception:
                        pass
                return render_template('dass21.html', students=students)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        
        # GET request - display the form
        students_raw = []
        try:
            students_raw = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
        except Exception as e:
            print(f"[DASS21] Error getting students: {e}")
            students_raw = []
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        # Convert to list
        students = []
        for student in students_raw:
            students.append(dict(student))
        
        return render_template('dass21.html', students=students)
    except Exception as e:
        print(f"[DASS21] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading DASS-21 page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/import_csv', methods=['GET', 'POST'])
@login_required
def import_csv():
    if request.method == 'POST':
        if 'confirm' in request.form:
            # Process confirmed import
            import_type = request.form.get('import_type')
            confirmed_data = request.form.get('confirmed_data')
            
            if not confirmed_data:
                flash('No data to import', 'error')
                return redirect(url_for('import_csv'))
            
            try:
                data = json.loads(confirmed_data)
                conn = get_db_connection()
                
                if import_type == 'students':
                    # Import students data
                    for row in data:
                        conn.execute('''
                            INSERT OR REPLACE INTO students 
                            (student_id, first_name, last_name, email, phone, department, programme, parent_contact)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            row.get('index_number', ''),
                            row.get('name', '').split()[0] if ' ' in row.get('name', '') else row.get('name', ''),
                            row.get('name', '').split()[-1] if ' ' in row.get('name', '') else '',
                            row.get('email', ''),
                            row.get('phone', ''),
                            row.get('department', ''),
                            row.get('programme', ''),
                            row.get('parent_contact', '')
                        ))
                
                elif import_type == 'appointments':
                    # Import appointments data
                    for row in data:
                        # Get student_id from student_name
                        student_name = row.get('student_name', '')
                        student = conn.execute(
                            'SELECT id FROM Student WHERE name = ? OR id = ?',
                            (student_name, student_name)
                        ).fetchone()
                        
                        if student:
                            conn.execute('''
                                INSERT INTO Appointment 
                                (student_id, date, time, purpose, counselor_id, status)
                                VALUES (?, ?, ?, ?, ?, ?)
                            ''', (
                                student['id'],
                                row.get('date', ''),
                                row.get('time', ''),
                                row.get('purpose', ''),
                                session.get('user_id'),
                                row.get('status', 'scheduled')
                            ))
                
                conn.commit()
                conn.close()
                
                flash(f'Successfully imported {len(data)} {import_type}', 'success')
                return redirect(url_for('dashboard'))
                
            except Exception as e:
                flash(f'Error importing data: {str(e)}', 'error')
                return redirect(url_for('import_csv'))
        
        else:
            # Handle file upload and preview
            import_type = request.form.get('import_type')
            csv_file = request.files.get('csv_file')
            
            if not import_type or not csv_file:
                flash('Please select import type and CSV file', 'error')
                return redirect(url_for('import_csv'))
            
            try:
                # Read CSV file
                csv_data = csv_file.read().decode('utf-8').splitlines()
                csv_reader = csv.DictReader(csv_data)
                
                preview_data = []
                errors = []
                
                # Validate and prepare data
                for row_num, row in enumerate(csv_reader, start=2):
                    if import_type == 'students':
                        # Validate student data
                        if not row.get('name') or not row.get('index_number'):
                            errors.append(f'Row {row_num}: Missing required fields (name, index_number)')
                        else:
                            preview_data.append({
                                'name': row.get('name', ''),
                                'index_number': row.get('index_number', ''),
                                'email': row.get('email', ''),
                                'phone': row.get('phone', ''),
                                'department': row.get('department', ''),
                                'programme': row.get('programme', ''),
                                'parent_contact': row.get('parent_contact', '')
                            })
                    
                    elif import_type == 'appointments':
                        # Validate appointment data
                        if not row.get('student_name') or not row.get('date') or not row.get('time'):
                            errors.append(f'Row {row_num}: Missing required fields (student_name, date, time)')
                        else:
                            preview_data.append({
                                'student_name': row.get('student_name', ''),
                                'date': row.get('date', ''),
                                'time': row.get('time', ''),
                                'counsellor': row.get('counsellor', ''),
                                'purpose': row.get('purpose', ''),
                                'status': row.get('status', 'scheduled')
                            })
                
                if not preview_data:
                    flash('No valid data found in CSV file', 'error')
                    return redirect(url_for('import_csv'))
                
                return render_template('import_csv.html', 
                                     preview_data=preview_data,
                                     headers=list(preview_data[0].keys()) if preview_data else [],
                                     errors=errors,
                                     import_type=import_type,
                                     confirmed_data=json.dumps(preview_data))
                
            except Exception as e:
                flash(f'Error reading CSV file: {str(e)}', 'error')
                return redirect(url_for('import_csv'))
    
    # GET request - display the form
    return render_template('import_csv.html')

@app.route('/intake', methods=['GET', 'POST'])
@login_required
def intake():
    if request.method == 'POST':
        # Get form data
        student_id = request.form.get('student_id')
        
        # Check if student exists
        conn = get_db_connection()
        student = conn.execute('SELECT * FROM Student WHERE id = ?', (student_id,)).fetchone()
        
        if not student:
            conn.close()
            flash('Student not found. Please check the ID or add the student first.', 'danger')
            return redirect(url_for('intake'))
        
        # Process intake form data
        intake_data = {
            'student_id': student_id,
            'date': request.form.get('date'),
            'presenting_issue': request.form.get('presenting_issue'),
            'background': request.form.get('background'),
            'mental_status': request.form.get('mental_status'),
            'risk_assessment': request.form.get('risk_assessment'),
            'diagnosis': request.form.get('diagnosis'),
            'treatment_plan': request.form.get('treatment_plan'),
            'counselor_id': session.get('user_id')
        }
        
        # Insert intake form data into database
        try:
            conn.execute('''
                INSERT INTO intake_forms (student_id, date, presenting_issue, background, 
                mental_status, risk_assessment, diagnosis, treatment_plan, counselor_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                intake_data['student_id'], intake_data['date'], intake_data['presenting_issue'],
                intake_data['background'], intake_data['mental_status'], intake_data['risk_assessment'],
                intake_data['diagnosis'], intake_data['treatment_plan'], intake_data['counselor_id']
            ))
            conn.commit()
            conn.close()
            
            flash('Intake form submitted successfully!', 'success')
            return redirect(url_for('student_profile', id=student_id))
        except Exception as e:
            conn.close()
            flash(f'Error submitting intake form: {str(e)}', 'danger')
            return redirect(url_for('intake'))
    
    # GET request - display the form
    return render_template('intake.html')

@app.route('/appointment', methods=['GET', 'POST'])
@login_required
def appointment():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        if request.method == 'POST':
            # Get form data
            student_id = request.form.get('student_id')
            appointment_date = request.form.get('date')
            appointment_time = request.form.get('time')
            purpose = request.form.get('purpose')
            counselor_id = request.form.get('counselor_id')

            students = []
            counsellors = []
            try:
                students = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
                counsellors = conn.execute('SELECT id, name FROM Counsellor ORDER BY name').fetchall()
            except Exception as e:
                print(f"[APPOINTMENT] Error getting dropdown data: {e}")

            # Validate data
            if not student_id or not appointment_date or not appointment_time or not counselor_id:
                try:
                    conn.close()
                except Exception:
                    pass
                flash('Please fill in all required fields', 'danger')
                return render_template('appointment.html', students=students, Counsellors=counsellors)

            # Check if student exists
            try:
                student = conn.execute('SELECT * FROM Student WHERE id = ?', (student_id,)).fetchone()
            except Exception as e:
                print(f"[APPOINTMENT] Error checking student: {e}")
                student = None

            if not student:
                try:
                    conn.close()
                except Exception:
                    pass
                flash('Student not found. Please check the ID or add the student first.', 'danger')
                return render_template('appointment.html', students=students, Counsellors=counsellors)

            # Save appointment to database
            try:
                conn.execute('''
                        INSERT INTO Appointment (student_id, date, time, purpose, Counsellor_id, status)
                    VALUES (?, ?, ?, ?, ?, 'scheduled')
                ''', (student_id, appointment_date, appointment_time, purpose, counselor_id))
                conn.commit()
                flash('Appointment scheduled successfully!', 'success')
                return redirect(url_for('manage_appointments'))
            except Exception as e:
                conn.rollback()
                print(f"[APPOINTMENT] Error saving appointment: {e}")
                flash(f'Error scheduling appointment: {str(e)}', 'danger')
                return render_template('appointment.html', students=students, Counsellors=counsellors)
            finally:
                try:
                    conn.close()
                except Exception:
                    pass
        
        # GET request - display the form
        students = []
        counsellors = []
        try:
            students = conn.execute('SELECT id, name, programme FROM Student ORDER BY name').fetchall()
            counsellors = conn.execute('SELECT id, name FROM Counsellor ORDER BY name').fetchall()
        except Exception as e:
            print(f"[APPOINTMENT] Error getting dropdown data: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass
    
        return render_template('appointment.html', students=students, Counsellors=counsellors)
    except Exception as e:
        print(f"[APPOINTMENT] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading appointment page. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/manage_appointments')
@login_required
def manage_appointments():
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
    
        appointments = []
        try:
            # Get all appointments with student and counsellor names
            appointments = conn.execute('''
                SELECT a.*, s.name as student_name, c.name as Counsellor_name
                FROM Appointment a
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                ORDER BY a.date DESC, a.time DESC
            ''').fetchall()
        except Exception as e:
            print(f"[APPOINTMENTS] Error getting appointments: {e}")
            appointments = []
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        return render_template('appointments.html', appointments=appointments)
    except Exception as e:
        print(f"[APPOINTMENTS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading appointments. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/update_appointment_status/<int:appointment_id>', methods=['POST'])
@login_required
def update_appointment_status(appointment_id):
    """Update the status of an appointment"""
    new_status = request.form.get('status')
    
    if not new_status:
        flash('Status is required', 'error')
        return redirect(url_for('manage_appointments'))
    
    # Valid statuses
    valid_statuses = ['Scheduled', 'In Session', 'Completed', 'Cancelled', 'Postponed']
    if new_status not in valid_statuses:
        flash('Invalid status', 'error')
        return redirect(url_for('manage_appointments'))
    
    conn = get_db_connection()
    try:
        # Check if appointment exists
        appointment = conn.execute('SELECT id FROM Appointment WHERE id = ?', (appointment_id,)).fetchone()
        
        if not appointment:
            flash('Appointment not found', 'error')
            return redirect(url_for('manage_appointments'))
        
        # Update the status
        conn.execute('''
            UPDATE Appointment 
            SET status = ? 
            WHERE id = ?
        ''', (new_status, appointment_id))
        conn.commit()
        
        flash(f'Appointment status updated to {new_status} successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error updating appointment status: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('manage_appointments'))

# ---------- Print Routes ----------
@app.route('/view_report/<int:report_id>')
@login_required
def view_report(report_id):
    conn = get_db_connection()
    
    # Get report details
    report = conn.execute('''
        SELECT * FROM reports WHERE id = ?
    ''', (report_id,)).fetchone()
    
    if not report:
        conn.close()
        flash('Report not found', 'error')
        return redirect(url_for('reports_list'))
    
    conn.close()
    # Convert Row to dict for easier template access
    report_dict = dict(report)
    return render_template('view_report.html', report=report_dict, now=datetime.utcnow())

@app.route('/download_report/<int:report_id>')
@login_required
def download_report(report_id):
    conn = get_db_connection()
    
    # Get report details
    report = conn.execute('''
        SELECT * FROM reports WHERE id = ?
    ''', (report_id,)).fetchone()
    
    if not report:
        conn.close()
        flash('Report not found', 'error')
        return redirect(url_for('reports_list'))
    
    conn.close()
    
    # For now, redirect to print report as download functionality
    # This can be enhanced to generate actual downloadable files
    return redirect(url_for('print_report', report_id=report_id))

@app.route('/delete_report/<int:report_id>', methods=['POST'])
@login_required
def delete_report(report_id):
    conn = get_db_connection()
    
    # Delete report from database
    try:
        result = conn.execute('DELETE FROM reports WHERE id = ?', (report_id,)).rowcount
        conn.commit()
        
        if result > 0:
            flash('Report deleted successfully!', 'success')
        else:
            flash('Report not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting report: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('reports_list'))

@app.route('/delete_student/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    """Delete a student and all related records"""
    conn = get_db_connection()
    
    try:
        # Check if student exists
        student = conn.execute('SELECT * FROM Student WHERE id = ?', (student_id,)).fetchone()
        if not student:
            flash('Student not found', 'error')
            return redirect(url_for('students'))
        
        # Delete related records first (due to foreign keys)
        # Delete referrals (through sessions through appointments)
        conn.execute('''
            DELETE FROM Referral 
            WHERE session_id IN (
                SELECT s.id FROM session s
                JOIN Appointment a ON s.appointment_id = a.id
                WHERE a.student_id = ?
            )
        ''', (student_id,))
        
        # Delete sessions
        conn.execute('''
            DELETE FROM session
            WHERE appointment_id IN (
                SELECT id FROM Appointment WHERE student_id = ?
            )
        ''', (student_id,))
        
        # Delete appointments
        conn.execute('DELETE FROM Appointment WHERE student_id = ?', (student_id,))
        
        # Delete assessments
        conn.execute('DELETE FROM DASS21 WHERE student_id = ?', (student_id,))
        conn.execute('DELETE FROM OutcomeQuestionnaire WHERE student_id = ?', (student_id,))
        
        # Finally delete the student
        result = conn.execute('DELETE FROM Student WHERE id = ?', (student_id,)).rowcount
        conn.commit()
        
        if result > 0:
            flash('Student and all related records deleted successfully!', 'success')
        else:
            flash('Student not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting student: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('students'))

@app.route('/delete_appointment/<int:appointment_id>', methods=['POST'])
@login_required
def delete_appointment(appointment_id):
    """Delete an appointment and related sessions"""
    conn = get_db_connection()
    
    try:
        # Check if appointment exists
        appointment = conn.execute('SELECT * FROM Appointment WHERE id = ?', (appointment_id,)).fetchone()
        if not appointment:
            flash('Appointment not found', 'error')
            return redirect(url_for('manage_appointments'))
        
        # Delete related sessions first
        conn.execute('DELETE FROM session WHERE appointment_id = ?', (appointment_id,))
        
        # Delete the appointment
        result = conn.execute('DELETE FROM Appointment WHERE id = ?', (appointment_id,)).rowcount
        conn.commit()
        
        if result > 0:
            flash('Appointment deleted successfully!', 'success')
        else:
            flash('Appointment not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting appointment: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('manage_appointments'))

@app.route('/delete_session/<int:session_id>', methods=['POST'])
@login_required
def delete_session(session_id):
    """Delete a session and related records"""
    conn = get_db_connection()
    
    try:
        # Check if session exists
        session_record = conn.execute('SELECT * FROM session WHERE id = ?', (session_id,)).fetchone()
        if not session_record:
            flash('Session not found', 'error')
            return redirect(url_for('sessions_list'))
        
        # Delete related referrals
        conn.execute('DELETE FROM Referral WHERE session_id = ?', (session_id,))
        
        # Delete case management records
        conn.execute('DELETE FROM CaseManagement WHERE session_id = ?', (session_id,))
        
        # Delete the session
        result = conn.execute('DELETE FROM session WHERE id = ?', (session_id,)).rowcount
        conn.commit()
        
        if result > 0:
            flash('Session deleted successfully!', 'success')
        else:
            flash('Session not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting session: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('sessions_list'))

@app.route('/delete_referral/<int:referral_id>', methods=['POST'])
@login_required
def delete_referral(referral_id):
    """Delete a referral"""
    conn = get_db_connection()
    
    try:
        result = conn.execute('DELETE FROM Referral WHERE id = ?', (referral_id,)).rowcount
        conn.commit()
        
        if result > 0:
            flash('Referral deleted successfully!', 'success')
        else:
            flash('Referral not found', 'error')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting referral: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('all_referrals'))

@app.route('/get_session/<int:session_id>')
@login_required
def get_session(session_id):
    """Get session details as JSON for the modal"""
    conn = get_db_connection()
    
    try:
        session_data = conn.execute('''
            SELECT sess.id, sess.session_type, sess.notes, sess.outcome, sess.created_at,
                   s.name as student_name,
                   c.name as Counsellor_name,
                   a.date, a.time, a.status as appointment_status
            FROM session sess
            LEFT JOIN Appointment a ON sess.appointment_id = a.id
            LEFT JOIN Student s ON a.student_id = s.id
            LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
            WHERE sess.id = ?
        ''', (session_id,)).fetchone()
        
        if not session_data:
            conn.close()
            return jsonify({'error': 'Session not found'}), 404
        
        # Convert to dict for JSON serialization
        session_dict = {
            'id': session_data['id'],
            'student_name': session_data['student_name'] or 'N/A',
            'Counsellor_name': session_data['Counsellor_name'] or 'N/A',
            'date': session_data['date'] or 'N/A',
            'time': session_data['time'] or 'N/A',
            'session_type': session_data['session_type'] or 'N/A',
            'appointment_status': session_data['appointment_status'] or 'N/A',
            'notes': session_data['notes'] or 'No notes provided',
            'outcome': session_data['outcome'] or 'N/A',
            'created_at': session_data['created_at']
        }
        
        conn.close()
        return jsonify(session_dict)
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/print_session/<int:session_id>')
@login_required
def print_session(session_id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
    
    # Get session details with student information
        session_data = None
        try:
            session_data = conn.execute('''
                SELECT sess.*, s.name as student_name, 
                       s.index_number, s.programme, s.contact, s.department,
                       c.name as Counsellor_name,
                       a.date, a.time, a.status as appointment_status
                FROM session sess
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                LEFT JOIN Counsellor c ON a.Counsellor_id = c.id
                WHERE sess.id = ?
            ''', (session_id,)).fetchone()
        except Exception as e:
            print(f"[PRINT_SESSION] Error getting session: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass
    
        if not session_data:
            flash('Session not found', 'error')
            return redirect(url_for('dashboard'))
        
        return render_template('print_session.html', session=session_data)
    except Exception as e:
        print(f"[PRINT_SESSION] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading session for printing.', 'error')
        return redirect(url_for('sessions_list'))

@app.route('/print_referral/<int:id>')
@login_required
def print_referral(id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
    
    # Get referral details with student information
        referral = None
        student_info = None
        try:
            referral = conn.execute('''
                SELECT r.*, s.id as student_db_id, s.name as student_name, 
                       s.index_number, s.contact as student_contact, s.department as student_department,
                       sess.created_at as session_date,
                       a.date as appointment_date, a.time as appointment_time
                FROM Referral r
                JOIN session sess ON r.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                JOIN Student s ON a.student_id = s.id
        WHERE r.id = ?
    ''', (id,)).fetchone()
    
            # Parse reasons from comma-separated string
            reasons_str = referral['reasons'] or ''
            referral_reasons_list = [r.strip() for r in reasons_str.split(',') if r.strip()]
            
            # Check for "Something Else" and extract text
            other_reason_text = None
            if referral_reasons_list:
                for idx, reason in enumerate(referral_reasons_list):
                    if reason.startswith('Something Else:'):
                        other_reason_text = reason.replace('Something Else:', '').strip()
                        referral_reasons_list[idx] = 'Something Else'  # Replace with just the checkbox name
                        break
            
            # Convert to dict for template and add professional ID
            referral_dict = dict(referral) if referral else {}
            if referral and referral_dict.get('student_db_id'):
                student_db_id = referral_dict.get('student_db_id', 0)
                referral_dict['professional_id'] = f"C{student_db_id:03d}"
            else:
                referral_dict['professional_id'] = 'N/A'
            referral_dict['referral_reasons_list'] = referral_reasons_list
            referral_dict['other_reason_text'] = other_reason_text
            student_department = referral.get('student_department', 'N/A') if referral else 'N/A'
            
        except Exception as e:
            print(f"[PRINT_REFERRAL] Error getting referral: {e}")
            import traceback
            traceback.print_exc()
            referral = None
            referral_dict = None
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        if not referral or not referral_dict:
            # Only flash error if it's a legitimate "not found" case (not a database error)
            if referral is None:
                flash('Referral not found', 'error')
            else:
                flash('Error loading referral details. Please try again.', 'error')
            return redirect(url_for('all_referrals'))
    
        return render_template('print_referral.html', 
                             referral=referral_dict,
                             referral_reasons_list=referral_reasons_list,
                             other_reason_text=other_reason_text,
                             student_department=student_department)
    except Exception as e:
        print(f"[PRINT_REFERRAL] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading referral for printing.', 'error')
        return redirect(url_for('all_referrals'))

@app.route('/print_case/<int:case_id>')
@login_required
def print_case(case_id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
    
    # Get case details with student information
        case = None
        try:
            case = conn.execute('''
                SELECT cm.*, s.name as student_name, 
                       s.index_number, s.programme, s.contact,
                       sess.created_at as session_date
                FROM CaseManagement cm
                JOIN session sess ON cm.session_id = sess.id
                JOIN Appointment a ON sess.appointment_id = a.id
                JOIN Student s ON a.student_id = s.id
                WHERE cm.id = ?
            ''', (case_id,)).fetchone()
        except Exception as e:
            print(f"[PRINT_CASE] Error getting case: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass
    
        if not case:
            flash('Case not found', 'error')
            return redirect(url_for('dashboard'))
        
        return render_template('print_case.html', case=case)
    except Exception as e:
        print(f"[PRINT_CASE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case for printing.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/case_notes_list')
@login_required
def case_notes_list():
    """Display all case notes linked to students"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        case_notes = []
        try:
            case_notes_raw = conn.execute('''
                SELECT cm.id, cm.session_id, cm.client_appearance, cm.problems, cm.interventions,
                       cm.recommendations, cm.next_visit_date, cm.counsellor_signature, cm.created_at,
                       s.id as student_db_id, s.name as student_name, s.index_number, s.programme,
                       sess.created_at as session_date
                FROM CaseManagement cm
                JOIN session sess ON cm.session_id = sess.id
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                ORDER BY cm.created_at DESC
            ''').fetchall()
            
            # Convert to list and add professional IDs
            for cn in case_notes_raw:
                cn_dict = dict(cn)
                student_db_id = cn_dict.get('student_db_id', 0)
                cn_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                case_notes.append(cn_dict)
        except Exception as e:
            print(f"[CASE_NOTES_LIST] Error getting case notes: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        return render_template('case_notes_list.html', case_notes=case_notes)
    except Exception as e:
        print(f"[CASE_NOTES_LIST] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case notes. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/dass21_list')
@login_required
def dass21_list():
    """Display all DASS-21 records linked to students"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        dass21_records = []
        try:
            dass21_raw = conn.execute('''
                SELECT d.id, d.student_id, d.depression_score, d.anxiety_score, d.stress_score,
                       d.created_at, d.completion_date,
                       s.id as student_db_id, s.name as student_name, s.index_number, s.programme
                FROM DASS21 d
                LEFT JOIN Student s ON d.student_id = s.id
                ORDER BY d.created_at DESC
            ''').fetchall()
            
            # Convert to list and add professional IDs and final scores
            for d in dass21_raw:
                d_dict = dict(d)
                student_db_id = d_dict.get('student_db_id', 0)
                d_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                # Calculate final scores (x2)
                d_dict['final_depression'] = (d_dict.get('depression_score', 0) or 0) * 2
                d_dict['final_anxiety'] = (d_dict.get('anxiety_score', 0) or 0) * 2
                d_dict['final_stress'] = (d_dict.get('stress_score', 0) or 0) * 2
                dass21_records.append(d_dict)
        except Exception as e:
            print(f"[DASS21_LIST] Error getting DASS-21 records: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        return render_template('dass21_list.html', dass21_records=dass21_records)
    except Exception as e:
        print(f"[DASS21_LIST] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading DASS-21 records. Please try again.', 'error')
        return redirect(url_for('dashboard'))

@app.route('/print_case_note/<int:case_id>')
@login_required
def print_case_note(case_id):
    """Print case note"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        case = None
        try:
            case = conn.execute('''
                SELECT cm.*, s.id as student_db_id, s.name as student_name, 
                       s.index_number, s.programme, s.contact, s.department,
                       sess.created_at as session_date
                FROM CaseManagement cm
                JOIN session sess ON cm.session_id = sess.id
                LEFT JOIN Appointment a ON sess.appointment_id = a.id
                LEFT JOIN Student s ON a.student_id = s.id
                WHERE cm.id = ?
            ''', (case_id,)).fetchone()
            
            if case:
                case_dict = dict(case)
                student_db_id = case_dict.get('student_db_id', 0)
                case_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                case = case_dict
        except Exception as e:
            print(f"[PRINT_CASE_NOTE] Error getting case: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        if not case:
            flash('Case note not found', 'error')
            return redirect(url_for('case_notes_list'))
        
        return render_template('print_case_note.html', case=case)
    except Exception as e:
        print(f"[PRINT_CASE_NOTE] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading case note for printing.', 'error')
        return redirect(url_for('case_notes_list'))

@app.route('/print_dass21/<int:dass21_id>')
@login_required
def print_dass21(dass21_id):
    """Print DASS-21 record"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        dass21_record = None
        try:
            dass21_raw = conn.execute('''
                SELECT d.*, s.id as student_db_id, s.name as student_name, 
                       s.index_number, s.programme, s.contact, s.department
                FROM DASS21 d
                LEFT JOIN Student s ON d.student_id = s.id
                WHERE d.id = ?
            ''', (dass21_id,)).fetchone()
            
            if dass21_raw:
                dass21_dict = dict(dass21_raw)
                student_db_id = dass21_dict.get('student_db_id', 0)
                dass21_dict['professional_id'] = f"C{student_db_id:03d}" if student_db_id else 'N/A'
                # Calculate final scores (x2)
                dass21_dict['final_depression'] = (dass21_dict.get('depression_score', 0) or 0) * 2
                dass21_dict['final_anxiety'] = (dass21_dict.get('anxiety_score', 0) or 0) * 2
                dass21_dict['final_stress'] = (dass21_dict.get('stress_score', 0) or 0) * 2
                dass21_record = dass21_dict
        except Exception as e:
            print(f"[PRINT_DASS21] Error getting DASS-21 record: {e}")
            import traceback
            traceback.print_exc()
        finally:
            try:
                conn.close()
            except Exception:
                pass
        
        if not dass21_record:
            flash('DASS-21 record not found', 'error')
            return redirect(url_for('dass21_list'))
        
        return render_template('print_dass21.html', dass21=dass21_record)
    except Exception as e:
        print(f"[PRINT_DASS21] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading DASS-21 record for printing.', 'error')
        return redirect(url_for('dass21_list'))

@app.route('/print_report/<int:report_id>')
@login_required
def print_report(report_id):
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
    
    # Get report details
        report = None
        try:
            report = conn.execute('''
                SELECT * FROM reports WHERE id = ?
            ''', (report_id,)).fetchone()
        except Exception as e:
            print(f"[PRINT_REPORT] Error getting report: {e}")
        finally:
            try:
                conn.close()
            except Exception:
                pass
    
        if not report:
            flash('Report not found', 'error')
            return redirect(url_for('reports_list'))
    
        # Convert Row to dict for easier template access
        report_dict = dict(report) if report else {}
        return render_template('print_report.html', report=report_dict, now=datetime.utcnow())
    except Exception as e:
        print(f"[PRINT_REPORT] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading report for printing.', 'error')
        return redirect(url_for('reports_list'))

@app.route('/statistics')
@login_required
def statistics():
    """Display comprehensive statistics with charts"""
    try:
        ensure_database_initialized()
        conn = get_db_connection()
        if conn is None:
            flash('Database connection failed. Please restart the application.', 'error')
            return redirect(url_for('dashboard'))
        
        # Initialize with defaults
        total_students = 0
        total_appointments = 0
        total_sessions = 0
        total_referrals = 0
        gender_stats = []
        programme_stats = []
        level_stats = []
        appointment_status_stats = []
        
        try:
            # Overall Statistics
            result = conn.execute('SELECT COUNT(*) as count FROM Student').fetchone()
            total_students = result['count'] if result else 0
            
            result = conn.execute('SELECT COUNT(*) as count FROM Appointment').fetchone()
            total_appointments = result['count'] if result else 0
            
            result = conn.execute('SELECT COUNT(*) as count FROM session').fetchone()
            total_sessions = result['count'] if result else 0
            
            try:
                result = conn.execute('SELECT COUNT(*) as count FROM Referral').fetchone()
                total_referrals = result['count'] if result else 0
            except:
                total_referrals = 0
            
            # Students by Gender
            try:
                gender_stats = conn.execute('''
                    SELECT COALESCE(gender, 'Not Specified') as gender, COUNT(*) as count 
                    FROM Student 
                    GROUP BY gender
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting gender stats: {e}")
                gender_stats = []
            
            # Students by Programme
            try:
                programme_stats = conn.execute('''
                    SELECT COALESCE(programme, 'Not Specified') as programme, COUNT(*) as count 
                    FROM Student 
                    GROUP BY programme
                    ORDER BY count DESC
                    LIMIT 10
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting programme stats: {e}")
                programme_stats = []
            
            # Students by Department
            try:
                department_stats = conn.execute('''
                    SELECT COALESCE(department, 'Not Specified') as department, COUNT(*) as count 
                    FROM Student 
                    GROUP BY department
                    ORDER BY count DESC
                    LIMIT 10
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting department stats: {e}")
                department_stats = []
            
            # Appointments by Status
            try:
                appointment_status_stats = conn.execute('''
                    SELECT COALESCE(status, 'Not Specified') as status, COUNT(*) as count 
                    FROM Appointment 
                    GROUP BY status
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting appointment status stats: {e}")
                appointment_status_stats = []
            
            # Appointments Over Time (Last 6 Months)
            try:
                appointment_timeline = conn.execute('''
                    SELECT strftime('%Y-%m', date) as month, COUNT(*) as count 
                    FROM Appointment 
                    WHERE date >= date('now', '-6 months')
                    GROUP BY month
                    ORDER BY month
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting appointment timeline: {e}")
                appointment_timeline = []
            
            # Sessions Over Time (Last 6 Months)
            try:
                session_timeline = conn.execute('''
                    SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
                    FROM session 
                    WHERE created_at >= date('now', '-6 months')
                    GROUP BY month
                    ORDER BY month
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting session timeline: {e}")
                session_timeline = []
            
            # Sessions by Type
            try:
                session_type_stats = conn.execute('''
                    SELECT COALESCE(session_type, 'Not Specified') as session_type, COUNT(*) as count 
                    FROM session 
                    GROUP BY session_type
                    ORDER BY count DESC
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting session type stats: {e}")
                session_type_stats = []
            
            # Age Distribution
            try:
                age_distribution = conn.execute('''
                    SELECT 
                        CASE 
                            WHEN age < 18 THEN 'Under 18'
                            WHEN age BETWEEN 18 AND 20 THEN '18-20'
                            WHEN age BETWEEN 21 AND 25 THEN '21-25'
                            WHEN age BETWEEN 26 AND 30 THEN '26-30'
                            WHEN age > 30 THEN 'Over 30'
                            ELSE 'Not Specified'
                        END as age_group,
                        COUNT(*) as count
                    FROM Student
                    GROUP BY age_group
                    ORDER BY 
                        CASE age_group
                            WHEN 'Under 18' THEN 1
                            WHEN '18-20' THEN 2
                            WHEN '21-25' THEN 3
                            WHEN '26-30' THEN 4
                            WHEN 'Over 30' THEN 5
                            ELSE 6
                        END
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting age distribution: {e}")
                age_distribution = []
            
            # Top Counsellors by Appointments
            try:
                top_counsellors = conn.execute('''
                    SELECT c.name, COUNT(a.id) as appointment_count
                    FROM Counsellor c
                    LEFT JOIN Appointment a ON c.id = a.Counsellor_id
                    GROUP BY c.id, c.name
                    ORDER BY appointment_count DESC
                    LIMIT 5
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting top counsellors: {e}")
                top_counsellors = []
            
            # Monthly New Students (Last 6 Months)
            try:
                new_students_timeline = conn.execute('''
                    SELECT strftime('%Y-%m', created_at) as month, COUNT(*) as count 
                    FROM Student 
                    WHERE created_at >= date('now', '-6 months')
                    GROUP BY month
                    ORDER BY month
                ''').fetchall()
            except Exception as e:
                print(f"[STATISTICS] Error getting new students timeline: {e}")
                new_students_timeline = []
            
        except Exception as e:
            print(f"[STATISTICS] Error in statistics queries: {e}")
        finally:
            conn.close()
        
        # Convert to dictionaries for JSON serialization (use safe defaults if missing)
        stats_data = {
            'overall': {
                'total_students': total_students,
                'total_appointments': total_appointments,
                'total_sessions': total_sessions,
                'total_referrals': total_referrals
            },
            'gender': [{'gender': row['gender'], 'count': row['count']} for row in gender_stats] if gender_stats else [],
            'programme': [{'programme': row['programme'], 'count': row['count']} for row in programme_stats] if programme_stats else [],
            'department': [{'department': row['department'], 'count': row['count']} for row in department_stats] if department_stats else [],
            'appointment_status': [{'status': row['status'], 'count': row['count']} for row in appointment_status_stats] if appointment_status_stats else [],
            'appointment_timeline': [{'month': row['month'], 'count': row['count']} for row in appointment_timeline] if appointment_timeline else [],
            'session_timeline': [{'month': row['month'], 'count': row['count']} for row in session_timeline] if session_timeline else [],
            'session_type': [{'session_type': row['session_type'], 'count': row['count']} for row in session_type_stats] if session_type_stats else [],
            'age_distribution': [{'age_group': row['age_group'], 'count': row['count']} for row in age_distribution] if age_distribution else [],
            'top_counsellors': [{'name': row['name'], 'count': row['appointment_count']} for row in top_counsellors] if top_counsellors else [],
            'new_students_timeline': [{'month': row['month'], 'count': row['count']} for row in new_students_timeline] if new_students_timeline else []
        }
        
        return render_template('statistics.html', stats_data=stats_data)
    
    except Exception as e:
        print(f"[STATISTICS] Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error loading statistics: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

# Add error handler for 500 errors - display them properly
@app.errorhandler(500)
def internal_error(error):
    import traceback
    error_trace = traceback.format_exc()
    print(f"[ERROR 500] {error}")
    print(f"[ERROR 500] Traceback:\n{error_trace}")
    
    # Try to log to file if running as EXE
    try:
        if getattr(sys, 'frozen', False):
            error_log_path = os.path.join(os.path.dirname(sys.executable), 'error_log.txt')
            with open(error_log_path, 'a') as f:
                f.write(f"\n=== ERROR {datetime.now()} ===\n")
                f.write(f"{error}\n")
                f.write(f"{error_trace}\n")
                f.write("=" * 50 + "\n")
            print(f"[ERROR] Details logged to: {error_log_path}")
    except:
        pass
    
    return '''
    <html>
    <head><title>Internal Server Error</title></head>
    <body>
        <h1>Internal Server Error</h1>
        <p>The server encountered an error. Details have been logged.</p>
        <p><a href="/welcome">Return to Login</a></p>
    </body>
    </html>
    ''', 500

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    error_trace = traceback.format_exc()
    print(f"[UNHANDLED ERROR] {e}")
    print(f"[UNHANDLED ERROR] Traceback:\n{error_trace}")
    return internal_error(e)

if __name__ == '__main__':
    # Initialize database FIRST before anything else
    print("Initializing database...")
    try:
        # Force database initialization at startup
        try:
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
        except:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        db_path = os.path.join(base_path, 'counseling.db')
        
        # Check and initialize
        if not os.path.exists(db_path):
            print(f"Database not found, creating at: {db_path}")
            import db_setup
            db_setup.init_db()
        else:
            # Verify Appointment table exists
            test_conn = sqlite3.connect(db_path)
            cursor = test_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Appointment'")
            if not cursor.fetchone():
                print("Appointment table missing, reinitializing database...")
                test_conn.close()
                import db_setup
                db_setup.init_db()
            else:
                test_conn.close()
                print("Database check passed - Appointment table exists")
    except Exception as e:
        print(f"WARNING: Database initialization issue: {e}")
        print("Attempting to continue anyway...")
    
    # Check if port 5000 is already in use and kill the process if needed
    import socket
    import subprocess
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('127.0.0.1', 5000))
        sock.close()
    except OSError as e:
        if e.errno == 10048 or (hasattr(e, 'winerror') and e.winerror == 10048):  # Port already in use
            print("WARNING: Port 5000 is already in use!")
            print("Attempting to kill the process using port 5000...")
            try:
                # Find process using port 5000
                result = subprocess.run(['netstat', '-ano'], capture_output=True, text=True, timeout=5)
                for line in result.stdout.split('\n'):
                    if ':5000' in line and 'LISTENING' in line:
                        parts = line.split()
                        if len(parts) > 4:
                            pid = parts[-1]
                            print(f"Found process {pid} using port 5000. Killing it...")
                            subprocess.run(['taskkill', '/F', '/PID', pid], 
                                         capture_output=True, timeout=5)
                            import time
                            time.sleep(1)  # Wait a bit for port to be released
                # Try binding again
                try:
                    sock2 = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock2.bind(('127.0.0.1', 5000))
                    sock2.close()
                    print("Port 5000 is now available!")
                except:
                    print("ERROR: Could not free port 5000.")
                    print("Please manually close other instances or restart your computer.")
                    is_exe = getattr(sys, 'frozen', False)
                    if not is_exe:
                        input("Press Enter to exit...")
                    sys.exit(1)
            except Exception as kill_error:
                print(f"Could not kill process: {kill_error}")
                print("Please manually close other instances.")
                is_exe = getattr(sys, 'frozen', False)
                if not is_exe:
                    input("Press Enter to exit...")
                sys.exit(1)
    
    # Log available routes for debugging
    print('=' * 60)
    print('AAMUSTED Counselling Management System')
    print('=' * 60)
    print('Starting server on http://127.0.0.1:5000')
    print('Registered routes:')
    for rule in app.url_map.iter_rules():
        if rule.endpoint not in ['static']:
            print(f"  - {rule.endpoint}: {rule}")
    print('=' * 60)
    print()
    
    # For EXE, don't use debug mode and open browser automatically
    import webbrowser
    import threading
    
    def open_browser():
        import time
        time.sleep(2.5)  # Wait longer for server to fully start
        try:
            # Try to open browser
            webbrowser.open('http://127.0.0.1:5000')
            print("Browser opened automatically!")
        except Exception as e:
            print(f"Could not open browser automatically: {e}")
            print("Please manually open: http://127.0.0.1:5000")
    
    # Open browser automatically (only if not in debug mode)
    is_exe = getattr(sys, 'frozen', False)
    if is_exe:
        browser_thread = threading.Thread(target=open_browser)
        browser_thread.daemon = True
        browser_thread.start()
        print("Server starting... Browser will open automatically.")
    else:
        print("Starting in development mode...")
        print("Open your browser and navigate to: http://127.0.0.1:5000")
    
    print()
    
    try:
        # Run app (debug=False for production EXE)
        app.run(debug=not is_exe, host='127.0.0.1', port=5000, use_reloader=False)
    except Exception as e:
        print(f"ERROR: Failed to start server: {e}")
        if not is_exe:
            input("Press Enter to exit...")
        sys.exit(1)